import sys
import os
import time
import re
from datetime import datetime, timedelta, date

import openpyxl
from openpyxl.styles import Alignment

from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                               QLabel, QLineEdit, QComboBox, QPushButton, QTableWidget, 
                               QTableWidgetItem, QHeaderView, QMessageBox, QGroupBox, 
                               QGridLayout, QDateEdit, QRadioButton, QButtonGroup, 
                               QListWidget)
from PySide6.QtCore import Qt, QDate


class JournalLogic:
    def __init__(self, filename):
        self.filename = filename
        self.wb = None
        self.START_ROW = 7
        self.HOURS_COLS = {'lecture': 12, 'practice': 13, 'lab': 14}
        self.selected_dates = []
        self.LOAD_TYPES = ["осн.", "почас.", "совм."]
    
    def date_to_datetime(self, date_obj):
        return datetime.combine(date_obj, datetime.min.time()) if isinstance(date_obj, date) and not isinstance(date_obj, datetime) else date_obj
    
    def determine_week_type(self, input_date):
        current_date = self.date_to_datetime(input_date)
        year = current_date.year
        
        september_1 = datetime(year, 9, 1)
        if current_date < september_1:
            september_1 = datetime(year-1, 9, 1)
        
        days_diff = (current_date - september_1).days
        return "числитель" if (days_diff // 7) % 2 == 0 else "знаменатель"
    
    def find_sheet_for_month(self, month):
        month_names = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",
                      7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        
        target_month = month_names.get(month)
        if not target_month:
            return None
        
        for sheet_name in self.wb.sheetnames:
            if re.search(r'\b' + re.escape(target_month) + r'\b', sheet_name):
                return sheet_name
        
        return None
    
    def generate_dates_by_period(self, start_date, end_date, week_type):
        """Генерирует даты по периоду и типу недели"""
        try:
            start_dt = start_date.toPython()
            end_dt = end_date.toPython()
            
            if start_dt >= end_dt:
                return False, "Дата начала должна быть раньше даты окончания"
            
            self.selected_dates.clear()
            generated_count = 0
            
            if week_type == "обе недели":
                current_date = start_dt
                while current_date <= end_dt:
                    sheet_name = self.find_sheet_for_month(current_date.month)
                    if sheet_name:
                        self.selected_dates.append({
                            'date': current_date, 'day': current_date.day, 'month': current_date.month,
                            'year': current_date.year, 'sheet': sheet_name,
                            'week_type': self.determine_week_type(current_date)
                        })
                        generated_count += 1
                    current_date += timedelta(days=7)
            else:
                current_date = start_dt
                while current_date <= end_dt:
                    current_week_type = self.determine_week_type(current_date)
                    if current_week_type == week_type:
                        sheet_name = self.find_sheet_for_month(current_date.month)
                        if sheet_name:
                            self.selected_dates.append({
                                'date': current_date, 'day': current_date.day, 'month': current_date.month,
                                'year': current_date.year, 'sheet': sheet_name, 'week_type': current_week_type
                            })
                            generated_count += 1
                        current_date += timedelta(days=7)
                    else:
                        current_date += timedelta(days=1)
            
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            
            if generated_count > 0:
                dates_list = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
                week_type_display = "числитель и знаменатель (каждую неделю)" if week_type == "обе недели" else week_type
                
                return True, (f"Сгенерировано {generated_count} дат\n"
                             f"Период: {start_dt.strftime('%d.%m.%Y')} - {end_dt.strftime('%d.%m.%Y')}\n"
                             f"Тип недели: {week_type_display}\n"
                             f"Даты: {dates_list}")
            else:
                return False, "В выбранном периоде нет дат"
                
        except Exception as e:
            return False, f"Ошибка генерации дат: {e}"
    
    def add_single_date(self, date_obj):
        """Добавляет одиночную дату"""
        try:
            date_py = date_obj.toPython()
            sheet_name = self.find_sheet_for_month(date_py.month)
            
            if not sheet_name:
                return False, f"Не найден лист для месяца {date_py.month}"
            
            for existing_date in self.selected_dates:
                if (existing_date['day'] == date_py.day and 
                    existing_date['month'] == date_py.month and 
                    existing_date['year'] == date_py.year):
                    return False, "Эта дата уже есть в списке"
            
            date_info = {
                'date': date_py, 
                'day': date_py.day, 
                'month': date_py.month,
                'year': date_py.year, 
                'sheet': sheet_name,
                'week_type': self.determine_week_type(date_py)
            }
            
            self.selected_dates.append(date_info)
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            
            return True, f"Дата {date_py.strftime('%d.%m.%Y')} добавлена"
            
        except Exception as e:
            return False, f"Ошибка добавления даты: {e}"
    
    def remove_selected_date(self, index):
        """Удаляет выбранную дату"""
        if index == -1:
            return False, "Выберите дату для удаления"
        
        if 0 <= index < len(self.selected_dates):
            removed_date = self.selected_dates.pop(index)
            return True, f"Дата {removed_date['day']}.{removed_date['month']:02d}.{removed_date['year']} удалена"
        else:
            return False, "Неверный индекс даты"
    
    def clear_dates(self):
        """Очищает все даты"""
        self.selected_dates.clear()
        return True, "Все даты очищены"
    
    def get_dates_info(self):
        """Возвращает информацию о выбранных датах"""
        count = len(self.selected_dates)
        if count > 0:
            dates_str = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
            sheets_count = {}
            for date in self.selected_dates:
                sheets_count[date['sheet']] = sheets_count.get(date['sheet'], 0) + 1
            
            sheets_info = ", ".join([f"{sheet}: {count}" for sheet, count in sheets_count.items()])
            week_types = set(date['week_type'] for date in self.selected_dates)
            week_types_info = f"Типы: {', '.join(week_types)}" if week_types else ""
            
            return f"Выбрано дат: {count} | Даты: {dates_str} | Листы: {sheets_info} {week_types_info}"
        else:
            return "Выбрано дат: 0"
    
    def get_dates_for_display(self):
        """Возвращает даты для отображения"""
        date_values = []
        display_texts = []
        
        for date_info in self.selected_dates:
            display_text = f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']} ({date_info['sheet']}, {date_info['week_type']})"
            display_texts.append(display_text)
            date_values.append(f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']}")
        
        return display_texts, date_values
    
    def safe_load_workbook(self):
        """Безопасная загрузка рабочей книги с повторными попытками"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if os.path.exists(self.filename):
                    self.wb = openpyxl.load_workbook(self.filename)
                    return True, "Файл успешно загружен"
                else:
                    return False, f"Файл {self.filename} не найден"
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
                else:
                    return False, (f"Нет доступа к файлу {self.filename}!\n"
                                 f"Убедитесь, что файл не открыт в другой программе.")
            except Exception as e:
                return False, f"Ошибка загрузки файла: {e}"
        return False, "Не удалось загрузить файл после нескольких попыток"
    
    def safe_save_workbook(self):
        """Безопасное сохранение рабочей книги с повторными попытками"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if self.wb:
                    self.wb.save(self.filename)
                    return True, "Файл успешно сохранен"
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
                else:
                    return False, (f"Нет доступа к файлу {self.filename}!\n"
                                 f"Убедитесь, что файл не открыт в другой программе.")
            except Exception as e:
                return False, f"Ошибка сохранения файла: {e}"
        return False, "Не удалось сохранить файл после нескольких попыток"
    
    def close_workbook(self):
        """Безопасное закрытие рабочей книги"""
        try:
            if hasattr(self, 'wb') and self.wb:
                self.wb.close()
                self.wb = None
                return True
        except Exception as e:
            print(f"Ошибка при закрытии файла: {e}")
            return False
    
    def load_workbook(self):
        """Загружает рабочую книгу Excel"""
        try:
            success, message = self.safe_load_workbook()
            if success:
                return True, "Файл успешно загружен"
            else:
                return False, message
        except Exception as e:
            return False, f"Ошибка загрузки файла: {e}"
    
    def get_sheets(self):
        """Возвращает список листов"""
        if self.wb:
            return self.wb.sheetnames
        return []
    
    def get_sheet_data(self, sheet_name):
        """Возвращает данные из листа"""
        if not self.wb or sheet_name not in self.wb.sheetnames:
            return []
        
        try:
            sheet = self.wb[sheet_name]
            data_rows = []
            
            # Для семестровых листов (осень/весна) используем другой формат отображения
            if sheet_name.lower() in ['осень', 'весна']:
                row = 5
                max_rows_to_check = 50
                
                while row <= max_rows_to_check:
                    try:
                        discipline = sheet[f'D{row}'].value
                        group = sheet[f'E{row}'].value
                        load_type = sheet[f'F{row}'].value
                        
                        # Проверяем, есть ли данные в строке
                        if discipline or group or load_type:
                            # Для семестровых листов получаем часы из колонок G, H, I
                            lecture = sheet[f'G{row}'].value or ''
                            practice = sheet[f'H{row}'].value or ''
                            lab = sheet[f'I{row}'].value or ''
                            
                            data_rows.append([
                                sheet_name,  # Вместо числа - название семестра
                                discipline or '',
                                group or '',
                                load_type or '',
                                lecture,
                                practice,
                                lab
                            ])
                    except:
                        pass
                    
                    row += 1
            else:
                # Для месячных листов используем стандартный формат
                row = self.START_ROW
                
                while sheet[f'E{row}'].value is not None:
                    day = sheet[f'E{row}'].value
                    if isinstance(day, (int, float)):
                        data_rows.append([
                            int(day),
                            sheet[f'F{row}'].value or '',
                            sheet[f'G{row}'].value or '',
                            sheet[f'H{row}'].value or '',
                            sheet.cell(row=row, column=12).value or '',
                            sheet.cell(row=row, column=13).value or '',
                            sheet.cell(row=row, column=14).value or ''
                        ])
                    row += 1
            
            return data_rows
                
        except Exception as e:
            print(f"Ошибка при чтении данных: {e}")
            return []
    
    def delete_entries(self, sheet_name, entries_to_delete):
        """Удаляет записи из листа"""
        try:
            if not self.wb or sheet_name not in self.wb.sheetnames:
                return False, "Лист не найден"
            
            sheet = self.wb[sheet_name]
            deleted_count = 0
            
            # Для семестровых и месячных листов разная логика удаления
            if sheet_name.lower() in ['осень', 'весна']:
                deleted_count = self._delete_season_entries(sheet, entries_to_delete)
            else:
                deleted_count = self._delete_monthly_entries(sheet, entries_to_delete)
            
            return True, f"Удалено записей: {deleted_count} из {len(entries_to_delete)}"
            
        except Exception as e:
            return False, f"Ошибка при удалении записей: {e}"
    
    def _delete_monthly_entries(self, sheet, entries_to_delete):
        """Удаляет записи из месячного листа"""
        rows_to_delete = []
        for entry_data in entries_to_delete:
            day = int(entry_data[0])
            discipline = entry_data[1]
            group = entry_data[2]
            
            row_num = self.START_ROW
            while sheet[f'E{row_num}'].value is not None:
                sheet_day = sheet[f'E{row_num}'].value
                sheet_discipline = sheet[f'F{row_num}'].value or ''
                sheet_group = sheet[f'G{row_num}'].value or ''
                
                if (isinstance(sheet_day, (int, float)) and int(sheet_day) == day and
                    sheet_discipline == discipline and sheet_group == group):
                    rows_to_delete.append(row_num)
                    break
                row_num += 1
        
        rows_to_delete.sort(reverse=True)
        for row_num in rows_to_delete:
            self._delete_monthly_row(sheet, row_num)
        
        return len(rows_to_delete)
    
    def _delete_season_entries(self, sheet, entries_to_delete):
        """Удаляет записи из семестрового листа"""
        rows_to_delete = []
        for entry_data in entries_to_delete:
            discipline = entry_data[1]
            group = entry_data[2]
            load_type = entry_data[3]
            
            row_num = 5
            max_rows_to_check = 50
            
            while row_num <= max_rows_to_check:
                try:
                    sheet_discipline = sheet[f'D{row_num}'].value or ''
                    sheet_group = sheet[f'E{row_num}'].value or ''
                    sheet_load_type = sheet[f'F{row_num}'].value or ''
                    
                    if (sheet_discipline == discipline and 
                        sheet_group == group and 
                        sheet_load_type == load_type):
                        rows_to_delete.append(row_num)
                        break
                except:
                    pass
                row_num += 1
        
        rows_to_delete.sort(reverse=True)
        for row_num in rows_to_delete:
            self._delete_season_row(sheet, row_num)
        
        return len(rows_to_delete)
    
    def _delete_monthly_row(self, sheet, row_num):
        """Удаляет строку из месячного листа"""
        max_row = self.START_ROW
        while sheet[f'E{max_row}'].value is not None:
            max_row += 1
        
        for row in range(row_num, max_row):
            sheet[f'E{row}'] = sheet[f'E{row + 1}'].value
            sheet[f'F{row}'] = sheet[f'F{row + 1}'].value
            sheet[f'G{row}'] = sheet[f'G{row + 1}'].value
            sheet[f'H{row}'] = sheet[f'H{row + 1}'].value
            sheet.cell(row=row, column=12).value = sheet.cell(row=row + 1, column=12).value
            sheet.cell(row=row, column=13).value = sheet.cell(row=row + 1, column=13).value
            sheet.cell(row=row, column=14).value = sheet.cell(row=row + 1, column=14).value
        
        # Очищаем последнюю строку
        sheet[f'E{max_row}'] = None
        sheet[f'F{max_row}'] = None
        sheet[f'G{max_row}'] = None
        sheet[f'H{max_row}'] = None
        sheet.cell(row=max_row, column=12).value = None
        sheet.cell(row=max_row, column=13).value = None
        sheet.cell(row=max_row, column=14).value = None
    
    def _delete_season_row(self, sheet, row_num):
        """Удаляет строку из семестрового листа"""
        max_row = 50  # Максимальная строка для проверки в семестровых листах
        
        for row in range(row_num, max_row):
            try:
                sheet[f'D{row}'] = sheet[f'D{row + 1}'].value
                sheet[f'E{row}'] = sheet[f'E{row + 1}'].value
                sheet[f'F{row}'] = sheet[f'F{row + 1}'].value
                sheet[f'G{row}'] = sheet[f'G{row + 1}'].value
                sheet[f'H{row}'] = sheet[f'H{row + 1}'].value
                sheet[f'I{row}'] = sheet[f'I{row + 1}'].value
            except:
                break
        
        # Очищаем последнюю строку
        try:
            sheet[f'D{max_row}'] = None
            sheet[f'E{max_row}'] = None
            sheet[f'F{max_row}'] = None
            sheet[f'G{max_row}'] = None
            sheet[f'H{max_row}'] = None
            sheet[f'I{max_row}'] = None
        except:
            pass
    
    def add_entries(self, data):
        """Добавляет записи в журнал"""
        try:
            if not self.wb or not self.selected_dates:
                return False, "Нет данных для добавления"
            
            # Проверяем обязательные поля
            if not all([data.get('discipline'), data.get('group'), data.get('load_type')]):
                return False, "Заполните все обязательные поля"
            
            # Проверяем, что есть хотя бы одни часы
            if data.get('lecture', 0) == 0 and data.get('practice', 0) == 0 and data.get('lab', 0) == 0:
                return False, "Заполните хотя бы одно поле: Лекции, Практические или Лабораторные"
            
            dates_by_sheet = {}
            for date_info in self.selected_dates:
                sheet_name = date_info['sheet']
                if sheet_name not in dates_by_sheet:
                    dates_by_sheet[sheet_name] = []
                dates_by_sheet[sheet_name].append(date_info)
            
            results = {}
            for sheet_name, dates in dates_by_sheet.items():
                if sheet_name not in self.wb.sheetnames:
                    continue
                    
                sheet = self.wb[sheet_name]
                dates.sort(key=lambda x: x['day'])
                
                added_rows = []
                for date_info in dates:
                    row = self._add_entry_to_sheet(sheet, date_info['day'], data)
                    if row:
                        added_rows.append(f"{date_info['day']}.{date_info['month']:02d}(стр.{row})")
                
                if added_rows:
                    results[sheet_name] = added_rows
            
            # Заполняем листы "Осень" и "Весна" с добавлением часов
            season_results = self._fill_season_sheets(data)
            
            # Сохраняем файл
            success, save_message = self.safe_save_workbook()
            if not success:
                return False, save_message
            
            # Формируем сообщение об успехе
            msg_lines = ["Записи добавлены:"]
            for sheet_name, dates in results.items():
                msg_lines.append(f"{sheet_name}: {', '.join(dates)}")
            
            if season_results:
                msg_lines.append("\nСеместровые листы:")
                for sheet_name, result in season_results.items():
                    msg_lines.append(f"{sheet_name}: {result}")
            
            if results or season_results:
                return True, "\n".join(msg_lines)
            else:
                return False, "Не удалось добавить записи"
                
        except ValueError as e:
            return False, "Проверьте числовые поля (Лекции, Практические, Лабораторные) - они должны содержать только числа"
        except Exception as e:
            return False, f"Ошибка при добавлении записей: {e}"
    
    def _add_entry_to_sheet(self, sheet, day, data):
        """Добавляет запись в лист и возвращает номер строки"""
        try:
            last_row = self.START_ROW
            while sheet[f'E{last_row}'].value is not None:
                last_row += 1
            
            existing_date_rows = []
            row = self.START_ROW
            while sheet[f'E{row}'].value is not None:
                existing_day = sheet[f'E{row}'].value
                if isinstance(existing_day, (int, float)) and int(existing_day) == day:
                    existing_date_rows.append(row)
                row += 1
            
            if existing_date_rows:
                last_date_row = max(existing_date_rows)
                insert_row = last_date_row + 1
                self._shift_rows_down(sheet, insert_row, last_row)
                self._fill_row_data(sheet, insert_row, day, data)
                return insert_row
            else:
                return self._insert_entry_sorted(sheet, day, data, last_row)
        except Exception as e:
            print(f"Ошибка при добавлении записи: {e}")
            return None
    
    def _shift_rows_down(self, sheet, start_row, last_row):
        """Сдвигает строки вниз"""
        for row in range(last_row, start_row - 1, -1):
            sheet[f'E{row+1}'] = sheet[f'E{row}'].value
            sheet[f'F{row+1}'] = sheet[f'F{row}'].value
            sheet[f'G{row+1}'] = sheet[f'G{row}'].value
            sheet[f'H{row+1}'] = sheet[f'H{row}'].value
            sheet.cell(row=row+1, column=12).value = sheet.cell(row=row, column=12).value
            sheet.cell(row=row+1, column=13).value = sheet.cell(row=row, column=13).value
            sheet.cell(row=row+1, column=14).value = sheet.cell(row=row, column=14).value
    
    def _insert_entry_sorted(self, sheet, day, data, last_row):
        """Вставляет запись в отсортированном порядке"""
        insert_row = self.START_ROW
        while sheet[f'E{insert_row}'].value is not None:
            existing_day = sheet[f'E{insert_row}'].value
            if isinstance(existing_day, (int, float)) and int(existing_day) > day:
                break
            insert_row += 1
        
        self._shift_rows_down(sheet, insert_row, last_row)
        self._fill_row_data(sheet, insert_row, day, data)
        return insert_row
    
    def _fill_row_data(self, sheet, row, day, data):
        """Заполняет строку данными"""
        sheet[f'E{row}'] = day
        sheet[f'F{row}'] = data['discipline']
        sheet[f'G{row}'] = data['group']
        sheet[f'H{row}'] = data['load_type']
        
        if data.get('lecture', 0) != 0:
            sheet.cell(row=row, column=12).value = data['lecture']
        if data.get('practice', 0) != 0:
            sheet.cell(row=row, column=13).value = data['practice']
        if data.get('lab', 0) != 0:
            sheet.cell(row=row, column=14).value = data['lab']
    
    def _fill_season_sheets(self, data):
        """Заполняет листы 'Осень' и 'Весна' данными и возвращает результаты"""
        season_results = {}
        
        try:
            if not self.selected_dates:
                return season_results
            
            months = set(date_info['month'] for date_info in self.selected_dates)
            
            autumn_months = {9, 10, 11, 12}
            spring_months = {1, 2, 3, 4, 5}
            
            fill_autumn = any(month in autumn_months for month in months)
            fill_spring = any(month in spring_months for month in months)
            
            if fill_autumn and 'осень' in self.wb.sheetnames:
                result = self._fill_season_sheet('осень', data)
                if result:
                    season_results['осень'] = result
            
            if fill_spring and 'весна' in self.wb.sheetnames:
                result = self._fill_season_sheet('весна', data)
                if result:
                    season_results['весна'] = result
                    
        except Exception as e:
            print(f"Ошибка при заполнении семестровых листов: {e}")
        
        return season_results
    
    def _fill_season_sheet(self, sheet_name, data):
        """Заполняет конкретный семестровый лист и возвращает результат"""
        try:
            sheet = self.wb[sheet_name]
            
            # Ищем существующую запись с такими же дисциплиной, группой и типом нагрузки
            existing_row = self._find_existing_season_entry(sheet, data)
            
            if existing_row:
                # Обновляем существующую запись - добавляем часы
                return self._update_season_entry(sheet, existing_row, data)
            else:
                # Создаем новую запись
                return self._create_new_season_entry(sheet, data)
            
        except Exception as e:
            return f"Ошибка: {str(e)}"
    
    def _find_existing_season_entry(self, sheet, data):
        """Ищет существующую запись в семестровом листе"""
        row = 5
        max_rows_to_check = 50
        
        while row <= max_rows_to_check:
            try:
                sheet_discipline = sheet[f'D{row}'].value or ''
                sheet_group = sheet[f'E{row}'].value or ''
                sheet_load_type = sheet[f'F{row}'].value or ''
                
                if (sheet_discipline == data['discipline'] and 
                    sheet_group == data['group'] and 
                    sheet_load_type == data['load_type']):
                    return row
            except:
                pass
            
            row += 1
        
        return None
    
    def _update_season_entry(self, sheet, row, data):
        """Обновляет существующую запись в семестровом листе, добавляя часы"""
        try:
            # Получаем текущие значения часов
            current_lecture = sheet[f'G{row}'].value or 0
            current_practice = sheet[f'H{row}'].value or 0
            current_lab = sheet[f'I{row}'].value or 0
            
            # Преобразуем в числа
            if isinstance(current_lecture, str):
                current_lecture = float(current_lecture) if current_lecture else 0
            if isinstance(current_practice, str):
                current_practice = float(current_practice) if current_practice else 0
            if isinstance(current_lab, str):
                current_lab = float(current_lab) if current_lab else 0
            
            # Суммируем часы
            new_lecture = current_lecture + data.get('lecture', 0)
            new_practice = current_practice + data.get('practice', 0)
            new_lab = current_lab + data.get('lab', 0)
            
            # Записываем обновленные значения
            sheet[f'G{row}'] = new_lecture if new_lecture != 0 else ''
            sheet[f'H{row}'] = new_practice if new_practice != 0 else ''
            sheet[f'I{row}'] = new_lab if new_lab != 0 else ''
            
            # Выравнивание
            for col in ['G', 'H', 'I']:
                sheet[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            return f"обновлена строка {row}: +{data.get('lecture', 0)}л, +{data.get('practice', 0)}п, +{data.get('lab', 0)}лаб"
            
        except Exception as e:
            return f"Ошибка обновления строки {row}: {str(e)}"
    
    def _create_new_season_entry(self, sheet, data):
        """Создает новую запись в семестровом листе"""
        row = 5
        max_rows_to_check = 50
        
        while row <= max_rows_to_check:
            try:
                cell_d = sheet[f'D{row}']
                cell_e = sheet[f'E{row}']
                cell_f = sheet[f'F{row}']
                
                def is_cell_empty(cell):
                    try:
                        return cell.value is None or str(cell.value).strip() == ''
                    except:
                        return True
                
                if (is_cell_empty(cell_d) and 
                    is_cell_empty(cell_e) and 
                    is_cell_empty(cell_f)):
                    break
            except:
                pass
            
            row += 1
        
        if row > max_rows_to_check:
            return f"Не найдено свободных строк (проверено до строки {max_rows_to_check})"
        
        try:
            # Записываем основные данные
            sheet[f'D{row}'] = data['discipline']
            sheet[f'E{row}'] = data['group']
            sheet[f'F{row}'] = data['load_type']
            
            # Записываем часы
            sheet[f'G{row}'] = data.get('lecture', 0) if data.get('lecture', 0) != 0 else ''
            sheet[f'H{row}'] = data.get('practice', 0) if data.get('practice', 0) != 0 else ''
            sheet[f'I{row}'] = data.get('lab', 0) if data.get('lab', 0) != 0 else ''
            
            # Выравнивание
            for col in ['D', 'E', 'F', 'G', 'H', 'I']:
                sheet[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            return f"строка {row}: {data['discipline']}, {data['group']}, {data['load_type']}, часы: {data.get('lecture', 0)}л/{data.get('practice', 0)}п/{data.get('lab', 0)}лаб"
            
        except Exception as write_error:
            return f"Ошибка записи в строку {row}: {str(write_error)}"


class JournalApp(QMainWindow):
    def __init__(self, filename):
        super().__init__()
        self.logic = JournalLogic(filename)
        self.setup_ui()
        self.load_workbook()
    
    def setup_ui(self):
        self.setWindowTitle("Журнал преподавателя")
        self.setGeometry(100, 100, 1200, 900)
        
        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Основной layout
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Период
        period_group = self.create_period_group()
        main_layout.addWidget(period_group)
        
        # Даты
        dates_group = self.create_dates_group()
        main_layout.addWidget(dates_group)
        
        # Поля ввода
        input_group = self.create_input_group()
        main_layout.addWidget(input_group)
        
        # Кнопка добавления
        add_button = QPushButton("Добавить записи")
        add_button.clicked.connect(self.add_entries)
        add_button.setMinimumHeight(35)
        add_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        main_layout.addWidget(add_button)
        
        # Просмотр данных
        view_group = self.create_view_group()
        main_layout.addWidget(view_group, 1)
    
    def create_period_group(self):
        group = QGroupBox("Выбор периода")
        layout = QGridLayout(group)
        
        # Начало периода
        layout.addWidget(QLabel("Начало периода:"), 0, 0)
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate(datetime.now().year, 9, 1))
        self.start_date.setDisplayFormat("dd.MM.yyyy")
        self.start_date.setCalendarPopup(True)
        layout.addWidget(self.start_date, 0, 1)
        
        # Конец периода
        layout.addWidget(QLabel("Конец периода:"), 0, 2)
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate(datetime.now().year, 12, 31))
        self.end_date.setDisplayFormat("dd.MM.yyyy")
        self.end_date.setCalendarPopup(True)
        layout.addWidget(self.end_date, 0, 3)
        
        # Тип недели
        layout.addWidget(QLabel("Тип недели:"), 0, 4)
        
        self.period_week_type = QButtonGroup(self)
        week_type_layout = QHBoxLayout()
        
        for text, value in [("Числитель", "числитель"), ("Знаменатель", "знаменатель"), ("Обе недели", "обе недели")]:
            radio = QRadioButton(text)
            radio.setProperty("value", value)
            self.period_week_type.addButton(radio)
            week_type_layout.addWidget(radio)
        
        self.period_week_type.buttons()[0].setChecked(True)
        layout.addLayout(week_type_layout, 0, 5, 1, 2)
        
        # Кнопки
        button_layout = QHBoxLayout()
        generate_btn = QPushButton("Сгенерировать даты по периоду")
        generate_btn.clicked.connect(self.generate_dates_by_period)
        button_layout.addWidget(generate_btn)
        
        clear_btn = QPushButton("Очистить даты")
        clear_btn.clicked.connect(self.clear_dates)
        button_layout.addWidget(clear_btn)
        
        layout.addLayout(button_layout, 1, 0, 1, 6)
        
        # Информация о датах
        self.dates_info_label = QLabel("Выбрано дат: 0")
        self.dates_info_label.setStyleSheet("color: blue; font-weight: bold;")
        layout.addWidget(self.dates_info_label, 2, 0, 1, 6)
        
        return group
    
    def create_dates_group(self):
        group = QGroupBox("Управление датами")
        layout = QHBoxLayout(group)
        
        # Левая часть - управление датами
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        
        # Добавление даты
        add_date_layout = QHBoxLayout()
        add_date_layout.addWidget(QLabel("Добавить дату:"))
        self.single_date = QDateEdit()
        self.single_date.setDate(QDate.currentDate())
        self.single_date.setDisplayFormat("dd.MM.yyyy")
        self.single_date.setCalendarPopup(True)
        add_date_layout.addWidget(self.single_date)
        
        add_btn = QPushButton("Добавить дату")
        add_btn.clicked.connect(self.add_single_date)
        add_date_layout.addWidget(add_btn)
        add_date_layout.addStretch()
        
        left_layout.addLayout(add_date_layout)
        
        # Удаление даты
        remove_date_layout = QHBoxLayout()
        remove_date_layout.addWidget(QLabel("Удалить дату:"))
        self.remove_date_combo = QComboBox()
        remove_date_layout.addWidget(self.remove_date_combo)
        
        remove_btn = QPushButton("Удалить дату")
        remove_btn.clicked.connect(self.remove_selected_date)
        remove_date_layout.addWidget(remove_btn)
        remove_date_layout.addStretch()
        
        left_layout.addLayout(remove_date_layout)
        
        # Список дат
        left_layout.addWidget(QLabel("Выбранные даты:"))
        self.dates_listbox = QListWidget()
        self.dates_listbox.setMaximumHeight(120)
        left_layout.addWidget(self.dates_listbox)
        
        # Кнопки управления
        manage_buttons_layout = QHBoxLayout()
        clear_all_btn = QPushButton("Очистить все даты")
        clear_all_btn.clicked.connect(self.clear_dates)
        manage_buttons_layout.addWidget(clear_all_btn)
        
        refresh_btn = QPushButton("Обновить список")
        refresh_btn.clicked.connect(self.update_dates_display)
        manage_buttons_layout.addWidget(refresh_btn)
        manage_buttons_layout.addStretch()
        
        left_layout.addLayout(manage_buttons_layout)
        
        layout.addWidget(left_widget, 1)
        
        return group
    
    def create_input_group(self):
        group = QGroupBox("Данные для добавления")
        layout = QGridLayout(group)
        
        self.entries = {}
        fields = [
            ("Дисциплина:", "discipline", "entry"), 
            ("Группа:", "group", "entry"),
            ("Вид нагрузки:", "load_type", "combobox"),
            ("Лекции:", "lecture", "entry"),
            ("Практические:", "practice", "entry"),
            ("Лабораторные:", "lab", "entry")
        ]
        
        for i, (label, field, field_type) in enumerate(fields):
            layout.addWidget(QLabel(label), i, 0)
            
            if field_type == "combobox":
                widget = QComboBox()
                widget.addItems(self.logic.LOAD_TYPES)
                self.entries[field] = widget
            else:
                widget = QLineEdit()
                self.entries[field] = widget
            
            layout.addWidget(widget, i, 1)
        
        return group
    
    def create_view_group(self):
        group = QGroupBox("Просмотр данных")
        layout = QVBoxLayout(group)
        
        # Панель управления
        control_layout = QHBoxLayout()
        
        control_layout.addWidget(QLabel("Лист:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.currentTextChanged.connect(self.show_data)
        control_layout.addWidget(self.sheet_combo)
        
        # Кнопки управления
        self.delete_btn = QPushButton("Удалить выбранные записи")
        self.delete_btn.clicked.connect(self.delete_selected_entries)
        self.delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 3px;
                padding: 5px 10px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        control_layout.addWidget(self.delete_btn)
        
        self.select_all_btn = QPushButton("Выбрать все")
        self.select_all_btn.clicked.connect(self.select_all_entries)
        control_layout.addWidget(self.select_all_btn)
        
        self.deselect_all_btn = QPushButton("Снять выделение")
        self.deselect_all_btn.clicked.connect(self.deselect_all_entries)
        control_layout.addWidget(self.deselect_all_btn)
        
        control_layout.addStretch()
        
        layout.addLayout(control_layout)
        
        # Информация о выборе
        self.selection_info = QLabel("Выбрано записей: 0")
        self.selection_info.setStyleSheet("color: green; font-weight: bold;")
        layout.addWidget(self.selection_info)
        
        # Таблица
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "Число", "Дисциплина", "Группа", "Нагрузка", "Лекции", "Практические", "Лабораторные"
        ])
        
        # Настройка таблицы
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.MultiSelection)
        self.table.itemSelectionChanged.connect(self.on_table_selection_changed)
        
        layout.addWidget(self.table, 1)
        
        return group
    
    def on_table_selection_changed(self):
        """Обновляет информацию о количестве выбранных записей"""
        selected_count = len(self.table.selectedItems()) // self.table.columnCount()
        self.selection_info.setText(f"Выбрано записей: {selected_count}")
    
    def select_all_entries(self):
        """Выбирает все записи в таблице"""
        self.table.selectAll()
    
    def deselect_all_entries(self):
        """Снимает выделение со всех записей"""
        self.table.clearSelection()
    
    def generate_dates_by_period(self):
        """Генерирует даты по периоду"""
        # Получаем выбранный тип недели
        target_week_type = None
        for btn in self.period_week_type.buttons():
            if btn.isChecked():
                target_week_type = btn.property("value")
                break
        
        if not target_week_type:
            QMessageBox.critical(self, "Ошибка", "Не выбран тип недели")
            return
        
        success, message = self.logic.generate_dates_by_period(
            self.start_date.date(), 
            self.end_date.date(), 
            target_week_type
        )
        
        if success:
            QMessageBox.information(self, "Успех", message)
        else:
            QMessageBox.critical(self, "Ошибка", message)
        
        self.update_dates_display()
    
    def add_single_date(self):
        """Добавляет одиночную дату"""
        success, message = self.logic.add_single_date(self.single_date.date())
        
        if success:
            QMessageBox.information(self, "Успех", message)
        else:
            QMessageBox.critical(self, "Ошибка", message)
        
        self.update_dates_display()
    
    def remove_selected_date(self):
        """Удаляет выбранную дату"""
        selected_index = self.remove_date_combo.currentIndex()
        success, message = self.logic.remove_selected_date(selected_index)
        
        if success:
            QMessageBox.information(self, "Успех", message)
        else:
            QMessageBox.critical(self, "Ошибка", message)
        
        self.update_dates_display()
    
    def clear_dates(self):
        """Очищает все даты"""
        success, message = self.logic.clear_dates()
        
        if success:
            QMessageBox.information(self, "Успех", message)
        else:
            QMessageBox.critical(self, "Ошибка", message)
        
        self.update_dates_display()
    
    def update_dates_display(self):
        """Обновляет отображение дат"""
        # Обновляем информацию о датах
        self.dates_info_label.setText(self.logic.get_dates_info())
        
        # Обновляем список дат
        display_texts, date_values = self.logic.get_dates_for_display()
        
        self.dates_listbox.clear()
        self.dates_listbox.addItems(display_texts)
        
        self.remove_date_combo.clear()
        self.remove_date_combo.addItems(date_values)
        if date_values:
            self.remove_date_combo.setCurrentIndex(0)
    
    def load_workbook(self):
        """Загружает рабочую книгу"""
        success, message = self.logic.load_workbook()
        
        if success:
            # Обновляем список листов
            sheets = self.logic.get_sheets()
            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheets)
            if sheets:
                self.sheet_combo.setCurrentIndex(0)
            
            self.show_data()
        else:
            QMessageBox.critical(self, "Ошибка", message)
    
    def show_data(self):
        """Показывает данные из выбранного листа"""
        sheet_name = self.sheet_combo.currentText()
        if not sheet_name:
            return
        
        data_rows = self.logic.get_sheet_data(sheet_name)
        
        self.table.setRowCount(len(data_rows))
        for i, row_data in enumerate(data_rows):
            for j, value in enumerate(row_data):
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(i, j, item)
        
        self.selection_info.setText("Выбрано записей: 0")
    
    def delete_selected_entries(self):
        """Удаляет выбранные записи"""
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            QMessageBox.warning(self, "Внимание", "Выберите записи для удаления")
            return
        
        # Получаем индексы выбранных строк
        selected_rows = set()
        for range in selected_ranges:
            for row in range.topRow(), range.bottomRow() + 1:
                selected_rows.add(row)
        
        selected_rows = sorted(selected_rows)
        if not selected_rows:
            return
        
        entries_to_delete = []
        for row in selected_rows:
            if row < self.table.rowCount():
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                entries_to_delete.append(row_data)
        
        confirm = QMessageBox.question(
            self, 
            "Подтверждение удаления", 
            f"Вы действительно хотите удалить {len(entries_to_delete)} записей?\n"
            f"Это действие нельзя отменить.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if confirm != QMessageBox.Yes:
            return
        
        sheet_name = self.sheet_combo.currentText()
        success, message = self.logic.delete_entries(sheet_name, entries_to_delete)
        
        if success:
            QMessageBox.information(self, "Успех", message)
            self.show_data()
        else:
            QMessageBox.critical(self, "Ошибка", message)
    
    def add_entries(self):
        """Добавляет записи в журнал"""
        # Собираем данные из полей ввода
        data = {
            'discipline': self.entries['discipline'].text(),
            'group': self.entries['group'].text(),
            'load_type': self.entries['load_type'].currentText(),
        }
        
        # Обрабатываем числовые поля
        try:
            lecture = self.entries['lecture'].text()
            practice = self.entries['practice'].text()
            lab = self.entries['lab'].text()
            
            data['lecture'] = float(lecture) if lecture else 0.0
            data['practice'] = float(practice) if practice else 0.0
            data['lab'] = float(lab) if lab else 0.0
        except ValueError:
            QMessageBox.critical(self, "Ошибка", "Проверьте числовые поля (Лекции, Практические, Лабораторные) - они должны содержать только числа")
            return
        
        success, message = self.logic.add_entries(data)
        
        if success:
            QMessageBox.information(self, "Успех", message)
            # Очищаем поля ввода
            for field in ['discipline', 'group', 'lecture', 'practice', 'lab']:
                if field in self.entries and isinstance(self.entries[field], QLineEdit):
                    self.entries[field].clear()
            
            if 'load_type' in self.entries:
                self.entries['load_type'].setCurrentIndex(0)
            
            self.show_data()
        else:
            QMessageBox.critical(self, "Ошибка", message)
    
    def closeEvent(self, event):
        """Обработчик закрытия окна"""
        self.logic.close_workbook()
        event.accept()


def main():
    app = QApplication(sys.argv)
    
    # Устанавливаем стиль приложения
    app.setStyle('Fusion')
    
    filename = "Тетрадь_ППС_2025_2026_каф_NN_Фамилия_ИО_оч_заоч.xltx"
    
    # Проверяем существование файла
    if not os.path.exists(filename):
        # Пробуем найти файл с другим расширением
        possible_extensions = ['.xltx', '.xlsx', '.xls']
        found = False
        for ext in possible_extensions:
            alt_filename = filename.replace('.xltx', ext)
            if os.path.exists(alt_filename):
                filename = alt_filename
                found = True
                break
        
        if not found:
            print(f"Файл не найден: {filename}")
            print(f"Текущая директория: {os.getcwd()}")
            print(f"Файлы в директории: {os.listdir('.')}")
            sys.exit(1)
    
    print(f"Загружаем файл: {filename}")
    
    window = JournalApp(filename)
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()