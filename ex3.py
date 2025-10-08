import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox

class JournalApp:
    def __init__(self, filename):
        self.filename = filename
        self.wb = None
        self.START_ROW = 7
        self.HOURS_COLS = {'lecture': 12, 'practice': 13, 'lab': 14}
        
        self.setup_gui()
        self.load_workbook()
    
    def setup_gui(self):
        self.root = tk.Tk()
        self.root.title("Журнал преподавателя")
        self.root.geometry("800x600")
        
        # Основной фрейм
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Выбор листа
        ttk.Label(main_frame, text="Выберите лист:").grid(row=0, column=0, sticky=tk.W)
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(main_frame, textvariable=self.sheet_var)
        self.sheet_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        
        # Поля ввода
        fields = [
            ("Число месяца:", "day"),
            ("Дисциплина:", "discipline"), 
            ("Группа:", "group"),
            ("Вид нагрузки:", "load_type"),
            ("Лекции:", "lecture"),
            ("Практические:", "practice"),
            ("Лабораторные:", "lab")
        ]
        
        self.entries = {}
        for i, (label, field) in enumerate(fields):
            ttk.Label(main_frame, text=label).grid(row=i+1, column=0, sticky=tk.W, pady=2)
            entry = ttk.Entry(main_frame, width=30)
            entry.grid(row=i+1, column=1, sticky=(tk.W, tk.E), padx=5, pady=2)
            self.entries[field] = entry
        
        # ОДНА КНОПКА для добавления и сохранения
        ttk.Button(main_frame, text="Добавить и сохранить", command=self.add_and_save).grid(row=8, column=0, columnspan=2, pady=10)
        
        # Таблица с данными
        self.tree = ttk.Treeview(main_frame, columns=("Число", "Дисциплина", "Группа", "Нагрузка", "Лекции", "Практические", "Лабораторные"), show="headings", height=15)
        
        columns = ["Число", "Дисциплина", "Группа", "Нагрузка", "Лекции", "Практические", "Лабораторные"]
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)
        
        self.tree.grid(row=9, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # Настройка растягивания
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(9, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
    
    def load_workbook(self):
        """Загрузка файла Excel"""
        try:
            self.wb = openpyxl.load_workbook(self.filename)
            sheets = self.wb.sheetnames
            self.sheet_combo['values'] = sheets
            if sheets:
                self.sheet_var.set(sheets[0])
            self.show_data()
        except FileNotFoundError:
            messagebox.showerror("Ошибка", f"Файл '{self.filename}' не найден")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка загрузки файла: {e}")
    
    def reload_workbook(self):
        """Перезагрузка файла после сохранения"""
        try:
            self.wb = openpyxl.load_workbook(self.filename)
            self.show_data()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка перезагрузки файла: {e}")
    
    def get_existing_days(self, sheet):
        days_data = []
        row = self.START_ROW
        
        while sheet[f'E{row}'].value is not None:
            day_value = sheet[f'E{row}'].value
            if isinstance(day_value, (int, float)):
                days_data.append({
                    'row': row,
                    'day': int(day_value),
                    'discipline': sheet[f'F{row}'].value,
                    'group': sheet[f'G{row}'].value,
                    'load_type': sheet[f'H{row}'].value,
                    **{key: sheet.cell(row=row, column=col).value 
                       for key, col in self.HOURS_COLS.items()}
                })
            row += 1
        
        days_data.sort(key=lambda x: x['day'])
        return days_data
    
    def find_insert_position(self, sheet, new_day):
        existing_days = self.get_existing_days(sheet)
        
        if not existing_days:
            return self.START_ROW
        
        for day_data in existing_days:
            if day_data['day'] == new_day:
                return day_data['row']
        
        for i, day_data in enumerate(existing_days):
            if new_day < day_data['day']:
                return self.START_ROW if i == 0 else existing_days[i-1]['row'] + 1
        
        return existing_days[-1]['row'] + 1
    
    def insert_sorted_entry(self, sheet, insert_row, data):
        existing_data = []
        row = insert_row
        
        while sheet[f'E{row}'].value is not None:
            cell_data = {
                'day': sheet[f'E{row}'].value,
                'discipline': sheet[f'F{row}'].value,
                'group': sheet[f'G{row}'].value,
                'load_type': sheet[f'H{row}'].value
            }
            for key, col in self.HOURS_COLS.items():
                cell_data[key] = sheet.cell(row=row, column=col).value
            
            existing_data.append(cell_data)
            
            for col in ['E', 'F', 'G', 'H']:
                sheet[f'{col}{row}'] = None
            for col in self.HOURS_COLS.values():
                sheet.cell(row=row, column=col).value = None
            
            row += 1
        
        sheet[f'E{insert_row}'] = data['day_number']
        sheet[f'F{insert_row}'] = data['discipline']
        sheet[f'G{insert_row}'] = data['group']
        sheet[f'H{insert_row}'] = data['load_type']
        
        for key, col in self.HOURS_COLS.items():
            sheet.cell(row=insert_row, column=col).value = data.get(key)
        
        for i, cell_data in enumerate(existing_data):
            new_row = insert_row + 1 + i
            sheet[f'E{new_row}'] = cell_data['day']
            sheet[f'F{new_row}'] = cell_data['discipline']
            sheet[f'G{new_row}'] = cell_data['group']
            sheet[f'H{new_row}'] = cell_data['load_type']
            for key, col in self.HOURS_COLS.items():
                sheet.cell(row=new_row, column=col).value = cell_data[key]
        
        return insert_row
    
    def add_and_save(self):
        """Одна функция для добавления записи и сохранения файла"""
        if not self.wb:
            messagebox.showerror("Ошибка", "Файл не загружен")
            return
        
        month = self.sheet_var.get()
        if not month:
            messagebox.showerror("Ошибка", "Выберите лист")
            return
        
        # Проверяем, что все обязательные поля заполнены
        if not all([self.entries['day'].get(), self.entries['discipline'].get(), 
                   self.entries['group'].get(), self.entries['load_type'].get()]):
            messagebox.showerror("Ошибка", "Заполните все обязательные поля")
            return
        
        try:
            data = {
                'day_number': int(self.entries['day'].get()),
                'discipline': self.entries['discipline'].get(),
                'group': self.entries['group'].get(),
                'load_type': self.entries['load_type'].get(),
                'lecture': float(self.entries['lecture'].get()) if self.entries['lecture'].get() else None,
                'practice': float(self.entries['practice'].get()) if self.entries['practice'].get() else None,
                'lab': float(self.entries['lab'].get()) if self.entries['lab'].get() else None
            }
        except ValueError:
            messagebox.showerror("Ошибка", "Проверьте правильность введенных данных")
            return
        
        if not 1 <= data['day_number'] <= 31:
            messagebox.showerror("Ошибка", "Число должно быть от 1 до 31")
            return
        
        try:
            sheet = self.wb[month]
            insert_row = self.find_insert_position(sheet, data['day_number'])
            result_row = self.insert_sorted_entry(sheet, insert_row, data)
            
            # Сохраняем файл
            self.wb.save(self.filename)
            
            # ПЕРЕЗАГРУЖАЕМ файл после сохранения
            self.reload_workbook()
            
            messagebox.showinfo("Успех", f"Запись добавлена в строку {result_row} и файл сохранен!")
            self.clear_entries()
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка: {e}")
            # Пытаемся перезагрузить файл при ошибке
            self.reload_workbook()
    
    def show_data(self):
        if not self.wb:
            return
        
        month = self.sheet_var.get()
        if not month or month not in self.wb.sheetnames:
            return
        
        # Очищаем таблицу
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Заполняем данными
        try:
            existing_days = self.get_existing_days(self.wb[month])
            
            for day in existing_days:
                self.tree.insert("", "end", values=(
                    day['day'],
                    day['discipline'] or '',
                    day['group'] or '',
                    day['load_type'] or '',
                    day['lecture'] or '',
                    day['practice'] or '',
                    day['lab'] or ''
                ))
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка отображения данных: {e}")
    
    def clear_entries(self):
        for entry in self.entries.values():
            entry.delete(0, tk.END)
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = JournalApp("ПРимер.xlsx")
    app.run()