import openpyxl
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, 
                              QLabel, QLineEdit, QComboBox, QPushButton, QTableWidget, QTableWidgetItem,
                              QHeaderView, QGroupBox, QFrame, QMessageBox, QFileDialog, QListWidget,
                              QListWidgetItem, QAbstractItemView, QMenu, QTextEdit, QScrollArea, QRadioButton,
                              QButtonGroup, QDateEdit, QProgressBar, QSplitter, QTabWidget, QCheckBox,
                              QDialog, QDialogButtonBox, QTextBrowser, QSizePolicy, QSpacerItem)
from PySide6.QtCore import Qt, QDate, QTimer, QPropertyAnimation, QEasingCurve, QRect, QSize
from PySide6.QtGui import QPalette, QColor, QFont, QPixmap, QMovie, QAction, QIcon, QPainter
from PySide6.QtCore import Property, QParallelAnimationGroup
from datetime import datetime, timedelta, date
import re
import os
import time
import sys
import json
from openpyxl.styles import Alignment

class JournalApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.filename = None
        self.wb = None
        self.START_ROW = 7
        self.HOURS_COLS = {'lecture': 12, 'practice': 13, 'lab': 14}
        self.selected_dates = []
        self.LOAD_TYPES = ["осн.", "почас.", "совм."]
        self.config_file = "app_config.json"
        
        self.load_config()
        self.setup_ui()
        
    def load_config(self):
        """Загружает конфигурацию приложения"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.filename = config.get('last_file')
        except Exception as e:
            print(f"Ошибка загрузки конфигурации: {e}")
    
    def save_config(self):
        """Сохраняет конфигурацию приложения"""
        try:
            config = {
                'last_file': self.filename
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Ошибка сохранения конфигурации: {e}")
    
    def setup_ui(self):
        """Настраивает пользовательский интерфейс"""
        self.setWindowTitle("Журнал преподавателя")
        self.setGeometry(100, 100, 1400, 900)
        
        # Настройка цветовой схемы
        self.setup_colors()
        
        # Создание центрального виджета
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Основной layout
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Создание виджетов
        self.create_menu_bar()
        self.create_file_section(main_layout)
        self.create_period_section(main_layout)
        self.create_input_section(main_layout)
        self.create_view_section(main_layout)
        
        # Если был сохранен последний файл, пытаемся загрузить его
        if self.filename and os.path.exists(self.filename):
            QTimer.singleShot(100, lambda: self.load_workbook(self.filename))
    
    def setup_colors(self):
        """Настраивает цветовую схему приложения"""
        palette = self.palette()
        palette.setColor(QPalette.Window, QColor(255, 230, 242))  # Розовый фон
        palette.setColor(QPalette.WindowText, QColor(139, 0, 139))  # ЧЕРНЫЙ текст (было 139, 0, 139)
        palette.setColor(QPalette.Base, QColor(255, 240, 245))  # Светло-розовый для полей ввода
        palette.setColor(QPalette.AlternateBase, QColor(255, 182, 193))  # Для чередующихся строк
        palette.setColor(QPalette.Button, QColor(255, 102, 178))  # Розовые кнопки
        palette.setColor(QPalette.ButtonText, QColor(0, 0, 0))  # ЧЕРНЫЙ текст кнопок (было 139, 0, 139)
        palette.setColor(QPalette.Highlight, QColor(152, 251, 152))  # Зеленая подсветка
        palette.setColor(QPalette.HighlightedText, QColor(0, 0, 0))  # ЧЕРНЫЙ текст подсветки (было 0, 100, 0)
        palette.setColor(QPalette.Text, QColor(0, 0, 0))  # ЧЕРНЫЙ основной текст
        palette.setColor(QPalette.BrightText, QColor(0, 0, 0))  # ЧЕРНЫЙ яркий текст
        self.setPalette(palette)
    
    def create_menu_bar(self):
        """Создает меню приложения"""
        menubar = self.menuBar()
        
        # Меню Файл
        file_menu = menubar.addMenu("Файл")
        open_action = QAction("Открыть файл Excel", self)
        open_action.triggered.connect(self.open_file)
        file_menu.addAction(open_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("Выход", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Меню Помощь
        help_menu = menubar.addMenu("Помощь")
        instructions_action = QAction("Инструкция", self)
        instructions_action.triggered.connect(self.show_instructions)
        help_menu.addAction(instructions_action)
        
        about_action = QAction("О программе", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
    
    def create_file_section(self, parent_layout):
        """Создает секцию выбора файла"""
        file_group = QGroupBox("Рабочий файл")
        file_layout = QHBoxLayout(file_group)
        
        # Поле пути к файлу
        self.file_path_label = QLabel("Файл не выбран")
        self.file_path_label.setStyleSheet("QLabel { background-color: #fff0f5; padding: 5px; border: 1px solid #ffb6c1; }")
        self.file_path_label.setMinimumHeight(30)
        file_layout.addWidget(self.file_path_label)
        
        # Кнопки управления файлом
        btn_layout = QHBoxLayout()
        self.select_file_btn = QPushButton("Выбрать файл")
        self.select_file_btn.setStyleSheet(self.get_action_button_style())
        self.select_file_btn.clicked.connect(self.open_file)
        btn_layout.addWidget(self.select_file_btn)
        
        file_layout.addLayout(btn_layout)
        
        parent_layout.addWidget(file_group)
    
    def create_period_section(self, parent_layout):
        """Создает секцию выбора периода"""
        period_group = QGroupBox("Выбор периода")
        period_layout = QGridLayout(period_group)
        
        # Даты
        period_layout.addWidget(QLabel("Начало периода:"), 0, 0)
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addMonths(-3))
        self.start_date.setCalendarPopup(True)
        self.start_date.setDisplayFormat("dd.MM.yyyy")
        period_layout.addWidget(self.start_date, 0, 1)
        
        period_layout.addWidget(QLabel("Конец периода:"), 0, 2)
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        self.end_date.setDisplayFormat("dd.MM.yyyy")
        period_layout.addWidget(self.end_date, 0, 3)
        
        # Тип недели
        period_layout.addWidget(QLabel("Тип недели:"), 0, 4)
        
        self.week_type_group = QButtonGroup(self)
        self.numerator_radio = QRadioButton("Числитель")
        self.denominator_radio = QRadioButton("Знаменатель")
        self.both_weeks_radio = QRadioButton("Обе недели")
        
        self.week_type_group.addButton(self.numerator_radio)
        self.week_type_group.addButton(self.denominator_radio)
        self.week_type_group.addButton(self.both_weeks_radio)
        
        self.numerator_radio.setChecked(True)
        
        radio_layout = QHBoxLayout()
        radio_layout.addWidget(self.numerator_radio)
        radio_layout.addWidget(self.denominator_radio)
        radio_layout.addWidget(self.both_weeks_radio)
        period_layout.addLayout(radio_layout, 0, 5)
        
        # Кнопки
        btn_layout = QHBoxLayout()
        self.generate_dates_btn = QPushButton("Сгенерировать даты по периоду")
        self.generate_dates_btn.setStyleSheet(self.get_action_button_style())
        self.generate_dates_btn.clicked.connect(self.generate_dates_by_period)
        btn_layout.addWidget(self.generate_dates_btn)
        
        self.clear_dates_btn = QPushButton("Очистить даты")
        self.clear_dates_btn.setStyleSheet(self.get_action_button_style())
        self.clear_dates_btn.clicked.connect(self.clear_dates)
        btn_layout.addWidget(self.clear_dates_btn)
        
        period_layout.addLayout(btn_layout, 1, 0, 1, 6)
        
        # Информация о датах
        self.dates_info_label = QLabel("Выбрано дат: 0")
        self.dates_info_label.setStyleSheet("QLabel { color: blue; font-weight: bold; }")
        period_layout.addWidget(self.dates_info_label, 2, 0, 1, 6)
        
        parent_layout.addWidget(period_group)
    
    def create_input_section(self, parent_layout):
        """Создает секцию ввода данных"""
        input_splitter = QSplitter(Qt.Horizontal)
        
        # Колонка 1: Управление датами
        dates_group = QGroupBox("Управление датами")
        dates_layout = QVBoxLayout(dates_group)
        
        # Добавление даты
        add_date_layout = QHBoxLayout()
        add_date_layout.addWidget(QLabel("Добавить дату:"))
        self.single_date = QDateEdit()
        self.single_date.setDate(QDate.currentDate())
        self.single_date.setCalendarPopup(True)
        self.single_date.setDisplayFormat("dd.MM.yyyy")
        add_date_layout.addWidget(self.single_date)
        
        self.add_date_btn = QPushButton("Добавить дату")
        self.add_date_btn.setStyleSheet(self.get_action_button_style())
        self.add_date_btn.clicked.connect(self.add_single_date)
        add_date_layout.addWidget(self.add_date_btn)
        
        dates_layout.addLayout(add_date_layout)
        
        # Удаление даты
        remove_date_layout = QHBoxLayout()
        remove_date_layout.addWidget(QLabel("Удалить дату:"))
        self.remove_date_combo = QComboBox()
        self.remove_date_combo.setMinimumWidth(150)
        remove_date_layout.addWidget(self.remove_date_combo)
        
        self.remove_date_btn = QPushButton("Удалить дату")
        self.remove_date_btn.setStyleSheet(self.get_action_button_style())
        self.remove_date_btn.clicked.connect(self.remove_selected_date)
        remove_date_layout.addWidget(self.remove_date_btn)
        
        dates_layout.addLayout(remove_date_layout)
        
        # Список выбранных дат
        dates_layout.addWidget(QLabel("Выбранные даты:"))
        self.dates_listbox = QListWidget()
        self.dates_listbox.setMinimumHeight(150)
        dates_layout.addWidget(self.dates_listbox)
        
        input_splitter.addWidget(dates_group)
        
        # Колонка 2: Поля ввода
        data_group = QGroupBox("Данные для заполнения")
        data_layout = QGridLayout(data_group)
        
        fields = [
            ("Дисциплина:", "discipline", "combobox"),
            ("Группа:", "group", "entry"),
            ("Вид нагрузки:", "load_type", "combobox"),
            ("Лекции:", "lecture", "entry"),
            ("Практические:", "practice", "entry"),
            ("Лабораторные:", "lab", "entry")
        ]
        
        self.entries = {}
        for i, (label, field, field_type) in enumerate(fields):
            data_layout.addWidget(QLabel(label), i, 0)
            
            if field_type == "combobox":
                if field == "discipline":
                    self.entries[field] = QComboBox()
                    self.entries[field].setEditable(True)
                else:
                    self.entries[field] = QComboBox()
                    self.entries[field].addItems(self.LOAD_TYPES)
            else:
                self.entries[field] = QLineEdit()
            
            data_layout.addWidget(self.entries[field], i, 1)
        
        input_splitter.addWidget(data_group)
        
        # Колонка 3: GIF/Анимация
        gif_group = QGroupBox("")
        gif_layout = QVBoxLayout(gif_group)
        
        # Загрузка и отображение GIF
        self.gif_label = QLabel("Журнал\nпреподавателя")
        self.gif_label.setAlignment(Qt.AlignCenter)
        self.gif_label.setStyleSheet("QLabel { font-size: 16px; font-weight: bold; color: #8b008b; }")
        self.gif_label.setMinimumSize(200, 200)
        gif_layout.addWidget(self.gif_label)
        
        # Пытаемся загрузить GIF
        self.load_and_display_gif()
        
        input_splitter.addWidget(gif_group)
        
        # Установка пропорций
        input_splitter.setSizes([300, 400, 200])
        
        parent_layout.addWidget(input_splitter)
        
        # Кнопка добавления записей
        self.add_entries_btn = QPushButton("Добавить записи")
        self.add_entries_btn.setStyleSheet(self.get_action_button_style())
        self.add_entries_btn.clicked.connect(self.add_entries)
        self.add_entries_btn.setEnabled(False)
        parent_layout.addWidget(self.add_entries_btn)
    
    def create_view_section(self, parent_layout):
        """Создает секцию просмотра данных"""
        view_group = QGroupBox("Просмотр и управление данными")
        view_layout = QVBoxLayout(view_group)
        
        # Панель управления просмотром
        control_layout = QHBoxLayout()
        
        control_layout.addWidget(QLabel("Лист:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.currentTextChanged.connect(self.show_data)
        control_layout.addWidget(self.sheet_combo)
        
        # Кнопки управления
        self.refresh_btn = QPushButton("Обновить данные")
        self.refresh_btn.setStyleSheet(self.get_action_button_style())
        self.refresh_btn.clicked.connect(self.show_data)
        control_layout.addWidget(self.refresh_btn)
        
        self.delete_btn = QPushButton("Удалить выбранные")
        self.delete_btn.setStyleSheet(self.get_danger_button_style())
        self.delete_btn.clicked.connect(self.delete_selected_entries)
        control_layout.addWidget(self.delete_btn)
        
        self.select_all_btn = QPushButton("Выбрать все")
        self.select_all_btn.setStyleSheet(self.get_action_button_style())
        self.select_all_btn.clicked.connect(self.select_all_entries)
        control_layout.addWidget(self.select_all_btn)
        
        self.deselect_btn = QPushButton("Снять выделение")
        self.deselect_btn.setStyleSheet(self.get_action_button_style())
        self.deselect_btn.clicked.connect(self.deselect_all_entries)
        control_layout.addWidget(self.deselect_btn)
        
        control_layout.addStretch()
        
        view_layout.addLayout(control_layout)
        
        # Информация о выборе
        self.selection_info = QLabel("Выбрано записей: 0")
        self.selection_info.setStyleSheet("QLabel { color: green; font-weight: bold; }")
        view_layout.addWidget(self.selection_info)
        
        # Таблица данных
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(7)
        self.table_widget.setHorizontalHeaderLabels(["Число", "Дисциплина", "Группа", "Нагрузка", "Лекции", "Практические", "Лабораторные"])
        
        # Настройка таблицы
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_widget.setSelectionMode(QAbstractItemView.MultiSelection)
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        self.table_widget.verticalHeader().setVisible(False)
        
        # Установка ширины колонок
        column_widths = [80, 250, 120, 100, 80, 100, 100]
        for i, width in enumerate(column_widths):
            self.table_widget.setColumnWidth(i, width)
        
        view_layout.addWidget(self.table_widget)
        
        parent_layout.addWidget(view_group)
    
    def get_action_button_style(self):
        """Возвращает стиль для кнопок действий"""
        return """
            QPushButton {
                background-color: #98fb98;
                color: #006400;
                border: 2px solid #90ee90;
                border-radius: 5px;
                padding: 5px 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #90ee90;
            }
            QPushButton:pressed {
                background-color: #7cfc00;
            }
        """
    
    def get_danger_button_style(self):
        """Возвращает стиль для опасных кнопок"""
        return """
            QPushButton {
                background-color: #ff6b6b;
                color: #8b0000;
                border: 2px solid #ff5252;
                border-radius: 5px;
                padding: 5px 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #ff5252;
            }
            QPushButton:pressed {
                background-color: #ff3838;
            }
        """
    
    def load_and_display_gif(self):
        """Загружает и отображает GIF изображение"""
        try:
            # Пытаемся найти GIF в разных местах
            possible_paths = [
                "Без названия.gif",
                "animation.gif",
                os.path.join(os.path.dirname(__file__), "Без названия.gif"),
                os.path.join(os.path.dirname(__file__), "animation.gif"),
            ]
            
            gif_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    gif_path = path
                    break
            
            if gif_path:
                self.movie = QMovie(gif_path)
                self.gif_label.setMovie(self.movie)
                self.movie.start()
            else:
                self.gif_label.setText("Журнал\nпреподавателя")
                
        except Exception as e:
            print(f"Ошибка загрузки GIF: {e}")
            self.gif_label.setText("Журнал\nпреподавателя")
    
    def open_file(self):
        """Открывает диалог выбора файла"""
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл Excel",
            "",
            "Excel files (*.xlsx *.xls *.xltx);;All files (*.*)"
        )
        
        if filename:
            self.load_workbook(filename)
    
    def load_workbook(self, filename):
        """Загружает рабочую книгу Excel"""
        try:
            self.close_workbook()
            self.wb = self.safe_load_workbook(filename)
            
            if self.wb:
                self.filename = filename
                self.file_path_label.setText(filename)
                self.add_entries_btn.setEnabled(True)
                
                # Обновляем список листов - только месячные листы
                all_sheets = self.wb.sheetnames
                monthly_sheets = self.filter_monthly_sheets(all_sheets)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(monthly_sheets)
                
                # Обновляем список дисциплин
                self.update_disciplines_list()
                
                # Сохраняем конфигурацию
                self.save_config()
                
                # Показываем данные
                self.show_data()
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки файла: {e}")
    
    def update_disciplines_list(self):
        """Обновляет список дисциплин из листов 'Осень' и 'Весна'"""
        if not self.wb:
            return
        
        disciplines = set()
        
        # Получаем дисциплины из листа 'Осень'
        if 'осень' in self.wb.sheetnames:
            sheet = self.wb['осень']
            row = 5
            while row <= 100:  # Проверяем до 100 строк
                try:
                    discipline = sheet[f'D{row}'].value
                    if discipline and isinstance(discipline, str) and discipline.strip():
                        disciplines.add(discipline.strip())
                except:
                    pass
                row += 2
        
        # Получаем дисциплины из листа 'Весна'
        if 'весна' in self.wb.sheetnames:
            sheet = self.wb['весна']
            row = 5
            while row <= 100:  # Проверяем до 100 строк
                try:
                    discipline = sheet[f'D{row}'].value
                    if discipline and isinstance(discipline, str) and discipline.strip():
                        disciplines.add(discipline.strip())
                except:
                    pass
                row += 2
        
        # Сортируем дисциплины по алфавиту
        sorted_disciplines = sorted(list(disciplines))
        
        # Обновляем комбобокс дисциплин
        if 'discipline' in self.entries:
            self.entries['discipline'].clear()
            self.entries['discipline'].addItems(sorted_disciplines)
    
    def filter_monthly_sheets(self, sheetnames):
        """Фильтрует листы, оставляя только месячные с номерами 01-12"""
        monthly_sheets = []
        for sheet in sheetnames:
            # Ищем листы, которые начинаются с номеров месяцев
            if re.match(r'^(09|10|11|12|01|02|03|04|05|06|07|08)', sheet):
                monthly_sheets.append(sheet)
        return monthly_sheets
    
    def show_instructions(self):
        """Показывает инструкцию по использованию"""
        instructions = """
ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ПРИЛОЖЕНИЯ "ЖУРНАЛ ПРЕПОДАВАТЕЛЯ"

1. ВЫБОР ФАЙЛА
   - Нажмите "Выбрать файл" для открытия существующего файла Excel

2. ВЫБОР ПЕРИОДА
   - Укажите начальную и конечную даты периода
   - Выберите тип недели: числитель, знаменатель или обе недели
   - Нажмите "Сгенерировать даты по периоду"

3. УПРАВЛЕНИЕ ДАТАМИ
   - Добавляйте отдельные даты через "Добавить дату"
   - Удаляйте даты через выпадающий список
   - Просматривайте выбранные даты в списке

4. ЗАПОЛНЕНИЕ ДАННЫХ
   - Введите название дисциплины (можно выбрать из списка или ввести новую)
   - Укажите группу
   - Выберите вид нагрузки
   - Заполните часы для лекций, практических и/или лабораторных работ

5. ДОБАВЛЕНИЕ ЗАПИСЕЙ
   - Нажмите "Добавить записи" для внесения данных в файл
   - Данные автоматически распределятся по соответствующим листам

6. ПРОСМОТР И УДАЛЕНИЕ
   - Выберите лист для просмотра данных
   - Выделите записи для удаления
   - Используйте кнопки управления для массовых операций

СОВЕТЫ:
- Всегда сохраняйте резервные копии важных файлов
- Проверяйте правильность введенных данных перед добавлением
- Используйте кнопку "Обновить данные" для актуальной информации
- Новые дисциплины автоматически добавляются в выпадающий список
        """
        self.show_info_dialog("Инструкция", instructions)
    
    def show_about(self):
        """Показывает информацию о программе"""
        about_text = """
ПРИЛОЖЕНИЕ "ЖУРНАЛ ПРЕПОДАВАТЕЛЯ"

Версия 2.1

Назначение:
Автоматизация заполнения журналов учебной нагрузки преподавателей

Возможности:
- Автоматическое распределение данных по месяцам
- Заполнение семестровых ведомостей
- Удобное управление датами занятий
- Просмотр и редактирование существующих записей
- Автоматическое сохранение списка дисциплин

Для корректной работы требуется:
- Файл Excel с соответствующими листами
- Права на запись в выбранный файл

Разработчики: RAMD        """
        self.show_info_dialog("О программе", about_text)
    
    def show_info_dialog(self, title, text):
        """Создает диалоговое окно с информацией"""
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setModal(True)
        dialog.resize(600, 500)
        
        layout = QVBoxLayout(dialog)
        
        text_edit = QTextBrowser()
        text_edit.setPlainText(text)
        layout.addWidget(text_edit)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok)
        button_box.accepted.connect(dialog.accept)
        layout.addWidget(button_box)
        
        dialog.exec()
    
    def close_workbook(self):
        """Безопасное закрытие рабочей книги"""
        try:
            if hasattr(self, 'wb') and self.wb:
                self.wb.close()
                self.wb = None
        except Exception as e:
            print(f"Ошибка при закрытии файла: {e}")
    
    def safe_save_workbook(self):
        """Безопасное сохранение рабочей книги с повторными попытками"""
        if not self.wb:
            QMessageBox.critical(self, "Ошибка", "Файл не загружен")
            return False
            
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if self.wb:
                    self.wb.save(self.filename)
                    return True
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
                else:
                    QMessageBox.critical(self, "Ошибка", 
                        f"Нет доступа к файлу {self.filename}!\n"
                        f"Убедитесь, что файл не открыт в другой программе.")
                    return False
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения файла: {e}")
                return False
        return False
    
    def safe_load_workbook(self, filename):
        """Безопасная загрузка рабочей книги с повторными попытками"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if os.path.exists(filename):
                    return openpyxl.load_workbook(filename)
                else:
                    QMessageBox.critical(self, "Ошибка", f"Файл {filename} не найден")
                    return None
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
                else:
                    QMessageBox.critical(self, "Ошибка", 
                        f"Нет доступа к файлу {filename}!\n"
                        f"Убедитесь, что файл не открыт в другой программе.")
                    return None
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки файла: {e}")
                return None
        return None

    def date_to_datetime(self, date_obj):
        return datetime.combine(date_obj, datetime.min.time()) if isinstance(date_obj, date) and not isinstance(date_obj, datetime) else date_obj

    def determine_week_type(self, input_date):
        current_date = self.date_to_datetime(input_date)
        year = current_date.year
        
        september_1 = datetime(year, 9, 1)
        if current_date < september_1:
            september_1 = datetime(year-1, 9, 1)
        
        days_diff = (current_date - september_1).days
        return "числитель" if (days_diff // 7) % 2 == 0 else "знаменатель"

    def find_sheet_for_month(self, month):
        if not self.wb:
            return None
            
        month_names = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",
                      7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        
        target_month = month_names.get(month)
        if not target_month:
            return None
        
        for sheet_name in self.wb.sheetnames:
            if re.search(r'\b' + re.escape(target_month) + r'\b', sheet_name):
                return sheet_name
        
        return None

    def generate_dates_by_period(self):
        try:
            start_dt = self.start_date.date().toPython()
            end_dt = self.end_date.date().toPython()
            
            # Определяем тип недели
            if self.numerator_radio.isChecked():
                target_week_type = "числитель"
            elif self.denominator_radio.isChecked():
                target_week_type = "знаменатель"
            else:
                target_week_type = "обе недели"
            
            if start_dt >= end_dt:
                QMessageBox.critical(self, "Ошибка", "Дата начала должна быть раньше даты окончания")
                return
            
            self.selected_dates.clear()
            generated_count = 0
            
            if target_week_type == "обе недели":
                current_date = start_dt
                while current_date <= end_dt:
                    sheet_name = self.find_sheet_for_month(current_date.month)
                    if sheet_name:
                        self.selected_dates.append({
                            'date': current_date, 'day': current_date.day, 'month': current_date.month,
                            'year': current_date.year, 'sheet': sheet_name,
                            'week_type': self.determine_week_type(current_date)
                        })
                        generated_count += 1
                    current_date += timedelta(days=7)
            else:
                current_date = start_dt
                while current_date <= end_dt:
                    week_type = self.determine_week_type(current_date)
                    if week_type == target_week_type:
                        sheet_name = self.find_sheet_for_month(current_date.month)
                        if sheet_name:
                            self.selected_dates.append({
                                'date': current_date, 'day': current_date.day, 'month': current_date.month,
                                'year': current_date.year, 'sheet': sheet_name, 'week_type': week_type
                            })
                            generated_count += 1
                        current_date += timedelta(days=7)
                    else:
                        current_date += timedelta(days=1)
            
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            self.update_dates_info()
            self.update_dates_display()
            
            if generated_count > 0:
                dates_list = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
                week_type_display = "числитель и знаменатель (каждую неделю)" if target_week_type == "обе недели" else target_week_type
                
                QMessageBox.information(self, "Успех", 
                    f"Сгенерировано {generated_count} дат\n"
                    f"Период: {start_dt.strftime('%d.%m.%Y')} - {end_dt.strftime('%d.%m.%Y')}\n"
                    f"Тип недели: {week_type_display}\n"
                    f"Даты: {dates_list}")
            else:
                QMessageBox.warning(self, "Внимание", "В выбранном периоде нет дат")
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка генерации дат: {e}")

    def add_single_date(self):
        try:
            date_obj = self.single_date.date().toPython()
            sheet_name = self.find_sheet_for_month(date_obj.month)
            
            if not sheet_name:
                QMessageBox.critical(self, "Ошибка", f"Не найден лист для месяца {date_obj.month}")
                return
            
            for existing_date in self.selected_dates:
                if (existing_date['day'] == date_obj.day and 
                    existing_date['month'] == date_obj.month and 
                    existing_date['year'] == date_obj.year):
                    QMessageBox.warning(self, "Внимание", "Эта дата уже есть в списке")
                    return
            
            date_info = {
                'date': date_obj, 
                'day': date_obj.day, 
                'month': date_obj.month,
                'year': date_obj.year, 
                'sheet': sheet_name,
                'week_type': self.determine_week_type(date_obj)
            }
            
            self.selected_dates.append(date_info)
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            
            self.update_dates_info()
            self.update_dates_display()
            
            QMessageBox.information(self, "Успех", f"Дата {date_obj.strftime('%d.%m.%Y')} добавлена")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка добавления даты: {e}")

    def remove_selected_date(self):
        selected_index = self.remove_date_combo.currentIndex()
        if selected_index == -1:
            QMessageBox.warning(self, "Внимание", "Выберите дату для удаления")
            return
        
        if 0 <= selected_index < len(self.selected_dates):
            removed_date = self.selected_dates.pop(selected_index)
            self.update_dates_info()
            self.update_dates_display()
            QMessageBox.information(self, "Успех", f"Дата {removed_date['day']}.{removed_date['month']:02d}.{removed_date['year']} удалена")

    def update_dates_display(self):
        self.dates_listbox.clear()
        date_values = []
        for date_info in self.selected_dates:
            display_text = f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']} ({date_info['sheet']}, {date_info['week_type']})"
            self.dates_listbox.addItem(display_text)
            date_values.append(f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']}")
        
        self.remove_date_combo.clear()
        self.remove_date_combo.addItems(date_values)

    def clear_dates(self):
        self.selected_dates.clear()
        self.update_dates_info()
        self.update_dates_display()
        QMessageBox.information(self, "Успех", "Все даты очищены")

    def update_dates_info(self):
        count = len(self.selected_dates)
        if count > 0:
            dates_str = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
            sheets_count = {}
            for date in self.selected_dates:
                sheets_count[date['sheet']] = sheets_count.get(date['sheet'], 0) + 1
            
            sheets_info = ", ".join([f"{sheet}: {count}" for sheet, count in sheets_count.items()])
            week_types = set(date['week_type'] for date in self.selected_dates)
            week_types_info = f"Типы: {', '.join(week_types)}" if week_types else ""
            
            self.dates_info_label.setText(f"Выбрано дат: {count} | Даты: {dates_str} | Листы: {sheets_info} {week_types_info}")
        else:
            self.dates_info_label.setText("Выбрано дат: 0")

    def show_data(self):
        """Показывает данные из выбранного листа"""
        if not self.wb:
            return
        
        sheet_name = self.sheet_combo.currentText()
        if sheet_name and sheet_name in self.wb.sheetnames:
            self.table_widget.setRowCount(0)
            
            try:
                sheet = self.wb[sheet_name]
                row = self.START_ROW
                
                while sheet[f'E{row}'].value is not None:
                    day = sheet[f'E{row}'].value
                    if isinstance(day, (int, float)):
                        current_row = self.table_widget.rowCount()
                        self.table_widget.insertRow(current_row)
                        
                        self.table_widget.setItem(current_row, 0, QTableWidgetItem(str(int(day))))
                        self.table_widget.setItem(current_row, 1, QTableWidgetItem(str(sheet[f'F{row}'].value or '')))
                        self.table_widget.setItem(current_row, 2, QTableWidgetItem(str(sheet[f'G{row}'].value or '')))
                        self.table_widget.setItem(current_row, 3, QTableWidgetItem(str(sheet[f'H{row}'].value or '')))
                        self.table_widget.setItem(current_row, 4, QTableWidgetItem(str(sheet.cell(row=row, column=12).value or '')))
                        self.table_widget.setItem(current_row, 5, QTableWidgetItem(str(sheet.cell(row=row, column=13).value or '')))
                        self.table_widget.setItem(current_row, 6, QTableWidgetItem(str(sheet.cell(row=row, column=14).value or '')))
                    
                    row += 1
                    
                self.update_selection_info()
                
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при чтении данных: {e}")

    def update_selection_info(self):
        """Обновляет информацию о количестве выбранных записей"""
        selected_count = len(self.table_widget.selectedItems()) // self.table_widget.columnCount()
        self.selection_info.setText(f"Выбрано записей: {selected_count}")

    def select_all_entries(self):
        """Выбирает все записи в таблице"""
        self.table_widget.selectAll()
        self.update_selection_info()

    def deselect_all_entries(self):
        """Снимает выделение со всех записей"""
        self.table_widget.clearSelection()
        self.update_selection_info()

    def delete_selected_entries(self):
        """Удаляет выбранные записи из таблицы и файла Excel"""
        if not self.wb:
            QMessageBox.critical(self, "Ошибка", "Файл не загружен")
            return
            
        selected_rows = set()
        for item in self.table_widget.selectedItems():
            selected_rows.add(item.row())
        
        if not selected_rows:
            QMessageBox.warning(self, "Внимание", "Выберите записи для удаления")
            return
        
        entries_to_delete = []
        for row in selected_rows:
            day = self.table_widget.item(row, 0).text()
            discipline = self.table_widget.item(row, 1).text()
            group = self.table_widget.item(row, 2).text()
            entries_to_delete.append((day, discipline, group))
        
        confirm = QMessageBox.question(
            self,
            "Подтверждение удаления", 
            f"Вы действительно хотите удалить {len(entries_to_delete)} записей?\n"
            f"Это действие нельзя отменить.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if confirm != QMessageBox.Yes:
            return
        
        try:
            # Закрываем и перезагружаем workbook для гарантии свежих данных
            self.close_workbook()
            self.wb = self.safe_load_workbook(self.filename)
            
            if not self.wb:
                return
            
            sheet_name = self.sheet_combo.currentText()
            if sheet_name not in self.wb.sheetnames:
                QMessageBox.critical(self, "Ошибка", "Лист не найден")
                return
            
            sheet = self.wb[sheet_name]
            deleted_count = 0
            
            rows_to_delete = []
            for entry_data in entries_to_delete:
                day = int(entry_data[0])
                discipline = entry_data[1]
                group = entry_data[2]
                
                row = self.START_ROW
                while sheet[f'E{row}'].value is not None:
                    sheet_day = sheet[f'E{row}'].value
                    sheet_discipline = sheet[f'F{row}'].value or ''
                    sheet_group = sheet[f'G{row}'].value or ''
                    
                    if (isinstance(sheet_day, (int, float)) and int(sheet_day) == day and
                        sheet_discipline == discipline and sheet_group == group):
                        rows_to_delete.append(row)
                        break
                    row += 1
            
            rows_to_delete.sort(reverse=True)
            for row_num in rows_to_delete:
                self.delete_row(sheet, row_num)
                deleted_count += 1
            
            # Сохраняем и перезагружаем файл
            if self.safe_save_workbook():
                self.show_data()
                QMessageBox.information(self, "Успех", f"Удалено записей: {deleted_count} из {len(entries_to_delete)}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении записей: {e}")

    def delete_row(self, sheet, row_num):
        """Удаляет строку из листа Excel и сдвигает остальные строки вверх"""
        max_row = self.START_ROW
        while sheet[f'E{max_row}'].value is not None:
            max_row += 1
        
        for row in range(row_num, max_row):
            sheet[f'E{row}'] = sheet[f'E{row + 1}'].value
            sheet[f'F{row}'] = sheet[f'F{row + 1}'].value
            sheet[f'G{row}'] = sheet[f'G{row + 1}'].value
            sheet[f'H{row}'] = sheet[f'H{row + 1}'].value
            sheet.cell(row=row, column=12).value = sheet.cell(row=row + 1, column=12).value
            sheet.cell(row=row, column=13).value = sheet.cell(row=row + 1, column=13).value
            sheet.cell(row=row, column=14).value = sheet.cell(row=row + 1, column=14).value
        
        # Очищаем последнюю строку (ВКЛЮЧАЯ КОЛОНКИ С ЧАСАМИ)
        sheet[f'E{max_row}'] = None
        sheet[f'F{max_row}'] = None
        sheet[f'G{max_row}'] = None
        sheet[f'H{max_row}'] = None
        sheet.cell(row=max_row, column=12).value = None
        sheet.cell(row=max_row, column=13).value = None
        sheet.cell(row=max_row, column=14).value = None

    def add_entries(self):
        """Добавляет записи в журнал"""
        if not self.wb:
            QMessageBox.critical(self, "Ошибка", "Файл не загружен")
            return
            
        if not all([self.wb, self.selected_dates, 
                   self.entries['discipline'].currentText() if hasattr(self.entries['discipline'], 'currentText') else self.entries['discipline'].text(),
                   self.entries['group'].text(), 
                   self.entries['load_type'].currentText() if hasattr(self.entries['load_type'], 'currentText') else self.entries['load_type'].text()]):
            QMessageBox.critical(self, "Ошибка", "Заполните все обязательные поля и сгенерируйте даты")
            return
        
        try:
            discipline = self.entries['discipline'].currentText() if hasattr(self.entries['discipline'], 'currentText') else self.entries['discipline'].text()
            group = self.entries['group'].text()
            load_type = self.entries['load_type'].currentText() if hasattr(self.entries['load_type'], 'currentText') else self.entries['load_type'].text()
            lecture = self.entries['lecture'].text()
            practice = self.entries['practice'].text()
            lab = self.entries['lab'].text()
            
            data = {
                'discipline': discipline,
                'group': group,
                'load_type': load_type,
                'lecture': float(lecture) if lecture else 0.0,
                'practice': float(practice) if practice else 0.0,
                'lab': float(lab) if lab else 0.0
            }
            
            if data['lecture'] == 0 and data['practice'] == 0 and data['lab'] == 0:
                QMessageBox.warning(self, "Внимание", "Заполните хотя бы одно поле: Лекции, Практические или Лабораторные")
                return
            
            # Закрываем и перезагружаем workbook для гарантии свежих данных
            self.close_workbook()
            self.wb = self.safe_load_workbook(self.filename)
            
            if not self.wb:
                return
            
            dates_by_sheet = {}
            for date_info in self.selected_dates:
                sheet_name = date_info['sheet']
                if sheet_name not in dates_by_sheet:
                    dates_by_sheet[sheet_name] = []
                dates_by_sheet[sheet_name].append(date_info)
            
            results = {}
            for sheet_name, dates in dates_by_sheet.items():
                if sheet_name not in self.wb.sheetnames:
                    continue
                    
                sheet = self.wb[sheet_name]
                dates.sort(key=lambda x: x['day'])
                
                added_rows = []
                for date_info in dates:
                    row = self.add_entry_to_sheet(sheet, date_info['day'], data)
                    if row:
                        added_rows.append(f"{date_info['day']}.{date_info['month']:02d}(стр.{row})")
                
                if added_rows:
                    results[sheet_name] = added_rows
            
            # Заполняем листы "Осень" и "Весна"
            season_results = self.fill_season_sheets(data)
            
            # Сохраняем файл
            if self.safe_save_workbook():
                # Обновляем отображение
                self.show_data()
                
                # Обновляем список дисциплин после добавления новых записей
                self.update_disciplines_list()
                
                msg_lines = ["Записи добавлены:"]
                for sheet_name, dates in results.items():
                    msg_lines.append(f"{sheet_name}: {', '.join(dates)}")
                
                if season_results:
                    msg_lines.append("\nСеместровые листы:")
                    for sheet_name, result in season_results.items():
                        msg_lines.append(f"{sheet_name}: {result}")
                
                if results or season_results:
                    QMessageBox.information(self, "Успех", "\n".join(msg_lines))
                    
                    # Очищаем поля ввода
                    for field in ['discipline', 'group', 'lecture', 'practice', 'lab']:
                        if field in self.entries:
                            if hasattr(self.entries[field], 'clear'):
                                self.entries[field].clear()
                    
                    if 'load_type' in self.entries:
                        self.entries['load_type'].setCurrentIndex(0)
                else:
                    QMessageBox.warning(self, "Внимание", "Не удалось добавить записи")
                
        except ValueError as e:
            QMessageBox.critical(self, "Ошибка", "Проверьте числовые поля (Лекции, Практические, Лабораторные) - они должны содержать только числа")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при добавлении записей: {e}")

    def fill_season_sheets(self, data):
        """Заполняет листы 'Осень' и 'Весна' данными и возвращает результаты"""
        season_results = {}
        
        try:
            if not self.selected_dates:
                return season_results
            
            months = set(date_info['month'] for date_info in self.selected_dates)
            
            autumn_months = {9, 10, 11, 12}
            spring_months = {1, 2, 3, 4, 5}
            
            fill_autumn = any(month in autumn_months for month in months)
            fill_spring = any(month in spring_months for month in months)
            
            if fill_autumn and 'осень' in self.wb.sheetnames:
                result = self.fill_season_sheet('осень', data)
                if result:
                    season_results['осень'] = result
            
            if fill_spring and 'весна' in self.wb.sheetnames:
                result = self.fill_season_sheet('весна', data)
                if result:
                    season_results['весна'] = result
                    
        except Exception as e:
            print(f"Ошибка при заполнении семестровых листов: {e}")
        
        return season_results

    def fill_season_sheet(self, sheet_name, data):
        """Заполняет конкретный семестровый лист и возвращает результат"""
        try:
            sheet = self.wb[sheet_name]
            
            row = 5
            max_rows_to_check = 50
            
            while row <= max_rows_to_check:
                try:
                    cell_d = sheet[f'D{row}']
                    cell_e = sheet[f'E{row}']
                    cell_f = sheet[f'F{row}']
                    
                    def is_cell_empty(cell):
                        try:
                            return cell.value is None or str(cell.value).strip() == ''
                        except:
                            return True
                    
                    if (is_cell_empty(cell_d) and 
                        is_cell_empty(cell_e) and 
                        is_cell_empty(cell_f)):
                        break
                except:
                    pass
                
                row += 2
            
            if row > max_rows_to_check:
                return f"Не найдено свободных строк (проверено до строки {max_rows_to_check})"
            
            try:
                sheet[f'D{row}'] = data['discipline']
                sheet[f'E{row}'] = data['group']
                sheet[f'F{row}'] = data['load_type']
                
                for col in ['D', 'E', 'F']:
                    sheet[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                return f"строка {row}: {data['discipline']}, {data['group']}, {data['load_type']}"
                
            except Exception as write_error:
                return f"Ошибка записи в строку {row}: {str(write_error)}"
            
        except Exception as e:
            return f"Ошибка: {str(e)}"

    def add_entry_to_sheet(self, sheet, day, data):
        """Добавляет запись в лист и возвращает номер строки"""
        try:
            last_row = self.START_ROW
            while sheet[f'E{last_row}'].value is not None:
                last_row += 1
            
            existing_date_rows = []
            row = self.START_ROW
            while sheet[f'E{row}'].value is not None:
                existing_day = sheet[f'E{row}'].value
                if isinstance(existing_day, (int, float)) and int(existing_day) == day:
                    existing_date_rows.append(row)
                row += 1
            
            if existing_date_rows:
                last_date_row = max(existing_date_rows)
                insert_row = last_date_row + 1
                self.shift_rows_down(sheet, insert_row, last_row)
                self.fill_row_data(sheet, insert_row, day, data)
                return insert_row
            else:
                return self.insert_entry_sorted(sheet, day, data, last_row)
        except Exception as e:
            print(f"Ошибка при добавлении записи: {e}")
            return None

    def shift_rows_down(self, sheet, start_row, last_row):
        """Сдвигает строки вниз"""
        for row in range(last_row, start_row - 1, -1):
            sheet[f'E{row+1}'] = sheet[f'E{row}'].value
            sheet[f'F{row+1}'] = sheet[f'F{row}'].value
            sheet[f'G{row+1}'] = sheet[f'G{row}'].value
            sheet[f'H{row+1}'] = sheet[f'H{row}'].value
            sheet.cell(row=row+1, column=12).value = sheet.cell(row=row, column=12).value
            sheet.cell(row=row+1, column=13).value = sheet.cell(row=row, column=13).value
            sheet.cell(row=row+1, column=14).value = sheet.cell(row=row, column=14).value
            
            # ОЧИЩАЕМ ИСХОДНУЮ СТРОКУ ПРИ СДВИГЕ (если это строка вставки)
            if row == start_row:
                sheet.cell(row=row, column=12).value = None
                sheet.cell(row=row, column=13).value = None
                sheet.cell(row=row, column=14).value = None

    def insert_entry_sorted(self, sheet, day, data, last_row):
        """Вставляет запись в отсортированном порядке"""
        insert_row = self.START_ROW
        while sheet[f'E{insert_row}'].value is not None:
            existing_day = sheet[f'E{insert_row}'].value
            if isinstance(existing_day, (int, float)) and int(existing_day) > day:
                break
            insert_row += 1
        
        self.shift_rows_down(sheet, insert_row, last_row)
        self.fill_row_data(sheet, insert_row, day, data)
        return insert_row

    def fill_row_data(self, sheet, row, day, data):
        """Заполняет строку данными - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
        sheet[f'E{row}'] = day
        sheet[f'F{row}'] = data['discipline']
        sheet[f'G{row}'] = data['group']
        sheet[f'H{row}'] = data['load_type']
        
        # ОЧИЩАЕМ ВСЕ КОЛОНКИ С ЧАСАМИ ПЕРЕД ЗАПОЛНЕНИЕМ
        sheet.cell(row=row, column=12).value = None  # Лекции (J)
        sheet.cell(row=row, column=13).value = None  # Практические (K)
        sheet.cell(row=row, column=14).value = None  # Лабораторные (L)
        
        # ЗАПОЛНЯЕМ ТОЛЬКО ТЕ КОЛОНКИ, ГДЕ ЕСТЬ ЧАСЫ
        if data['lecture'] != 0:
            sheet.cell(row=row, column=12).value = data['lecture']
        if data['practice'] != 0:
            sheet.cell(row=row, column=13).value = data['practice']
        if data['lab'] != 0:
            sheet.cell(row=row, column=14).value = data['lab']

    def closeEvent(self, event):
        """Обработчик закрытия приложения"""
        self.close_workbook()
        self.save_config()
        event.accept()

def main():
    """Основная функция запуска приложения"""
    try:
        app = QApplication(sys.argv)
        
        # Установка стиля приложения
        app.setStyle('Fusion')
        
        window = JournalApp()
        window.show()
        
        sys.exit(app.exec())
    except Exception as e:
        print(f"Не удалось запустить приложение: {e}")
        QMessageBox.critical(None, "Ошибка запуска", f"Не удалось запустить приложение: {e}")

if __name__ == "__main__":
    main()