import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import Calendar, DateEntry
from datetime import datetime, timedelta, date
import re
from PIL import Image, ImageTk
from openpyxl.styles import Alignment
import os
import time
import sys
import json

class JournalApp:
    def __init__(self):
        self.filename = None
        self.wb = None
        self.START_ROW = 7
        self.HOURS_COLS = {'lecture': 12, 'practice': 13, 'lab': 14}
        self.selected_dates = []
        self.LOAD_TYPES = ["осн.", "почас.", "совм."]
        self.config_file = "app_config.json"
        
        self.load_config()
        self.setup_gui()
        
    def load_config(self):
        """Загружает конфигурацию приложения"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.filename = config.get('last_file')
        except Exception as e:
            print(f"Ошибка загрузки конфигурации: {e}")
    
    def save_config(self):
        """Сохраняет конфигурацию приложения"""
        try:
            config = {
                'last_file': self.filename
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Ошибка сохранения конфигурации: {e}")
    
    def setup_gui(self):
        self.root = tk.Tk()
        self.root.title("Журнал преподавателя")
        self.root.geometry("1200x900")
        
        # Настройка цветовой схемы - розовые и зеленые тона
        self.setup_colors()
        
        # Обработчик закрытия окна
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Создаем меню
        self.create_menu()
        
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Фрейм выбора файла
        file_frame = self.create_file_frame(main_frame)
        file_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Выбор периода
        period_frame = self.create_period_frame(main_frame)
        period_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Три колонки: управление датами, поля ввода, GIF
        columns_frame = ttk.Frame(main_frame)
        columns_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Колонка 1: Управление датами
        dates_frame = self.create_dates_frame(columns_frame)
        dates_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))
        
        # Колонка 2: Поля ввода
        input_frame = self.create_input_frame(columns_frame)
        input_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))
        
        # Колонка 3: GIF
        gif_frame = self.create_gif_frame(columns_frame)
        gif_frame.grid(row=0, column=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Кнопка добавления записей
        self.add_btn = ttk.Button(main_frame, text="Добавить записи", command=self.add_entries, state="disabled")
        self.add_btn.grid(row=3, column=0, columnspan=3, pady=15)
        
        # Просмотр данных
        view_frame = self.create_view_frame(main_frame)
        view_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # Настройка растягивания
        self.configure_grid(main_frame, columns_frame, view_frame)
        
        # Если был сохранен последний файл, пытаемся загрузить его
        if self.filename and os.path.exists(self.filename):
            self.load_workbook(self.filename)
    
    def setup_colors(self):
        """Настраивает цветовую схему приложения"""
        style = ttk.Style()
        
        # Розовая цветовая схема для основных элементов
        style.configure("TFrame", background="#ffe6f2")
        style.configure("TLabel", background="#ffe6f2", foreground="#8b008b")
        style.configure("TButton", background="#ff66b2", foreground="#8b008b")
        style.configure("TLabelframe", background="#ffe6f2", foreground="#8b008b")
        style.configure("TLabelframe.Label", background="#ffe6f2", foreground="#8b008b")
        style.configure("TEntry", fieldbackground="#fff0f5")
        style.configure("TCombobox", fieldbackground="#fff0f5")
        
        # Зеленая цветовая схема для кнопок действий
        style.configure("Action.TButton", background="#98fb98", foreground="#006400")
        style.configure("Danger.TButton", background="#ff6b6b", foreground="#8b0000")
        
        # Настройка для Treeview
        style.configure("Treeview", background="#fff0f5", fieldbackground="#fff0f5", foreground="#8b008b")
        style.configure("Treeview.Heading", background="#ffb6c1", foreground="#8b008b")
        
        # Настройка главного окна
        self.root.configure(background="#ffe6f2")
    
    def create_menu(self):
        """Создает меню приложения"""
        menubar = tk.Menu(self.root, bg="#ffb6c1", fg="#8b008b")
        
        # Меню Файл
        file_menu = tk.Menu(menubar, tearoff=0, bg="#ffe6f2", fg="#8b008b")
        file_menu.add_command(label="Открыть файл Excel", command=self.open_file)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.on_closing)
        menubar.add_cascade(label="Файл", menu=file_menu)
        
        # Меню Помощь
        help_menu = tk.Menu(menubar, tearoff=0, bg="#ffe6f2", fg="#8b008b")
        help_menu.add_command(label="Инструкция", command=self.show_instructions)
        help_menu.add_command(label="О программе", command=self.show_about)
        menubar.add_cascade(label="Помощь", menu=help_menu)
        
        self.root.config(menu=menubar)
    
    def create_file_frame(self, parent):
        """Создает фрейм выбора файла"""
        frame = ttk.LabelFrame(parent, text="Рабочий файл", padding="10")
        
        # Поле пути к файлу
        self.file_path_var = tk.StringVar()
        file_entry = ttk.Entry(frame, textvariable=self.file_path_var, width=80, state="readonly")
        file_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        # Кнопки управления файлом
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=0, column=1, sticky=tk.W)
        
        ttk.Button(btn_frame, text="Выбрать файл", command=self.open_file, style="Action.TButton").pack(side=tk.LEFT, padx=2)
        
        # Информация о статусе
        self.file_status_label = ttk.Label(frame, text="Файл не выбран", foreground="red")
        self.file_status_label.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))
        
        frame.columnconfigure(0, weight=1)
        
        return frame
    
    def open_file(self):
        """Открывает диалог выбора файла"""
        filename = filedialog.askopenfilename(
            title="Выберите файл Excel",
            filetypes=[
                ("Excel files", "*.xlsx *.xls *.xltx"),
                ("All files", "*.*")
            ]
        )
        
        if filename:
            self.load_workbook(filename)
    
    def load_workbook(self, filename):
        """Загружает рабочую книгу Excel"""
        try:
            self.close_workbook()
            self.wb = self.safe_load_workbook(filename)
            
            if self.wb:
                self.filename = filename
                self.file_path_var.set(filename)
                self.file_status_label.config(text="Файл успешно загружен", foreground="green")
                self.add_btn.config(state="normal")
                
                # Обновляем список листов - только месячные листы
                all_sheets = self.wb.sheetnames
                monthly_sheets = self.filter_monthly_sheets(all_sheets)
                self.sheet_combo['values'] = monthly_sheets
                if monthly_sheets:
                    self.sheet_var.set(monthly_sheets[0])
                self.show_data()
                
                # Обновляем список дисциплин
                self.update_disciplines_list()
                
                # Сохраняем конфигурацию
                self.save_config()
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка загрузки файла: {e}")
            self.file_status_label.config(text="Ошибка загрузки файла", foreground="red")
    
    def update_disciplines_list(self):
        """Обновляет список дисциплин из листов 'Осень' и 'Весна'"""
        if not self.wb:
            return
        
        disciplines = set()
        
        # Получаем дисциплины из листа 'Осень'
        if 'осень' in self.wb.sheetnames:
            sheet = self.wb['осень']
            row = 5
            while row <= 100:  # Проверяем до 100 строк
                try:
                    discipline = sheet[f'D{row}'].value
                    if discipline and isinstance(discipline, str) and discipline.strip():
                        disciplines.add(discipline.strip())
                except:
                    pass
                row += 2
        
        # Получаем дисциплины из листа 'Весна'
        if 'весна' in self.wb.sheetnames:
            sheet = self.wb['весна']
            row = 5
            while row <= 100:  # Проверяем до 100 строк
                try:
                    discipline = sheet[f'D{row}'].value
                    if discipline and isinstance(discipline, str) and discipline.strip():
                        disciplines.add(discipline.strip())
                except:
                    pass
                row += 2
        
        # Сортируем дисциплины по алфавиту
        sorted_disciplines = sorted(list(disciplines))
        
        # Обновляем комбобокс дисциплин
        if hasattr(self, 'entries') and 'discipline' in self.entries:
            self.entries['discipline']['values'] = sorted_disciplines
    
    def filter_monthly_sheets(self, sheetnames):
        """Фильтрует листы, оставляя только месячные с номерами 01-12"""
        monthly_sheets = []
        for sheet in sheetnames:
            # Ищем листы, которые начинаются с номеров месяцев
            if re.match(r'^(09|10|11|12|01|02|03|04|05|06|07|08)', sheet):
                monthly_sheets.append(sheet)
        return monthly_sheets
    
    def show_instructions(self):
        """Показывает инструкцию по использованию"""
        instructions = """
ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ПРИЛОЖЕНИЯ "ЖУРНАЛ ПРЕПОДАВАТЕЛЯ"

1. ВЫБОР ФАЙЛА
   - Нажмите "Выбрать файл" для открытия существующего файла Excel

2. ВЫБОР ПЕРИОДА
   - Укажите начальную и конечную даты периода
   - Выберите тип недели: числитель, знаменатель или обе недели
   - Нажмите "Сгенерировать даты по периоду"

3. УПРАВЛЕНИЕ ДАТАМИ
   - Добавляйте отдельные даты через "Добавить дату"
   - Удаляйте даты через выпадающий список
   - Просматривайте выбранные даты в списке

4. ЗАПОЛНЕНИЕ ДАННЫХ
   - Введите название дисциплины (можно выбрать из списка или ввести новую)
   - Укажите группу
   - Выберите вид нагрузки
   - Заполните часы для лекций, практических и/или лабораторных работ

5. ДОБАВЛЕНИЕ ЗАПИСЕЙ
   - Нажмите "Добавить записи" для внесения данных в файл
   - Данные автоматически распределятся по соответствующим листам

6. ПРОСМОТР И УДАЛЕНИЕ
   - Выберите лист для просмотра данных
   - Выделите записи для удаления
   - Используйте кнопки управления для массовых операций

СОВЕТЫ:
- Всегда сохраняйте резервные копии важных файлов
- Проверяйте правильность введенных данных перед добавлением
- Используйте кнопку "Обновить данные" для актуальной информации
- Новые дисциплины автоматически добавляются в выпадающий список
        """
        self.show_info_window("Инструкция", instructions)
    
    def show_about(self):
        """Показывает информацию о программе"""
        about_text = """
ПРИЛОЖЕНИЕ "ЖУРНАЛ ПРЕПОДАВАТЕЛЯ"

Версия 2.1

Назначение:
Автоматизация заполнения журналов учебной нагрузки преподавателей

Возможности:
- Автоматическое распределение данных по месяцам
- Заполнение семестровых ведомостей
- Удобное управление датами занятий
- Просмотр и редактирование существующих записей
- Автоматическое сохранение списка дисциплин

Для корректной работы требуется:
- Файл Excel с соответствующими листами
- Права на запись в выбранный файл

Разработчики: RAMD        """
        self.show_info_window("О программе", about_text)
    
    def show_info_window(self, title, text):
        """Создает окно с информацией"""
        window = tk.Toplevel(self.root)
        window.title(title)
        window.geometry("600x500")
        window.transient(self.root)
        window.grab_set()
        window.configure(background="#ffe6f2")
        
        frame = ttk.Frame(window, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        text_widget = tk.Text(frame, wrap=tk.WORD, width=70, height=30, bg="#fff0f5", fg="#8b008b")
        text_widget.insert(1.0, text)
        text_widget.config(state=tk.DISABLED)
        text_widget.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        ttk.Button(frame, text="Закрыть", command=window.destroy, style="Action.TButton").grid(row=1, column=0, pady=10)
        
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        window.columnconfigure(0, weight=1)
        window.rowconfigure(0, weight=1)

    def on_closing(self):
        """Обработчик закрытия приложения"""
        self.close_workbook()
        self.save_config()
        self.root.destroy()
    
    def close_workbook(self):
        """Безопасное закрытие рабочей книги"""
        try:
            if hasattr(self, 'wb') and self.wb:
                self.wb.close()
                self.wb = None
        except Exception as e:
            print(f"Ошибка при закрытии файла: {e}")
    
    def safe_save_workbook(self):
        """Безопасное сохранение рабочей книги с повторными попытками"""
        if not self.wb:
            messagebox.showerror("Ошибка", "Файл не загружен")
            return False
            
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if self.wb:
                    self.wb.save(self.filename)
                    return True
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
                else:
                    messagebox.showerror("Ошибка", 
                        f"Нет доступа к файлу {self.filename}!\n"
                        f"Убедитесь, что файл не открыт в другой программе.")
                    return False
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка сохранения файла: {e}")
                return False
        return False
    
    def safe_load_workbook(self, filename):
        """Безопасная загрузка рабочей книги с повторными попытками"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if os.path.exists(filename):
                    return openpyxl.load_workbook(filename)
                else:
                    messagebox.showerror("Ошибка", f"Файл {filename} не найден")
                    return None
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
                else:
                    messagebox.showerror("Ошибка", 
                        f"Нет доступа к файлу {filename}!\n"
                        f"Убедитесь, что файл не открыт в другой программе.")
                    return None
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка загрузки файла: {e}")
                return None
        return None

    def create_period_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="Выбор периода", padding="10")
        
        # Даты
        ttk.Label(frame, text="Начало периода:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.start_date = DateEntry(frame, width=12, background='#ff66b2', foreground='white', 
                                  borderwidth=2, date_pattern='dd.mm.yyyy', locale='ru_RU')
        self.start_date.set_date(datetime(datetime.now().year, 9, 1))
        self.start_date.grid(row=0, column=1, sticky=tk.W, padx=(0, 15))
        
        ttk.Label(frame, text="Конец периода:").grid(row=0, column=2, sticky=tk.W, padx=(0, 5))
        self.end_date = DateEntry(frame, width=12, background='#ff66b2', foreground='white',
                                borderwidth=2, date_pattern='dd.mm.yyyy', locale='ru_RU')
        self.end_date.set_date(datetime(datetime.now().year, 12, 31))
        self.end_date.grid(row=0, column=3, sticky=tk.W, padx=(0, 15))
        
        # Тип недели
        ttk.Label(frame, text="Тип недели:").grid(row=0, column=4, sticky=tk.W, padx=(0, 5))
        self.period_week_type = tk.StringVar(value="числитель")
        
        week_type_frame = ttk.Frame(frame)
        week_type_frame.grid(row=0, column=5, columnspan=3, sticky=tk.W)
        
        for text, value in [("Числитель", "числитель"), ("Знаменатель", "знаменатель"), ("Обе недели", "обе недели")]:
            ttk.Radiobutton(week_type_frame, text=text, variable=self.period_week_type, value=value).pack(side=tk.LEFT, padx=(0, 5))
        
        # Кнопки
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=1, column=0, columnspan=7, pady=10, sticky=tk.W)
        
        ttk.Button(btn_frame, text="Сгенерировать даты по периоду", command=self.generate_dates_by_period, style="Action.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Очистить даты", command=self.clear_dates, style="Action.TButton").pack(side=tk.LEFT, padx=5)
        
        # Информация о датах
        self.dates_info_label = ttk.Label(frame, text="Выбрано дат: 0", foreground="blue", font=('Arial', 10, 'bold'))
        self.dates_info_label.grid(row=2, column=0, columnspan=7, sticky=tk.W, pady=2)
        
        return frame

    def create_dates_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="Управление датами", padding="10")
        
        # Добавление даты
        ttk.Label(frame, text="Добавить дату:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.single_date = DateEntry(frame, width=12, background='#ff66b2', foreground='white',
                                   borderwidth=2, date_pattern='dd.mm.yyyy', locale='ru_RU')
        self.single_date.grid(row=0, column=1, sticky=tk.W, padx=(0, 15))
        
        ttk.Button(frame, text="Добавить дату", command=self.add_single_date, style="Action.TButton").grid(row=0, column=2, sticky=tk.W, padx=5)
        
        # Удаление даты
        ttk.Label(frame, text="Удалить дату:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5), pady=(10, 0))
        self.remove_date_combo = ttk.Combobox(frame, width=15, state="readonly")
        self.remove_date_combo.grid(row=1, column=1, sticky=tk.W, padx=(0, 15), pady=(10, 0))
        
        ttk.Button(frame, text="Удалить дату", command=self.remove_selected_date, style="Action.TButton").grid(row=1, column=2, sticky=tk.W, padx=5, pady=(10, 0))
        
        # Список выбранных дат
        ttk.Label(frame, text="Выбранные даты:").grid(row=2, column=0, sticky=tk.W, pady=(10, 5))
        
        self.dates_listbox = tk.Listbox(frame, width=40, height=8, bg="#fff0f5", fg="#8b008b")
        self.dates_listbox.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # Кнопки управления списком
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=5, sticky=tk.W)
        
        ttk.Button(btn_frame, text="Очистить все даты", command=self.clear_dates, style="Action.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Обновить список", command=self.update_dates_display, style="Action.TButton").pack(side=tk.LEFT, padx=5)
        
        return frame

    def create_input_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="Данные для заполнения", padding="10")
        
        fields = [
            ("Дисциплина:", "discipline", "combobox"),  # Изменено на combobox
            ("Группа:", "group", "entry"),
            ("Вид нагрузки:", "load_type", "combobox"),
            ("Лекции:", "lecture", "entry"),
            ("Практические:", "practice", "entry"),
            ("Лабораторные:", "lab", "entry")
        ]
        
        self.entries = {}
        for i, (label, field, field_type) in enumerate(fields):
            ttk.Label(frame, text=label).grid(row=i, column=0, sticky=tk.W, pady=5, padx=(0, 10))
            
            if field_type == "combobox":
                if field == "discipline":
                    # Комбобокс для дисциплин с возможностью ввода своего значения
                    self.entries[field] = ttk.Combobox(frame, width=25)
                    self.entries[field].config(values=[])  # Начально пустой список
                else:
                    self.entries[field] = ttk.Combobox(frame, values=self.LOAD_TYPES, width=25)
                    self.entries[field].set(self.LOAD_TYPES[0])
            else:
                # Используем обычный tk.Entry вместо ttk.Entry для лучшей поддержки вставки
                self.entries[field] = tk.Entry(frame, width=30, bg="#fff0f5", fg="#8b008b")
            
            self.entries[field].grid(row=i, column=1, sticky=(tk.W, tk.E), pady=5)
            
            # Разрешаем вставку для полей "Дисциплина" и "Группа"
            if field in ['discipline', 'group']:
                # Привязываем комбинации клавиш для обеих раскладок
                self.entries[field].bind('<Control-KeyPress>', self.handle_keypress)
                self.entries[field].bind('<Control-Key>', self.handle_keypress)
                
                # Также добавляем контекстное меню для надежности
                self.create_context_menu(self.entries[field])
        
        # Настройка растягивания для колонки с полями ввода
        frame.columnconfigure(1, weight=1)
        
        return frame

    def create_context_menu(self, widget):
        """Создает контекстное меню для поля ввода"""
        context_menu = tk.Menu(widget, tearoff=0)
        context_menu.add_command(label="Вырезать", command=lambda: self.cut_text(widget))
        context_menu.add_command(label="Копировать", command=lambda: self.copy_text(widget))
        context_menu.add_command(label="Вставить", command=lambda: self.paste_text(widget))
        context_menu.add_command(label="Отменить", command=lambda: self.undo_text(widget))
        context_menu.add_separator()
        context_menu.add_command(label="Выделить все", command=lambda: self.select_all_text(widget))
        
        widget.bind("<Button-3>", lambda event: self.show_context_menu(event, context_menu))

    def show_context_menu(self, event, menu):
        """Показывает контекстное меню"""
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def handle_keypress(self, event):
        """Обработчик комбинаций клавиш для обеих раскладок"""
        # Получаем код клавиши
        keycode = event.keycode
        
        # В Windows коды клавиш для Ctrl+C, Ctrl+V, Ctrl+X, Ctrl+A, Ctrl+Z
        # одинаковы независимо от раскладки
        if event.state & 0x4:  # Проверяем, что нажат Ctrl
            if keycode == 67:  # Ctrl+C (копирование) - 67 для английской C, работает и для русской С
                self.copy_text(event)
                return "break"
            elif keycode == 86:  # Ctrl+V (вставка) - 86 для английской V, работает и для русской М
                self.paste_text(event)
                return "break"
            elif keycode == 88:  # Ctrl+X (вырезание) - 88 для английской X, работает и для русской Ч
                self.cut_text(event)
                return "break"
            elif keycode == 65:  # Ctrl+A (выделение всего) - 65 для английской A, работает и для русской Ф
                self.select_all_text(event)
                return "break"
            elif keycode == 90:  # Ctrl+Z (отмена) - 90 для английской Z, работает и для русской Я
                self.undo_text(event)
                return "break"

    def copy_text(self, event_or_widget):
        """Копирование текста"""
        try:
            if isinstance(event_or_widget, tk.Event):
                widget = event_or_widget.widget
            else:
                widget = event_or_widget
            
            if widget.select_present():
                selected_text = widget.selection_get()
                widget.clipboard_clear()
                widget.clipboard_append(selected_text)
        except Exception as e:
            print(f"Ошибка копирования: {e}")

    def paste_text(self, event_or_widget):
        """Вставка текста"""
        try:
            if isinstance(event_or_widget, tk.Event):
                widget = event_or_widget.widget
            else:
                widget = event_or_widget
            
            clipboard_text = widget.clipboard_get()
            
            if widget.select_present():
                # Заменяем выделенный текст
                start = widget.index(tk.SEL_FIRST)
                end = widget.index(tk.SEL_LAST)
                widget.delete(start, end)
                widget.insert(start, clipboard_text)
            else:
                # Вставляем в позицию курсора
                widget.insert(tk.INSERT, clipboard_text)
        except Exception as e:
            print(f"Ошибка вставки: {e}")

    def cut_text(self, event_or_widget):
        """Вырезание текста"""
        try:
            if isinstance(event_or_widget, tk.Event):
                widget = event_or_widget.widget
            else:
                widget = event_or_widget
            
            if widget.select_present():
                self.copy_text(widget)  # Сначала копируем
                start = widget.index(tk.SEL_FIRST)
                end = widget.index(tk.SEL_LAST)
                widget.delete(start, end)  # Затем удаляем
        except Exception as e:
            print(f"Ошибка вырезания: {e}")

    def select_all_text(self, event_or_widget):
        """Выделение всего текста"""
        try:
            if isinstance(event_or_widget, tk.Event):
                widget = event_or_widget.widget
            else:
                widget = event_or_widget
            
            widget.select_range(0, tk.END)
            widget.icursor(tk.END)
        except Exception as e:
            print(f"Ошибка выделения текста: {e}")

    def undo_text(self, event_or_widget):
        """Отмена последнего действия"""
        try:
            if isinstance(event_or_widget, tk.Event):
                widget = event_or_widget.widget
            else:
                widget = event_or_widget
            
            # Пытаемся использовать стандартный механизм отмены
            widget.event_generate('<<Undo>>')
        except Exception as e:
            print(f"Ошибка отмены: {e}")

    def create_gif_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="", padding="10")
        
        # Загрузка и отображение GIF
        self.load_and_display_gif(frame)
        
        return frame

    def load_and_display_gif(self, parent):
        """Загружает и отображает GIF изображение"""
        try:
            # Пытаемся найти GIF в разных местах
            possible_paths = [
                "Без названия.gif",
                "animation.gif",
                os.path.join(os.path.dirname(__file__), "Без названия.gif"),
                os.path.join(os.path.dirname(__file__), "animation.gif"),
            ]
            
            gif_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    gif_path = path
                    break
            
            if gif_path:
                self.gif_image = Image.open(gif_path)
                
                self.gif_frames = []
                try:
                    while True:
                        frame = self.gif_image.copy()
                        frame = frame.resize((200, 200), Image.Resampling.LANCZOS)
                        self.gif_frames.append(ImageTk.PhotoImage(frame))
                        self.gif_image.seek(len(self.gif_frames))
                except EOFError:
                    pass
                    
                if not self.gif_frames:
                    frame = self.gif_image.copy()
                    frame = frame.resize((200, 200), Image.Resampling.LANCZOS)
                    self.gif_frames = [ImageTk.PhotoImage(frame)]
                
                self.gif_label = ttk.Label(parent)
                self.gif_label.pack(expand=True)
                
                if len(self.gif_frames) > 1:
                    self.current_frame = 0
                    self.animate_gif()
                else:
                    self.gif_label.configure(image=self.gif_frames[0])
            else:
                # Если GIF не найден, показываем информационную метку
                info_label = ttk.Label(parent, text="Журнал\nстрадателя", 
                                      font=('Arial', 16, 'bold'), justify=tk.CENTER)
                info_label.pack(expand=True)
                
        except Exception as e:
            # В случае ошибки показываем текстовую метку
            info_label = ttk.Label(parent, text="Журнал\nпреподавателя", 
                                  font=('Arial', 16, 'bold'), justify=tk.CENTER)
            info_label.pack(expand=True)

    def animate_gif(self):
        """Анимирует GIF изображение"""
        if hasattr(self, 'gif_frames') and self.gif_frames and hasattr(self, 'root'):
            try:
                self.gif_label.configure(image=self.gif_frames[self.current_frame])
                self.current_frame = (self.current_frame + 1) % len(self.gif_frames)
                self.root.after(200, self.animate_gif)
            except Exception as e:
                print(f"Ошибка анимации GIF: {e}")

    def create_view_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="Просмотр и управление данными", padding="10")
        
        # Панель управления просмотром
        control_frame = ttk.Frame(frame)
        control_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(control_frame, text="Лист:").grid(row=0, column=0, sticky=tk.W)
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(control_frame, textvariable=self.sheet_var, width=25, state="readonly")
        self.sheet_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        self.sheet_combo.bind('<<ComboboxSelected>>', self.show_data)
        
        # Кнопки управления
        btn_frame = ttk.Frame(control_frame)
        btn_frame.grid(row=0, column=2, padx=(20, 0), sticky=tk.W)
        
        ttk.Button(btn_frame, text="Обновить данные", 
                  command=self.show_data, style="Action.TButton").pack(side=tk.LEFT, padx=2)
        
        ttk.Button(btn_frame, text="Удалить выбранные", 
                  command=self.delete_selected_entries, style="Danger.TButton").pack(side=tk.LEFT, padx=2)
        
        ttk.Button(btn_frame, text="Выбрать все", 
                  command=self.select_all_entries, style="Action.TButton").pack(side=tk.LEFT, padx=2)
        
        ttk.Button(btn_frame, text="Снять выделение", 
                  command=self.deselect_all_entries, style="Action.TButton").pack(side=tk.LEFT, padx=2)
        
        # Информация о выборе
        self.selection_info = ttk.Label(control_frame, text="Выбрано записей: 0", foreground="green", font=('Arial', 9, 'bold'))
        self.selection_info.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        # Таблица с увеличенной высотой
        columns = ("Число", "Дисциплина", "Группа", "Нагрузка", "Лекции", "Практические", "Лабораторные")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings", height=15, selectmode="extended")
        
        # Устанавливаем ширину колонок
        column_widths = {
            "Число": 80,
            "Дисциплина": 250,
            "Группа": 120,
            "Нагрузка": 100,
            "Лекции": 80,
            "Практические": 100,
            "Лабораторные": 100
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=column_widths.get(col, 100))
        
        # Размещаем таблицу
        self.tree.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # Привязываем обработчик выбора
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        
        # Вертикальная полоса прокрутки
        scrollbar_y = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_y.grid(row=2, column=2, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=scrollbar_y.set)
        
        # Горизонтальная полоса прокрутки
        scrollbar_x = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        scrollbar_x.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E))
        self.tree.configure(xscrollcommand=scrollbar_x.set)
        
        return frame

    def configure_grid(self, main_frame, columns_frame, view_frame):
        # Настройка растягивания для основного фрейма
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.columnconfigure(2, weight=1)
        main_frame.rowconfigure(4, weight=1)  # Строка с просмотром данных
        
        # Настройка растягивания для фрейма с колонками
        columns_frame.columnconfigure(0, weight=1)  # Управление датами
        columns_frame.columnconfigure(1, weight=1)  # Поля ввода
        columns_frame.columnconfigure(2, weight=0)  # GIF (фиксированная ширина)
        
        # Настройка растягивания для фрейма просмотра
        view_frame.columnconfigure(0, weight=1)
        view_frame.columnconfigure(1, weight=1)
        view_frame.rowconfigure(2, weight=1)
        
        # Настройка растягивания для главного окна
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    def on_tree_select(self, event):
        """Обновляет информацию о количестве выбранных записей"""
        selected_count = len(self.tree.selection())
        self.selection_info.config(text=f"Выбрано записей: {selected_count}")

    def select_all_entries(self):
        """Выбирает все записи в таблице"""
        all_items = self.tree.get_children()
        self.tree.selection_set(all_items)
        self.on_tree_select(None)

    def deselect_all_entries(self):
        """Снимает выделение со всех записей"""
        self.tree.selection_remove(self.tree.selection())
        self.on_tree_select(None)

    def delete_selected_entries(self):
        """Удаляет выбранные записи из таблицы и файла Excel"""
        if not self.wb:
            messagebox.showerror("Ошибка", "Файл не загружен")
            return
            
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Внимание", "Выберите записи для удаления")
            return
        
        entries_to_delete = []
        for item in selected_items:
            item_data = self.tree.item(item, 'values')
            if item_data:
                entries_to_delete.append(item_data)
        
        if not entries_to_delete:
            return
        
        confirm = messagebox.askyesno(
            "Подтверждение удаления", 
            f"Вы действительно хотите удалить {len(entries_to_delete)} записей?\n"
            f"Это действие нельзя отменить."
        )
        
        if not confirm:
            return
        
        try:
            # Закрываем и перезагружаем workbook для гарантии свежих данных
            self.close_workbook()
            self.wb = self.safe_load_workbook(self.filename)
            
            if not self.wb:
                return
            
            sheet_name = self.sheet_var.get()
            if sheet_name not in self.wb.sheetnames:
                messagebox.showerror("Ошибка", "Лист не найден")
                return
            
            sheet = self.wb[sheet_name]
            deleted_count = 0
            
            rows_to_delete = []
            for entry_data in entries_to_delete:
                day = int(entry_data[0])
                discipline = entry_data[1]
                group = entry_data[2]
                
                row = self.START_ROW
                while sheet[f'E{row}'].value is not None:
                    sheet_day = sheet[f'E{row}'].value
                    sheet_discipline = sheet[f'F{row}'].value or ''
                    sheet_group = sheet[f'G{row}'].value or ''
                    
                    if (isinstance(sheet_day, (int, float)) and int(sheet_day) == day and
                        sheet_discipline == discipline and sheet_group == group):
                        rows_to_delete.append(row)
                        break
                    row += 1
            
            rows_to_delete.sort(reverse=True)
            for row_num in rows_to_delete:
                self.delete_row(sheet, row_num)
                deleted_count += 1
            
            # Сохраняем и перезагружаем файл
            if self.safe_save_workbook():
                self.show_data()
                messagebox.showinfo("Успех", f"Удалено записей: {deleted_count} из {len(entries_to_delete)}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при удалении записей: {e}")

    def delete_row(self, sheet, row_num):
        """Удаляет строку из листа Excel и сдвигает остальные строки вверх"""
        max_row = self.START_ROW
        while sheet[f'E{max_row}'].value is not None:
            max_row += 1
        
        for row in range(row_num, max_row):
            sheet[f'E{row}'] = sheet[f'E{row + 1}'].value
            sheet[f'F{row}'] = sheet[f'F{row + 1}'].value
            sheet[f'G{row}'] = sheet[f'G{row + 1}'].value
            sheet[f'H{row}'] = sheet[f'H{row + 1}'].value
            sheet.cell(row=row, column=12).value = sheet.cell(row=row + 1, column=12).value
            sheet.cell(row=row, column=13).value = sheet.cell(row=row + 1, column=13).value
            sheet.cell(row=row, column=14).value = sheet.cell(row=row + 1, column=14).value
        
        # Очищаем последнюю строку (ВКЛЮЧАЯ КОЛОНКИ С ЧАСАМИ)
        sheet[f'E{max_row}'] = None
        sheet[f'F{max_row}'] = None
        sheet[f'G{max_row}'] = None
        sheet[f'H{max_row}'] = None
        sheet.cell(row=max_row, column=12).value = None
        sheet.cell(row=max_row, column=13).value = None
        sheet.cell(row=max_row, column=14).value = None

    def date_to_datetime(self, date_obj):
        return datetime.combine(date_obj, datetime.min.time()) if isinstance(date_obj, date) and not isinstance(date_obj, datetime) else date_obj

    def determine_week_type(self, input_date):
        current_date = self.date_to_datetime(input_date)
        year = current_date.year
        
        september_1 = datetime(year, 9, 1)
        if current_date < september_1:
            september_1 = datetime(year-1, 9, 1)
        
        days_diff = (current_date - september_1).days
        return "числитель" if (days_diff // 7) % 2 == 0 else "знаменатель"

    def find_sheet_for_month(self, month):
        if not self.wb:
            return None
            
        month_names = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",
                      7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        
        target_month = month_names.get(month)
        if not target_month:
            return None
        
        for sheet_name in self.wb.sheetnames:
            if re.search(r'\b' + re.escape(target_month) + r'\b', sheet_name):
                return sheet_name
        
        return None

    def generate_dates_by_period(self):
        try:
            start_dt = self.date_to_datetime(self.start_date.get_date())
            end_dt = self.date_to_datetime(self.end_date.get_date())
            target_week_type = self.period_week_type.get()
            
            if start_dt >= end_dt:
                messagebox.showerror("Ошибка", "Дата начала должна быть раньше даты окончания")
                return
            
            self.selected_dates.clear()
            generated_count = 0
            
            if target_week_type == "обе недели":
                current_date = start_dt
                while current_date <= end_dt:
                    sheet_name = self.find_sheet_for_month(current_date.month)
                    if sheet_name:
                        self.selected_dates.append({
                            'date': current_date, 'day': current_date.day, 'month': current_date.month,
                            'year': current_date.year, 'sheet': sheet_name,
                            'week_type': self.determine_week_type(current_date)
                        })
                        generated_count += 1
                    current_date += timedelta(days=7)
            else:
                current_date = start_dt
                while current_date <= end_dt:
                    week_type = self.determine_week_type(current_date)
                    if week_type == target_week_type:
                        sheet_name = self.find_sheet_for_month(current_date.month)
                        if sheet_name:
                            self.selected_dates.append({
                                'date': current_date, 'day': current_date.day, 'month': current_date.month,
                                'year': current_date.year, 'sheet': sheet_name, 'week_type': week_type
                            })
                            generated_count += 1
                        current_date += timedelta(days=7)
                    else:
                        current_date += timedelta(days=1)
            
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            self.update_dates_info()
            self.update_dates_display()
            
            if generated_count > 0:
                dates_list = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
                week_type_display = "числитель и знаменатель (каждую неделю)" if target_week_type == "обе недели" else target_week_type
                
                messagebox.showinfo("Успех", 
                    f"Сгенерировано {generated_count} дат\n"
                    f"Период: {start_dt.strftime('%d.%m.%Y')} - {end_dt.strftime('%d.%m.%Y')}\n"
                    f"Тип недели: {week_type_display}\n"
                    f"Даты: {dates_list}")
            else:
                messagebox.showwarning("Внимание", "В выбранном периоде нет дат")
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка генерации дат: {e}")

    def add_single_date(self):
        try:
            date_obj = self.date_to_datetime(self.single_date.get_date())
            sheet_name = self.find_sheet_for_month(date_obj.month)
            
            if not sheet_name:
                messagebox.showerror("Ошибка", f"Не найден лист для месяца {date_obj.month}")
                return
            
            for existing_date in self.selected_dates:
                if (existing_date['day'] == date_obj.day and 
                    existing_date['month'] == date_obj.month and 
                    existing_date['year'] == date_obj.year):
                    messagebox.showwarning("Внимание", "Эта дата уже есть в списке")
                    return
            
            date_info = {
                'date': date_obj, 
                'day': date_obj.day, 
                'month': date_obj.month,
                'year': date_obj.year, 
                'sheet': sheet_name,
                'week_type': self.determine_week_type(date_obj)
            }
            
            self.selected_dates.append(date_info)
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            
            self.update_dates_info()
            self.update_dates_display()
            
            messagebox.showinfo("Успех", f"Дата {date_obj.strftime('%d.%m.%Y')} добавлена")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка добавления даты: {e}")

    def remove_selected_date(self):
        selected_index = self.remove_date_combo.current()
        if selected_index == -1:
            messagebox.showwarning("Внимание", "Выберите дату для удаления")
            return
        
        if 0 <= selected_index < len(self.selected_dates):
            removed_date = self.selected_dates.pop(selected_index)
            self.update_dates_info()
            self.update_dates_display()
            messagebox.showinfo("Успех", f"Дата {removed_date['day']}.{removed_date['month']:02d}.{removed_date['year']} удалена")

    def update_dates_display(self):
        self.dates_listbox.delete(0, tk.END)
        for date_info in self.selected_dates:
            display_text = f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']} ({date_info['sheet']}, {date_info['week_type']})"
            self.dates_listbox.insert(tk.END, display_text)
        
        date_values = [f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']}" for date_info in self.selected_dates]
        self.remove_date_combo['values'] = date_values
        if date_values:
            self.remove_date_combo.set(date_values[0])
        else:
            self.remove_date_combo.set('')

    def clear_dates(self):
        self.selected_dates.clear()
        self.update_dates_info()
        self.update_dates_display()
        messagebox.showinfo("Успех", "Все даты очищены")

    def update_dates_info(self):
        count = len(self.selected_dates)
        if count > 0:
            dates_str = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
            sheets_count = {}
            for date in self.selected_dates:
                sheets_count[date['sheet']] = sheets_count.get(date['sheet'], 0) + 1
            
            sheets_info = ", ".join([f"{sheet}: {count}" for sheet, count in sheets_count.items()])
            week_types = set(date['week_type'] for date in self.selected_dates)
            week_types_info = f"Типы: {', '.join(week_types)}" if week_types else ""
            
            self.dates_info_label.config(text=f"Выбрано дат: {count} | Даты: {dates_str} | Листы: {sheets_info} {week_types_info}")
        else:
            self.dates_info_label.config(text="Выбрано дат: 0")

    def show_data(self, event=None):
        """Показывает данные из выбранного листа"""
        if not self.wb:
            return
        
        sheet_name = self.sheet_var.get()
        if sheet_name in self.wb.sheetnames:
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            try:
                sheet = self.wb[sheet_name]
                row = self.START_ROW
                
                while sheet[f'E{row}'].value is not None:
                    day = sheet[f'E{row}'].value
                    if isinstance(day, (int, float)):
                        self.tree.insert("", "end", values=(
                            int(day),
                            sheet[f'F{row}'].value or '',
                            sheet[f'G{row}'].value or '',
                            sheet[f'H{row}'].value or '',
                            sheet.cell(row=row, column=12).value or '',
                            sheet.cell(row=row, column=13).value or '',
                            sheet.cell(row=row, column=14).value or ''
                        ))
                    row += 1
                    
                self.selection_info.config(text="Выбрано записей: 0")
                
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при чтении данных: {e}")

    def add_entries(self):
        """Добавляет записи в журнал"""
        if not self.wb:
            messagebox.showerror("Ошибка", "Файл не загружен")
            return
            
        if not all([self.wb, self.selected_dates, self.entries['discipline'].get(), 
                   self.entries['group'].get(), self.entries['load_type'].get()]):
            messagebox.showerror("Ошибка", "Заполните все обязательные поля и сгенерируйте даты")
            return
        
        try:
            lecture = self.entries['lecture'].get()
            practice = self.entries['practice'].get()
            lab = self.entries['lab'].get()
            
            data = {
                'discipline': self.entries['discipline'].get(),
                'group': self.entries['group'].get(),
                'load_type': self.entries['load_type'].get(),
                'lecture': float(lecture) if lecture else 0.0,
                'practice': float(practice) if practice else 0.0,
                'lab': float(lab) if lab else 0.0
            }
            
            if data['lecture'] == 0 and data['practice'] == 0 and data['lab'] == 0:
                messagebox.showwarning("Внимание", "Заполните хотя бы одно поле: Лекции, Практические или Лабораторные")
                return
            
            # Закрываем и перезагружаем workbook для гарантии свежих данных
            self.close_workbook()
            self.wb = self.safe_load_workbook(self.filename)
            
            if not self.wb:
                return
            
            dates_by_sheet = {}
            for date_info in self.selected_dates:
                sheet_name = date_info['sheet']
                if sheet_name not in dates_by_sheet:
                    dates_by_sheet[sheet_name] = []
                dates_by_sheet[sheet_name].append(date_info)
            
            results = {}
            for sheet_name, dates in dates_by_sheet.items():
                if sheet_name not in self.wb.sheetnames:
                    continue
                    
                sheet = self.wb[sheet_name]
                dates.sort(key=lambda x: x['day'])
                
                added_rows = []
                for date_info in dates:
                    row = self.add_entry_to_sheet(sheet, date_info['day'], data)
                    if row:
                        added_rows.append(f"{date_info['day']}.{date_info['month']:02d}(стр.{row})")
                
                if added_rows:
                    results[sheet_name] = added_rows
            
            # Заполняем листы "Осень" и "Весна"
            season_results = self.fill_season_sheets(data)
            
            # Сохраняем файл
            if self.safe_save_workbook():
                # Обновляем отображение
                self.show_data()
                
                # Обновляем список дисциплин после добавления новых записей
                self.update_disciplines_list()
                
                msg_lines = ["Записи добавлены:"]
                for sheet_name, dates in results.items():
                    msg_lines.append(f"{sheet_name}: {', '.join(dates)}")
                
                if season_results:
                    msg_lines.append("\nСеместровые листы:")
                    for sheet_name, result in season_results.items():
                        msg_lines.append(f"{sheet_name}: {result}")
                
                if results or season_results:
                    messagebox.showinfo("Успех", "\n".join(msg_lines))
                    
                    # Очищаем поля ввода
                    for field in ['discipline', 'group', 'lecture', 'practice', 'lab']:
                        if field in self.entries:
                            if isinstance(self.entries[field], tk.Entry) or isinstance(self.entries[field], ttk.Entry):
                                self.entries[field].delete(0, tk.END)
                    
                    if 'load_type' in self.entries:
                        self.entries['load_type'].set(self.LOAD_TYPES[0])
                else:
                    messagebox.showwarning("Внимание", "Не удалось добавить записи")
                
        except ValueError as e:
            messagebox.showerror("Ошибка", "Проверьте числовые поля (Лекции, Практические, Лабораторные) - они должны содержать только числа")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при добавлении записей: {e}")

    def fill_season_sheets(self, data):
        """Заполняет листы 'Осень' и 'Весна' данными и возвращает результаты"""
        season_results = {}
        
        try:
            if not self.selected_dates:
                return season_results
            
            months = set(date_info['month'] for date_info in self.selected_dates)
            
            autumn_months = {9, 10, 11, 12}
            spring_months = {1, 2, 3, 4, 5}
            
            fill_autumn = any(month in autumn_months for month in months)
            fill_spring = any(month in spring_months for month in months)
            
            if fill_autumn and 'осень' in self.wb.sheetnames:
                result = self.fill_season_sheet('осень', data)
                if result:
                    season_results['осень'] = result
            
            if fill_spring and 'весна' in self.wb.sheetnames:
                result = self.fill_season_sheet('весна', data)
                if result:
                    season_results['весна'] = result
                    
        except Exception as e:
            print(f"Ошибка при заполнении семестровых листов: {e}")
        
        return season_results

    def fill_season_sheet(self, sheet_name, data):
        """Заполняет конкретный семестровый лист и возвращает результат"""
        try:
            sheet = self.wb[sheet_name]
            
            row = 5
            max_rows_to_check = 50
            
            while row <= max_rows_to_check:
                try:
                    cell_d = sheet[f'D{row}']
                    cell_e = sheet[f'E{row}']
                    cell_f = sheet[f'F{row}']
                    
                    def is_cell_empty(cell):
                        try:
                            return cell.value is None or str(cell.value).strip() == ''
                        except:
                            return True
                    
                    if (is_cell_empty(cell_d) and 
                        is_cell_empty(cell_e) and 
                        is_cell_empty(cell_f)):
                        break
                except:
                    pass
                
                row += 2
            
            if row > max_rows_to_check:
                return f"Не найдено свободных строк (проверено до строки {max_rows_to_check})"
            
            try:
                sheet[f'D{row}'] = data['discipline']
                sheet[f'E{row}'] = data['group']
                sheet[f'F{row}'] = data['load_type']
                
                for col in ['D', 'E', 'F']:
                    sheet[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                return f"строка {row}: {data['discipline']}, {data['group']}, {data['load_type']}"
                
            except Exception as write_error:
                return f"Ошибка записи в строку {row}: {str(write_error)}"
            
        except Exception as e:
            return f"Ошибка: {str(e)}"

    def add_entry_to_sheet(self, sheet, day, data):
        """Добавляет запись в лист и возвращает номер строки"""
        try:
            last_row = self.START_ROW
            while sheet[f'E{last_row}'].value is not None:
                last_row += 1
            
            existing_date_rows = []
            row = self.START_ROW
            while sheet[f'E{row}'].value is not None:
                existing_day = sheet[f'E{row}'].value
                if isinstance(existing_day, (int, float)) and int(existing_day) == day:
                    existing_date_rows.append(row)
                row += 1
            
            if existing_date_rows:
                last_date_row = max(existing_date_rows)
                insert_row = last_date_row + 1
                self.shift_rows_down(sheet, insert_row, last_row)
                self.fill_row_data(sheet, insert_row, day, data)
                return insert_row
            else:
                return self.insert_entry_sorted(sheet, day, data, last_row)
        except Exception as e:
            print(f"Ошибка при добавлении записи: {e}")
            return None

    def shift_rows_down(self, sheet, start_row, last_row):
        """Сдвигает строки вниз"""
        for row in range(last_row, start_row - 1, -1):
            sheet[f'E{row+1}'] = sheet[f'E{row}'].value
            sheet[f'F{row+1}'] = sheet[f'F{row}'].value
            sheet[f'G{row+1}'] = sheet[f'G{row}'].value
            sheet[f'H{row+1}'] = sheet[f'H{row}'].value
            sheet.cell(row=row+1, column=12).value = sheet.cell(row=row, column=12).value
            sheet.cell(row=row+1, column=13).value = sheet.cell(row=row, column=13).value
            sheet.cell(row=row+1, column=14).value = sheet.cell(row=row, column=14).value
            
            # ОЧИЩАЕМ ИСХОДНУЮ СТРОКУ ПРИ СДВИГЕ (если это строка вставки)
            if row == start_row:
                sheet.cell(row=row, column=12).value = None
                sheet.cell(row=row, column=13).value = None
                sheet.cell(row=row, column=14).value = None

    def insert_entry_sorted(self, sheet, day, data, last_row):
        """Вставляет запись в отсортированном порядке"""
        insert_row = self.START_ROW
        while sheet[f'E{insert_row}'].value is not None:
            existing_day = sheet[f'E{insert_row}'].value
            if isinstance(existing_day, (int, float)) and int(existing_day) > day:
                break
            insert_row += 1
        
        self.shift_rows_down(sheet, insert_row, last_row)
        self.fill_row_data(sheet, insert_row, day, data)
        return insert_row

    def fill_row_data(self, sheet, row, day, data):
        """Заполняет строку данными - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
        sheet[f'E{row}'] = day
        sheet[f'F{row}'] = data['discipline']
        sheet[f'G{row}'] = data['group']
        sheet[f'H{row}'] = data['load_type']
        
        # ОЧИЩАЕМ ВСЕ КОЛОНКИ С ЧАСАМИ ПЕРЕД ЗАПОЛНЕНИЕМ
        sheet.cell(row=row, column=12).value = None  # Лекции (J)
        sheet.cell(row=row, column=13).value = None  # Практические (K)
        sheet.cell(row=row, column=14).value = None  # Лабораторные (L)
        
        # ЗАПОЛНЯЕМ ТОЛЬКО ТЕ КОЛОНКИ, ГДЕ ЕСТЬ ЧАСЫ
        if data['lecture'] != 0:
            sheet.cell(row=row, column=12).value = data['lecture']
        if data['practice'] != 0:
            sheet.cell(row=row, column=13).value = data['practice']
        if data['lab'] != 0:
            sheet.cell(row=row, column=14).value = data['lab']

    def run(self):
        """Запускает приложение"""
        try:
            self.root.mainloop()
        except Exception as e:
            messagebox.showerror("Критическая ошибка", f"Приложение завершилось с ошибкой: {e}")
        finally:
            self.close_workbook()
            self.save_config()

def main():
    """Основная функция запуска приложения"""
    try:
        app = JournalApp()
        app.run()
    except Exception as e:
        print(f"Не удалось запустить приложение: {e}")
        messagebox.showerror("Ошибка запуска", f"Не удалось запустить приложение: {e}")

if __name__ == "__main__":
    main()