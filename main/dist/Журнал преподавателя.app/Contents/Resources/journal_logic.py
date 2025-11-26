import openpyxl
from PySide6.QtWidgets import QMessageBox, QFileDialog
from PySide6.QtCore import QDate
from datetime import datetime, timedelta, date
import re
import os
import time
import json
from openpyxl.styles import Alignment

class JournalLogic:
    def __init__(self):
        self.filename = None
        self.wb = None
        self.START_ROW = 7
        self.HOURS_COLS = {'lecture': 12, 'practice': 13, 'lab': 14}
        self.selected_dates = []
        self.LOAD_TYPES = ["осн.", "почас.", "совм."]
        self.config_file = "app_config.json"
        self.ui = None
        self.load_config()
    
    def set_ui(self, ui):
        """Устанавливает ссылку на UI"""
        self.ui = ui
    
    def load_config(self):
        """Загружает конфигурацию приложения"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.filename = config.get('last_file')
        except Exception as e:
            print(f"Ошибка загрузки конфигурации: {e}")
    
    def save_config(self):
        """Сохраняет конфигурацию приложения"""
        try:
            config = {
                'last_file': self.filename
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Ошибка сохранения конфигурации: {e}")
    
    def load_and_display_gif(self):
        """Загружает и отображает GIF изображение"""
        try:
            possible_paths = [
                "Без названия.gif",
                "animation.gif",
                os.path.join(os.path.dirname(__file__), "Без названия.gif"),
                os.path.join(os.path.dirname(__file__), "animation.gif"),
            ]
            
            gif_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    gif_path = path
                    break
            
            if gif_path and self.ui:
                self.ui.movie = QMovie(gif_path)
                self.ui.gif_label.setMovie(self.ui.movie)
                self.ui.movie.start()
            elif self.ui:
                self.ui.gif_label.setText("Журнал\nпреподавателя")
                
        except Exception as e:
            print(f"Ошибка загрузки GIF: {e}")
            if self.ui:
                self.ui.gif_label.setText("Журнал\nпреподавателя")
    
    def open_file(self):
        """Открывает диалог выбора файла"""
        filename, _ = QFileDialog.getOpenFileName(
            self.ui,
            "Выберите файл Excel",
            "",
            "Excel files (*.xlsx *.xls *.xltx);;All files (*.*)"
        )
        
        if filename:
            self.load_workbook(filename)
    
    def load_workbook(self, filename):
        """Загружает рабочую книгу Excel"""
        try:
            self.close_workbook()
            self.wb = self.safe_load_workbook(filename)
            
            if self.wb:
                self.filename = filename
                if self.ui:
                    self.ui.file_path_label.setText(filename)
                    self.ui.add_entries_btn.setEnabled(True)
                
                # Обновляем список листов - только месячные листы
                all_sheets = self.wb.sheetnames
                monthly_sheets = self.filter_monthly_sheets(all_sheets)
                if self.ui:
                    self.ui.sheet_combo.clear()
                    self.ui.sheet_combo.addItems(monthly_sheets)
                
                # Обновляем список дисциплин
                self.update_disciplines_list()
                
                # Сохраняем конфигурацию
                self.save_config()
                
                # Показываем данные
                self.show_data()
                
        except Exception as e:
            QMessageBox.critical(self.ui, "Ошибка", f"Ошибка загрузки файла: {e}")
    
    def update_disciplines_list(self):
        """Обновляет список дисциплин из листов 'Осень' и 'Весна'"""
        if not self.wb or not self.ui:
            return
        
        disciplines = set()
        
        # Получаем дисциплины из листа 'Осень'
        if 'осень' in self.wb.sheetnames:
            sheet = self.wb['осень']
            row = 5
            while row <= 100:
                try:
                    discipline = sheet[f'D{row}'].value
                    if discipline and isinstance(discipline, str) and discipline.strip():
                        disciplines.add(discipline.strip())
                except:
                    pass
                row += 2
        
        # Получаем дисциплины из листа 'Весна'
        if 'весна' in self.wb.sheetnames:
            sheet = self.wb['весна']
            row = 5
            while row <= 100:
                try:
                    discipline = sheet[f'D{row}'].value
                    if discipline and isinstance(discipline, str) and discipline.strip():
                        disciplines.add(discipline.strip())
                except:
                    pass
                row += 2
        
        # Сортируем дисциплины по алфавиту
        sorted_disciplines = sorted(list(disciplines))
        
        # Обновляем комбобокс дисциплин
        if 'discipline' in self.ui.entries:
            self.ui.entries['discipline'].clear()
            self.ui.entries['discipline'].addItems(sorted_disciplines)
    
    def filter_monthly_sheets(self, sheetnames):
        """Фильтрует листы, оставляя только месячные с номерами 01-12"""
        monthly_sheets = []
        for sheet in sheetnames:
            if re.match(r'^(09|10|11|12|01|02|03|04|05|06|07|08)', sheet):
                monthly_sheets.append(sheet)
        return monthly_sheets
    
    def show_instructions(self):
        """Показывает инструкцию по использованию"""
        instructions = """
ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ПРИЛОЖЕНИЯ "ЖУРНАЛ ПРЕПОДАВАТЕЛЯ"

1. ВЫБОР ФАЙЛА
   - Нажмите "Выбрать файл" для открытия существующего файла Excel

2. ВЫБОР ПЕРИОДА
   - Укажите начальную и конечную даты периода
   - Выберите тип недели: числитель, знаменатель или обе недели
   - Нажмите "Сгенерировать даты по периоду"

3. УПРАВЛЕНИЕ ДАТАМИ
   - Добавляйте отдельные даты через "Добавить дату"
   - Удаляйте даты через выпадающий список
   - Просматривайте выбранные даты в списке

4. ЗАПОЛНЕНИЕ ДАННЫХ
   - Введите название дисциплины (можно выбрать из списка или ввести новую)
   - Укажите группу
   - Выберите вид нагрузки
   - Заполните часы для лекций, практических и/или лабораторных работ

5. ДОБАВЛЕНИЕ ЗАПИСЕЙ
   - Нажмите "Добавить записи" для внесения данных в файл
   - Данные автоматически распределятся по соответствующим листам

6. ПРОСМОТР И УДАЛЕНИЕ
   - Выберите лист для просмотра данных
   - Выделите записи для удаления
   - Используйте кнопки управления для массовых операций

СОВЕТЫ:
- Всегда сохраняйте резервные копии важных файлов
- Проверяйте правильность введенных данных перед добавлением
- Используйте кнопку "Обновить данные" для актуальной информации
- Новые дисциплины автоматически добавляются в выпадающий список
        """
        if self.ui:
            self.ui.show_info_dialog("Инструкция", instructions)
    
    def show_about(self):
        """Показывает информацию о программе"""
        about_text = """
ПРИЛОЖЕНИЕ "ЖУРНАЛ ПРЕПОДАВАТЕЛЯ"

Назначение:
Автоматизация заполнения журналов учебной нагрузки преподавателей

Возможности:
- Автоматическое распределение данных по месяцам
- Заполнение семестровых ведомостей
- Удобное управление датами занятий
- Просмотр и редактирование существующих записей
- Автоматическое сохранение списка дисциплин

Для корректной работы требуется:
- Файл Excel с соответствующими листами
- Права на запись в выбранный файл

Разработчики: RAMD        """
        if self.ui:
            self.ui.show_info_dialog("О программе", about_text)
    
    def close_workbook(self):
        """Безопасное закрытие рабочей книги"""
        try:
            if hasattr(self, 'wb') and self.wb:
                self.wb.close()
                self.wb = None
        except Exception as e:
            print(f"Ошибка при закрытии файла: {e}")
    
    def safe_save_workbook(self):
        """Безопасное сохранение рабочей книги с повторными попытками"""
        if not self.wb:
            QMessageBox.critical(self.ui, "Ошибка", "Файл не загружен")
            return False
            
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if self.wb:
                    self.wb.save(self.filename)
                    return True
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
                else:
                    QMessageBox.critical(self.ui, "Ошибка", 
                        f"Нет доступа к файлу {self.filename}!\n"
                        f"Убедитесь, что файл не открыт в другой программе.")
                    return False
            except Exception as e:
                QMessageBox.critical(self.ui, "Ошибка", f"Ошибка сохранения файла: {e}")
                return False
        return False
    
    def safe_load_workbook(self, filename):
        """Безопасная загрузка рабочей книги с повторными попытками"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if os.path.exists(filename):
                    return openpyxl.load_workbook(filename)
                else:
                    QMessageBox.critical(self.ui, "Ошибка", f"Файл {filename} не найден")
                    return None
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
                else:
                    QMessageBox.critical(self.ui, "Ошибка", 
                        f"Нет доступа к файлу {filename}!\n"
                        f"Убедитесь, что файл не открыт в другой программе.")
                    return None
            except Exception as e:
                QMessageBox.critical(self.ui, "Ошибка", f"Ошибка загрузки файла: {e}")
                return None
        return None

    def date_to_datetime(self, date_obj):
        return datetime.combine(date_obj, datetime.min.time()) if isinstance(date_obj, date) and not isinstance(date_obj, datetime) else date_obj

    def determine_week_type(self, input_date):
        current_date = self.date_to_datetime(input_date)
        year = current_date.year
        
        september_1 = datetime(year, 9, 1)
        if current_date < september_1:
            september_1 = datetime(year-1, 9, 1)
        
        days_diff = (current_date - september_1).days
        return "числитель" if (days_diff // 7) % 2 == 0 else "знаменатель"

    def find_sheet_for_month(self, month):
        if not self.wb:
            return None
            
        month_names = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",
                      7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        
        target_month = month_names.get(month)
        if not target_month:
            return None
        
        for sheet_name in self.wb.sheetnames:
            if re.search(r'\b' + re.escape(target_month) + r'\b', sheet_name):
                return sheet_name
        
        return None

    def generate_dates_by_period(self):
        if not self.ui:
            return
            
        try:
            start_dt = self.ui.start_date.date().toPython()
            end_dt = self.ui.end_date.date().toPython()
            
            # Определяем тип недели
            if self.ui.numerator_radio.isChecked():
                target_week_type = "числитель"
            elif self.ui.denominator_radio.isChecked():
                target_week_type = "знаменатель"
            else:
                target_week_type = "обе недели"
            
            if start_dt >= end_dt:
                QMessageBox.critical(self.ui, "Ошибка", "Дата начала должна быть раньше даты окончания")
                return
            
            self.selected_dates.clear()
            generated_count = 0
            
            if target_week_type == "обе недели":
                current_date = start_dt
                while current_date <= end_dt:
                    sheet_name = self.find_sheet_for_month(current_date.month)
                    if sheet_name:
                        self.selected_dates.append({
                            'date': current_date, 'day': current_date.day, 'month': current_date.month,
                            'year': current_date.year, 'sheet': sheet_name,
                            'week_type': self.determine_week_type(current_date)
                        })
                        generated_count += 1
                    current_date += timedelta(days=7)
            else:
                current_date = start_dt
                while current_date <= end_dt:
                    week_type = self.determine_week_type(current_date)
                    if week_type == target_week_type:
                        sheet_name = self.find_sheet_for_month(current_date.month)
                        if sheet_name:
                            self.selected_dates.append({
                                'date': current_date, 'day': current_date.day, 'month': current_date.month,
                                'year': current_date.year, 'sheet': sheet_name, 'week_type': week_type
                            })
                            generated_count += 1
                        current_date += timedelta(days=7)
                    else:
                        current_date += timedelta(days=1)
            
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            self.update_dates_info()
            self.update_dates_display()
            
            if generated_count > 0:
                dates_list = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
                week_type_display = "числитель и знаменатель (каждую неделю)" if target_week_type == "обе недели" else target_week_type
                
                QMessageBox.information(self.ui, "Успех", 
                    f"Сгенерировано {generated_count} дат\n"
                    f"Период: {start_dt.strftime('%d.%m.%Y')} - {end_dt.strftime('%d.%m.%Y')}\n"
                    f"Тип недели: {week_type_display}\n"
                    f"Даты: {dates_list}")
            else:
                QMessageBox.warning(self.ui, "Внимание", "В выбранном периоде нет дат")
                
        except Exception as e:
            QMessageBox.critical(self.ui, "Ошибка", f"Ошибка генерации дат: {e}")

    def add_single_date(self):
        if not self.ui:
            return
            
        try:
            date_obj = self.ui.single_date.date().toPython()
            sheet_name = self.find_sheet_for_month(date_obj.month)
            
            if not sheet_name:
                QMessageBox.critical(self.ui, "Ошибка", f"Не найден лист для месяца {date_obj.month}")
                return
            
            for existing_date in self.selected_dates:
                if (existing_date['day'] == date_obj.day and 
                    existing_date['month'] == date_obj.month and 
                    existing_date['year'] == date_obj.year):
                    QMessageBox.warning(self.ui, "Внимание", "Эта дата уже есть в списке")
                    return
            
            date_info = {
                'date': date_obj, 
                'day': date_obj.day, 
                'month': date_obj.month,
                'year': date_obj.year, 
                'sheet': sheet_name,
                'week_type': self.determine_week_type(date_obj)
            }
            
            self.selected_dates.append(date_info)
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            
            self.update_dates_info()
            self.update_dates_display()
            
            QMessageBox.information(self.ui, "Успех", f"Дата {date_obj.strftime('%d.%m.%Y')} добавлена")
            
        except Exception as e:
            QMessageBox.critical(self.ui, "Ошибка", f"Ошибка добавления даты: {e}")

    def remove_selected_date(self):
        if not self.ui:
            return
            
        selected_index = self.ui.remove_date_combo.currentIndex()
        if selected_index == -1:
            QMessageBox.warning(self.ui, "Внимание", "Выберите дату для удаления")
            return
        
        if 0 <= selected_index < len(self.selected_dates):
            removed_date = self.selected_dates.pop(selected_index)
            self.update_dates_info()
            self.update_dates_display()
            QMessageBox.information(self.ui, "Успех", f"Дата {removed_date['day']}.{removed_date['month']:02d}.{removed_date['year']} удалена")

    def update_dates_display(self):
        if not self.ui:
            return
            
        self.ui.dates_listbox.clear()
        date_values = []
        for date_info in self.selected_dates:
            display_text = f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']} ({date_info['sheet']}, {date_info['week_type']})"
            self.ui.dates_listbox.addItem(display_text)
            date_values.append(f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']}")
        
        self.ui.remove_date_combo.clear()
        self.ui.remove_date_combo.addItems(date_values)

    def clear_dates(self):
        self.selected_dates.clear()
        self.update_dates_info()
        self.update_dates_display()
        if self.ui:
            QMessageBox.information(self.ui, "Успех", "Все даты очищены")

    def update_dates_info(self):
        if not self.ui:
            return
            
        count = len(self.selected_dates)
        if count > 0:
            dates_str = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
            sheets_count = {}
            for date in self.selected_dates:
                sheets_count[date['sheet']] = sheets_count.get(date['sheet'], 0) + 1
            
            sheets_info = ", ".join([f"{sheet}: {count}" for sheet, count in sheets_count.items()])
            week_types = set(date['week_type'] for date in self.selected_dates)
            week_types_info = f"Типы: {', '.join(week_types)}" if week_types else ""
            
            self.ui.dates_info_label.setText(f"Даты: {count} | {dates_str} | Листы: {sheets_info} {week_types_info}")
        else:
            self.ui.dates_info_label.setText("Даты: 0")

    def show_data(self):
        """Показывает данные из выбранного листа"""
        if not self.wb or not self.ui:
            return
        
        sheet_name = self.ui.sheet_combo.currentText()
        if sheet_name and sheet_name in self.wb.sheetnames:
            self.ui.table_widget.setRowCount(0)
            
            try:
                sheet = self.wb[sheet_name]
                row = self.START_ROW
                
                while sheet[f'E{row}'].value is not None:
                    day = sheet[f'E{row}'].value
                    if isinstance(day, (int, float)):
                        current_row = self.ui.table_widget.rowCount()
                        self.ui.table_widget.insertRow(current_row)
                        
                        self.ui.table_widget.setItem(current_row, 0, QTableWidgetItem(str(int(day))))
                        self.ui.table_widget.setItem(current_row, 1, QTableWidgetItem(str(sheet[f'F{row}'].value or '')))
                        self.ui.table_widget.setItem(current_row, 2, QTableWidgetItem(str(sheet[f'G{row}'].value or '')))
                        self.ui.table_widget.setItem(current_row, 3, QTableWidgetItem(str(sheet[f'H{row}'].value or '')))
                        self.ui.table_widget.setItem(current_row, 4, QTableWidgetItem(str(sheet.cell(row=row, column=12).value or '')))
                        self.ui.table_widget.setItem(current_row, 5, QTableWidgetItem(str(sheet.cell(row=row, column=13).value or '')))
                        self.ui.table_widget.setItem(current_row, 6, QTableWidgetItem(str(sheet.cell(row=row, column=14).value or '')))
                    
                    row += 1
                    
                self.update_selection_info()
                
            except Exception as e:
                QMessageBox.critical(self.ui, "Ошибка", f"Ошибка при чтении данных: {e}")

    def update_selection_info(self):
        """Обновляет информацию о количестве выбранных записей"""
        if not self.ui:
            return
            
        selected_count = len(self.ui.table_widget.selectedItems()) // self.ui.table_widget.columnCount()
        self.ui.selection_info.setText(f"Выбрано: {selected_count}")

    def select_all_entries(self):
        """Выбирает все записи в таблице"""
        if self.ui:
            self.ui.table_widget.selectAll()
            self.update_selection_info()

    def deselect_all_entries(self):
        """Снимает выделение со всех записей"""
        if self.ui:
            self.ui.table_widget.clearSelection()
            self.update_selection_info()

    def delete_selected_entries(self):
        """Удаляет выбранные записи из таблицы и файла Excel"""
        if not self.wb:
            QMessageBox.critical(self.ui, "Ошибка", "Файл не загружен")
            return
            
        selected_rows = set()
        for item in self.ui.table_widget.selectedItems():
            selected_rows.add(item.row())
        
        if not selected_rows:
            QMessageBox.warning(self.ui, "Внимание", "Выберите записи для удаления")
            return
        
        entries_to_delete = []
        for row in selected_rows:
            day = self.ui.table_widget.item(row, 0).text()
            discipline = self.ui.table_widget.item(row, 1).text()
            group = self.ui.table_widget.item(row, 2).text()
            entries_to_delete.append((day, discipline, group))
        
        confirm = QMessageBox.question(
            self.ui,
            "Подтверждение удаления", 
            f"Вы действительно хотите удалить {len(entries_to_delete)} записей?\n"
            f"Это действие нельзя отменить.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if confirm != QMessageBox.Yes:
            return
        
        try:
            # Закрываем и перезагружаем workbook для гарантии свежих данных
            self.close_workbook()
            self.wb = self.safe_load_workbook(self.filename)
            
            if not self.wb:
                return
            
            sheet_name = self.ui.sheet_combo.currentText()
            if sheet_name not in self.wb.sheetnames:
                QMessageBox.critical(self.ui, "Ошибка", "Лист не найден")
                return
            
            sheet = self.wb[sheet_name]
            deleted_count = 0
            
            rows_to_delete = []
            for entry_data in entries_to_delete:
                day = int(entry_data[0])
                discipline = entry_data[1]
                group = entry_data[2]
                
                row = self.START_ROW
                while sheet[f'E{row}'].value is not None:
                    sheet_day = sheet[f'E{row}'].value
                    sheet_discipline = sheet[f'F{row}'].value or ''
                    sheet_group = sheet[f'G{row}'].value or ''
                    
                    if (isinstance(sheet_day, (int, float)) and int(sheet_day) == day and
                        sheet_discipline == discipline and sheet_group == group):
                        rows_to_delete.append(row)
                        break
                    row += 1
            
            rows_to_delete.sort(reverse=True)
            for row_num in rows_to_delete:
                self.delete_row(sheet, row_num)
                deleted_count += 1
            
            # Сохраняем и перезагружаем файл
            if self.safe_save_workbook():
                self.show_data()
                QMessageBox.information(self.ui, "Успех", f"Удалено записей: {deleted_count} из {len(entries_to_delete)}")
            
        except Exception as e:
            QMessageBox.critical(self.ui, "Ошибка", f"Ошибка при удалении записей: {e}")

    def delete_row(self, sheet, row_num):
        """Удаляет строку из листа Excel и сдвигает остальные строки вверх"""
        max_row = self.START_ROW
        while sheet[f'E{max_row}'].value is not None:
            max_row += 1
        
        for row in range(row_num, max_row):
            sheet[f'E{row}'] = sheet[f'E{row + 1}'].value
            sheet[f'F{row}'] = sheet[f'F{row + 1}'].value
            sheet[f'G{row}'] = sheet[f'G{row + 1}'].value
            sheet[f'H{row}'] = sheet[f'H{row + 1}'].value
            sheet.cell(row=row, column=12).value = sheet.cell(row=row + 1, column=12).value
            sheet.cell(row=row, column=13).value = sheet.cell(row=row + 1, column=13).value
            sheet.cell(row=row, column=14).value = sheet.cell(row=row + 1, column=14).value
        
        # Очищаем последнюю строку (ВКЛЮЧАЯ КОЛОНКИ С ЧАСАМИ)
        sheet[f'E{max_row}'] = None
        sheet[f'F{max_row}'] = None
        sheet[f'G{max_row}'] = None
        sheet[f'H{max_row}'] = None
        sheet.cell(row=max_row, column=12).value = None
        sheet.cell(row=max_row, column=13).value = None
        sheet.cell(row=max_row, column=14).value = None

    def add_entries(self):
        """Добавляет записи в журнал"""
        if not self.wb or not self.ui:
            QMessageBox.critical(self.ui, "Ошибка", "Файл не загружен")
            return
            
        if not all([self.wb, self.selected_dates, 
                   self.ui.entries['discipline'].currentText() if hasattr(self.ui.entries['discipline'], 'currentText') else self.ui.entries['discipline'].text(),
                   self.ui.entries['group'].text(), 
                   self.ui.entries['load_type'].currentText() if hasattr(self.ui.entries['load_type'], 'currentText') else self.ui.entries['load_type'].text()]):
            QMessageBox.critical(self.ui, "Ошибка", "Заполните все обязательные поля и сгенерируйте даты")
            return
        
        try:
            discipline = self.ui.entries['discipline'].currentText() if hasattr(self.ui.entries['discipline'], 'currentText') else self.ui.entries['discipline'].text()
            group = self.ui.entries['group'].text()
            load_type = self.ui.entries['load_type'].currentText() if hasattr(self.ui.entries['load_type'], 'currentText') else self.ui.entries['load_type'].text()
            lecture = self.ui.entries['lecture'].text()
            practice = self.ui.entries['practice'].text()
            lab = self.ui.entries['lab'].text()
            
            data = {
                'discipline': discipline,
                'group': group,
                'load_type': load_type,
                'lecture': float(lecture) if lecture else 0.0,
                'practice': float(practice) if practice else 0.0,
                'lab': float(lab) if lab else 0.0
            }
            
            if data['lecture'] == 0 and data['practice'] == 0 and data['lab'] == 0:
                QMessageBox.warning(self.ui, "Внимание", "Заполните хотя бы одно поле: Лекции, Практические или Лабораторные")
                return
            
            # Закрываем и перезагружаем workbook для гарантии свежих данных
            self.close_workbook()
            self.wb = self.safe_load_workbook(self.filename)
            
            if not self.wb:
                return
            
            dates_by_sheet = {}
            for date_info in self.selected_dates:
                sheet_name = date_info['sheet']
                if sheet_name not in dates_by_sheet:
                    dates_by_sheet[sheet_name] = []
                dates_by_sheet[sheet_name].append(date_info)
            
            results = {}
            for sheet_name, dates in dates_by_sheet.items():
                if sheet_name not in self.wb.sheetnames:
                    continue
                    
                sheet = self.wb[sheet_name]
                dates.sort(key=lambda x: x['day'])
                
                added_rows = []
                for date_info in dates:
                    row = self.add_entry_to_sheet(sheet, date_info['day'], data)
                    if row:
                        added_rows.append(f"{date_info['day']}.{date_info['month']:02d}(стр.{row})")
                
                if added_rows:
                    results[sheet_name] = added_rows
            
            # Заполняем листы "Осень" и "Весна"
            season_results = self.fill_season_sheets(data)
            
            # Сохраняем файл
            if self.safe_save_workbook():
                # Обновляем отображение
                self.show_data()
                
                # Обновляем список дисциплин после добавления новых записей
                self.update_disciplines_list()
                
                msg_lines = ["Записи добавлены:"]
                for sheet_name, dates in results.items():
                    msg_lines.append(f"{sheet_name}: {', '.join(dates)}")
                
                if season_results:
                    msg_lines.append("\nСеместровые листы:")
                    for sheet_name, result in season_results.items():
                        msg_lines.append(f"{sheet_name}: {result}")
                
                if results or season_results:
                    QMessageBox.information(self.ui, "Успех", "\n".join(msg_lines))
                    
                    # Очищаем поля ввода
                    for field in ['discipline', 'group', 'lecture', 'practice', 'lab']:
                        if field in self.ui.entries:
                            if hasattr(self.ui.entries[field], 'clear'):
                                self.ui.entries[field].clear()
                    
                    if 'load_type' in self.ui.entries:
                        self.ui.entries['load_type'].setCurrentIndex(0)
                else:
                    QMessageBox.warning(self.ui, "Внимание", "Не удалось добавить записи")
                
        except ValueError as e:
            QMessageBox.critical(self.ui, "Ошибка", "Проверьте числовые поля (Лекции, Практические, Лабораторные) - они должны содержать только числа")
        except Exception as e:
            QMessageBox.critical(self.ui, "Ошибка", f"Ошибка при добавлении записей: {e}")

    def fill_season_sheets(self, data):
        """Заполняет листы 'Осень' и 'Весна' данными и возвращает результаты"""
        season_results = {}
        
        try:
            if not self.selected_dates:
                return season_results
            
            months = set(date_info['month'] for date_info in self.selected_dates)
            
            autumn_months = {9, 10, 11, 12}
            spring_months = {1, 2, 3, 4, 5}
            
            fill_autumn = any(month in autumn_months for month in months)
            fill_spring = any(month in spring_months for month in months)
            
            if fill_autumn and 'осень' in self.wb.sheetnames:
                result = self.fill_season_sheet('осень', data)
                if result:
                    season_results['осень'] = result
            
            if fill_spring and 'весна' in self.wb.sheetnames:
                result = self.fill_season_sheet('весна', data)
                if result:
                    season_results['весна'] = result
                    
        except Exception as e:
            print(f"Ошибка при заполнении семестровых листов: {e}")
        
        return season_results

    def fill_season_sheet(self, sheet_name, data):
        """Заполняет конкретный семестровый лист и возвращает результат"""
        try:
            sheet = self.wb[sheet_name]
            
            row = 5
            max_rows_to_check = 50
            
            while row <= max_rows_to_check:
                try:
                    cell_d = sheet[f'D{row}']
                    cell_e = sheet[f'E{row}']
                    cell_f = sheet[f'F{row}']
                    
                    def is_cell_empty(cell):
                        try:
                            return cell.value is None or str(cell.value).strip() == ''
                        except:
                            return True
                    
                    if (is_cell_empty(cell_d) and 
                        is_cell_empty(cell_e) and 
                        is_cell_empty(cell_f)):
                        break
                except:
                    pass
                
                row += 2
            
            if row > max_rows_to_check:
                return f"Не найдено свободных строк (проверено до строки {max_rows_to_check})"
            
            try:
                sheet[f'D{row}'] = data['discipline']
                sheet[f'E{row}'] = data['group']
                sheet[f'F{row}'] = data['load_type']
                
                for col in ['D', 'E', 'F']:
                    sheet[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                return f"строка {row}: {data['discipline']}, {data['group']}, {data['load_type']}"
                
            except Exception as write_error:
                return f"Ошибка записи в строку {row}: {str(write_error)}"
            
        except Exception as e:
            return f"Ошибка: {str(e)}"

    def add_entry_to_sheet(self, sheet, day, data):
        """Добавляет запись в лист и возвращает номер строки"""
        try:
            last_row = self.START_ROW
            while sheet[f'E{last_row}'].value is not None:
                last_row += 1
            
            existing_date_rows = []
            row = self.START_ROW
            while sheet[f'E{row}'].value is not None:
                existing_day = sheet[f'E{row}'].value
                if isinstance(existing_day, (int, float)) and int(existing_day) == day:
                    existing_date_rows.append(row)
                row += 1
            
            if existing_date_rows:
                last_date_row = max(existing_date_rows)
                insert_row = last_date_row + 1
                self.shift_rows_down(sheet, insert_row, last_row)
                self.fill_row_data(sheet, insert_row, day, data)
                return insert_row
            else:
                return self.insert_entry_sorted(sheet, day, data, last_row)
        except Exception as e:
            print(f"Ошибка при добавлении записи: {e}")
            return None

    def shift_rows_down(self, sheet, start_row, last_row):
        """Сдвигает строки вниз"""
        for row in range(last_row, start_row - 1, -1):
            sheet[f'E{row+1}'] = sheet[f'E{row}'].value
            sheet[f'F{row+1}'] = sheet[f'F{row}'].value
            sheet[f'G{row+1}'] = sheet[f'G{row}'].value
            sheet[f'H{row+1}'] = sheet[f'H{row}'].value
            sheet.cell(row=row+1, column=12).value = sheet.cell(row=row, column=12).value
            sheet.cell(row=row+1, column=13).value = sheet.cell(row=row, column=13).value
            sheet.cell(row=row+1, column=14).value = sheet.cell(row=row, column=14).value
            
            # ОЧИЩАЕМ ИСХОДНУЮ СТРОКУ ПРИ СДВИГЕ (если это строка вставки)
            if row == start_row:
                sheet.cell(row=row, column=12).value = None
                sheet.cell(row=row, column=13).value = None
                sheet.cell(row=row, column=14).value = None

    def insert_entry_sorted(self, sheet, day, data, last_row):
        """Вставляет запись в отсортированном порядке"""
        insert_row = self.START_ROW
        while sheet[f'E{insert_row}'].value is not None:
            existing_day = sheet[f'E{insert_row}'].value
            if isinstance(existing_day, (int, float)) and int(existing_day) > day:
                break
            insert_row += 1
        
        self.shift_rows_down(sheet, insert_row, last_row)
        self.fill_row_data(sheet, insert_row, day, data)
        return insert_row

    def fill_row_data(self, sheet, row, day, data):
        """Заполняет строку данными - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
        sheet[f'E{row}'] = day
        sheet[f'F{row}'] = data['discipline']
        sheet[f'G{row}'] = data['group']
        sheet[f'H{row}'] = data['load_type']
        
        # ОЧИЩАЕМ ВСЕ КОЛОНКИ С ЧАСАМИ ПЕРЕД ЗАПОЛНЕНИЕМ
        sheet.cell(row=row, column=12).value = None  # Лекции (J)
        sheet.cell(row=row, column=13).value = None  # Практические (K)
        sheet.cell(row=row, column=14).value = None  # Лабораторные (L)
        
        # ЗАПОЛНЯЕМ ТОЛЬКО ТЕ КОЛОНКИ, ГДЕ ЕСТЬ ЧАСЫ
        if data['lecture'] != 0:
            sheet.cell(row=row, column=12).value = data['lecture']
        if data['practice'] != 0:
            sheet.cell(row=row, column=13).value = data['practice']
        if data['lab'] != 0:
            sheet.cell(row=row, column=14).value = data['lab']