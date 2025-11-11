import sys
import os
import openpyxl
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
    QWidget, QLabel, QLineEdit, QComboBox, QPushButton, 
    QTableWidget, QTableWidgetItem, QDateEdit, QListWidget,
    QMessageBox, QGroupBox, QHeaderView,
    QAbstractItemView, QFileDialog
)
from PySide6.QtCore import Qt, QDate
from datetime import datetime, timedelta
import re

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

class JournalApp(QMainWindow):
    def __init__(self, filename=None):
        super().__init__()
        self.filename = filename
        self.wb = None
        self.START_ROW = 7
        self.HOURS_COLS = {'lecture': 12, 'practice': 13, 'lab': 14}
        self.selected_dates = []
        self.LOAD_TYPES = ["осн.", "почас.", "совм."]
        
        self.setup_ui()
        if self.filename and os.path.exists(self.filename):
            self.load_workbook()
        else:
            self.ask_for_journal_file()
    
    def ask_for_journal_file(self):
        """Ask user to select journal file if not found"""
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл журнала",
            "",
            "Excel Files (*.xlsx *.xls *.xltx);;All Files (*)"
        )
        
        if filename:
            self.filename = filename
            self.load_workbook()
        else:
            QMessageBox.warning(self, "Внимание", 
                              "Файл журнала не найден.\nПожалуйста, выберите файл 'Тетрадь_ППС_2025_2026_каф_NN_Фамилия_ИО_оч_заоч.xltx'")
    
    def setup_ui(self):
        self.setWindowTitle("Журнал преподавателя")
        self.setGeometry(100, 100, 1200, 900)
        
        # Основной виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Основной layout
        main_layout = QVBoxLayout(central_widget)
        
        # Информация о файле
        self.file_info_label = QLabel("Файл журнала: не загружен")
        self.file_info_label.setStyleSheet("color: red; font-weight: bold; padding: 5px;")
        main_layout.addWidget(self.file_info_label)
        
        # Кнопка выбора файла
        self.select_file_btn = QPushButton("Выбрать файл журнала")
        self.select_file_btn.clicked.connect(self.ask_for_journal_file)
        self.select_file_btn.setStyleSheet("QPushButton { background-color: #2196F3; color: white; padding: 5px; }")
        main_layout.addWidget(self.select_file_btn)
        
        # Период
        period_group = self.create_period_group()
        main_layout.addWidget(period_group)
        
        # Даты
        dates_group = self.create_dates_group()
        main_layout.addWidget(dates_group)
        
        # Поля ввода
        input_group = self.create_input_group()
        main_layout.addWidget(input_group)
        
        # Кнопка добавления
        self.add_btn = QPushButton("Добавить записи")
        self.add_btn.clicked.connect(self.add_entries)
        self.add_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }")
        main_layout.addWidget(self.add_btn)
        
        # Просмотр данных
        view_group = self.create_view_group()
        main_layout.addWidget(view_group, 1)
    
    def create_period_group(self):
        group = QGroupBox("Выбор периода")
        layout = QVBoxLayout()
        
        # Первая строка - даты и тип недели
        first_row = QHBoxLayout()
        
        # Начало периода
        first_row.addWidget(QLabel("Начало периода:"))
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate(datetime.now().year, 9, 1))
        self.start_date.setDisplayFormat("dd.MM.yyyy")
        first_row.addWidget(self.start_date)
        
        # Конец периода
        first_row.addWidget(QLabel("Конец периода:"))
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate(datetime.now().year, 12, 31))
        self.end_date.setDisplayFormat("dd.MM.yyyy")
        first_row.addWidget(self.end_date)
        
        # Тип недели
        first_row.addWidget(QLabel("Тип недели:"))
        self.week_type_combo = QComboBox()
        self.week_type_combo.addItems(["числитель", "знаменатель", "обе недели"])
        first_row.addWidget(self.week_type_combo)
        
        first_row.addStretch()
        layout.addLayout(first_row)
        
        # Вторая строка - кнопки и информация
        second_row = QHBoxLayout()
        
        # Кнопки
        self.gen_dates_btn = QPushButton("Сгенерировать даты по периоду")
        self.gen_dates_btn.clicked.connect(self.generate_dates_by_period)
        second_row.addWidget(self.gen_dates_btn)
        
        self.clear_dates_btn = QPushButton("Очистить даты")
        self.clear_dates_btn.clicked.connect(self.clear_dates)
        second_row.addWidget(self.clear_dates_btn)
        
        second_row.addStretch()
        
        # Информация о датах
        self.dates_info_label = QLabel("Выбрано дат: 0")
        self.dates_info_label.setStyleSheet("color: blue; font-weight: bold;")
        second_row.addWidget(self.dates_info_label)
        
        layout.addLayout(second_row)
        group.setLayout(layout)
        return group
    
    def create_dates_group(self):
        group = QGroupBox("Управление датами")
        layout = QVBoxLayout()
        
        # Добавление даты
        add_layout = QHBoxLayout()
        add_layout.addWidget(QLabel("Добавить дату:"))
        self.single_date = QDateEdit()
        self.single_date.setDate(QDate.currentDate())
        self.single_date.setDisplayFormat("dd.MM.yyyy")
        add_layout.addWidget(self.single_date)
        
        self.add_date_btn = QPushButton("Добавить дату")
        self.add_date_btn.clicked.connect(self.add_single_date)
        add_layout.addWidget(self.add_date_btn)
        add_layout.addStretch()
        layout.addLayout(add_layout)
        
        # Удаление даты
        remove_layout = QHBoxLayout()
        remove_layout.addWidget(QLabel("Удалить дату:"))
        self.remove_date_combo = QComboBox()
        remove_layout.addWidget(self.remove_date_combo)
        
        self.remove_date_btn = QPushButton("Удалить дату")
        self.remove_date_btn.clicked.connect(self.remove_selected_date)
        remove_layout.addWidget(self.remove_date_btn)
        remove_layout.addStretch()
        layout.addLayout(remove_layout)
        
        # Список дат
        layout.addWidget(QLabel("Выбранные даты:"))
        self.dates_list = QListWidget()
        layout.addWidget(self.dates_list)
        
        # Кнопки управления списком
        list_buttons_layout = QHBoxLayout()
        self.clear_all_btn = QPushButton("Очистить все даты")
        self.clear_all_btn.clicked.connect(self.clear_dates)
        list_buttons_layout.addWidget(self.clear_all_btn)
        
        self.refresh_list_btn = QPushButton("Обновить список")
        self.refresh_list_btn.clicked.connect(self.update_dates_display)
        list_buttons_layout.addWidget(self.refresh_list_btn)
        list_buttons_layout.addStretch()
        layout.addLayout(list_buttons_layout)
        
        group.setLayout(layout)
        return group
    
    def create_input_group(self):
        group = QGroupBox("Данные для ввода")
        layout = QVBoxLayout()
        
        # Дисциплина
        disc_layout = QHBoxLayout()
        disc_layout.addWidget(QLabel("Дисциплина:"))
        self.discipline_edit = QLineEdit()
        disc_layout.addWidget(self.discipline_edit, 1)
        layout.addLayout(disc_layout)
        
        # Группа
        group_layout = QHBoxLayout()
        group_layout.addWidget(QLabel("Группа:"))
        self.group_edit = QLineEdit()
        group_layout.addWidget(self.group_edit, 1)
        layout.addLayout(group_layout)
        
        # Тип нагрузки
        load_layout = QHBoxLayout()
        load_layout.addWidget(QLabel("Вид нагрузки:"))
        self.load_type_combo = QComboBox()
        self.load_type_combo.addItems(self.LOAD_TYPES)
        load_layout.addWidget(self.load_type_combo)
        layout.addLayout(load_layout)
        
        # Часы
        hours_layout = QHBoxLayout()
        hours_layout.addWidget(QLabel("Лекции:"))
        self.lecture_edit = QLineEdit()
        self.lecture_edit.setPlaceholderText("0")
        hours_layout.addWidget(self.lecture_edit)
        
        hours_layout.addWidget(QLabel("Практические:"))
        self.practice_edit = QLineEdit()
        self.practice_edit.setPlaceholderText("0")
        hours_layout.addWidget(self.practice_edit)
        
        hours_layout.addWidget(QLabel("Лабораторные:"))
        self.lab_edit = QLineEdit()
        self.lab_edit.setPlaceholderText("0")
        hours_layout.addWidget(self.lab_edit)
        layout.addLayout(hours_layout)
        
        group.setLayout(layout)
        return group
    
    def create_view_group(self):
        group = QGroupBox("Просмотр данных")
        layout = QVBoxLayout()
        
        # Панель управления
        control_layout = QHBoxLayout()
        
        # Выбор листа
        control_layout.addWidget(QLabel("Лист:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.currentTextChanged.connect(self.show_data)
        control_layout.addWidget(self.sheet_combo)
        
        # Кнопки управления
        self.delete_btn = QPushButton("Удалить выбранные записи")
        self.delete_btn.clicked.connect(self.delete_selected_entries)
        self.delete_btn.setStyleSheet("QPushButton { background-color: #ff6b6b; color: black; }")
        control_layout.addWidget(self.delete_btn)
        
        self.select_all_btn = QPushButton("Выбрать все")
        self.select_all_btn.clicked.connect(self.select_all_entries)
        control_layout.addWidget(self.select_all_btn)
        
        self.deselect_all_btn = QPushButton("Снять выделение")
        self.deselect_all_btn.clicked.connect(self.deselect_all_entries)
        control_layout.addWidget(self.deselect_all_btn)
        
        control_layout.addStretch()
        layout.addLayout(control_layout)
        
        # Информация о выборе
        self.selection_info = QLabel("Выбрано записей: 0")
        self.selection_info.setStyleSheet("color: green; font-weight: bold;")
        layout.addWidget(self.selection_info)
        
        # Таблица
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        headers = ["Число", "Дисциплина", "Группа", "Нагрузка", "Лекции", "Практические", "Лабораторные"]
        self.table.setHorizontalHeaderLabels(headers)
        
        # Настройка таблицы
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # Автоматическое растягивание колонок
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        
        # Обработчик выбора
        self.table.itemSelectionChanged.connect(self.on_table_selection_changed)
        
        layout.addWidget(self.table, 1)
        group.setLayout(layout)
        return group

    # МЕТОДЫ ЛОГИКИ
    def date_to_datetime(self, date_obj):
        if isinstance(date_obj, QDate):
            return datetime(date_obj.year(), date_obj.month(), date_obj.day())
        return date_obj

    def determine_week_type(self, input_date):
        current_date = self.date_to_datetime(input_date)
        year = current_date.year
        
        september_1 = datetime(year, 9, 1)
        if current_date < september_1:
            september_1 = datetime(year-1, 9, 1)
        
        days_diff = (current_date - september_1).days
        return "числитель" if (days_diff // 7) % 2 == 0 else "знаменатель"

    def find_sheet_for_month(self, month):
        month_names = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",
                      7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        
        target_month = month_names.get(month)
        if not target_month:
            return None
        
        for sheet_name in self.wb.sheetnames:
            if re.search(r'\b' + re.escape(target_month) + r'\b', sheet_name):
                return sheet_name
        
        return None

    def generate_dates_by_period(self):
        try:
            start_dt = self.date_to_datetime(self.start_date.date())
            end_dt = self.date_to_datetime(self.end_date.date())
            target_week_type = self.week_type_combo.currentText()
            
            if start_dt >= end_dt:
                QMessageBox.critical(self, "Ошибка", "Дата начала должна быть раньше даты окончания")
                return
            
            self.selected_dates.clear()
            generated_count = 0
            
            if target_week_type == "обе недели":
                current_date = start_dt
                while current_date <= end_dt:
                    sheet_name = self.find_sheet_for_month(current_date.month)
                    if sheet_name:
                        self.selected_dates.append({
                            'date': current_date, 'day': current_date.day, 'month': current_date.month,
                            'year': current_date.year, 'sheet': sheet_name,
                            'week_type': self.determine_week_type(current_date)
                        })
                        generated_count += 1
                    current_date += timedelta(days=7)
            else:
                current_date = start_dt
                while current_date <= end_dt:
                    week_type = self.determine_week_type(current_date)
                    if week_type == target_week_type:
                        sheet_name = self.find_sheet_for_month(current_date.month)
                        if sheet_name:
                            self.selected_dates.append({
                                'date': current_date, 'day': current_date.day, 'month': current_date.month,
                                'year': current_date.year, 'sheet': sheet_name, 'week_type': week_type
                            })
                            generated_count += 1
                        current_date += timedelta(days=7)
                    else:
                        current_date += timedelta(days=1)
            
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            self.update_dates_info()
            self.update_dates_display()
            
            if generated_count > 0:
                dates_list = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
                week_type_display = "числитель и знаменатель (каждую неделю)" if target_week_type == "обе недели" else target_week_type
                
                QMessageBox.information(self, "Успех", 
                    f"Сгенерировано {generated_count} дат\n"
                    f"Период: {start_dt.strftime('%d.%m.%Y')} - {end_dt.strftime('%d.%m.%Y')}\n"
                    f"Тип недели: {week_type_display}\n"
                    f"Даты: {dates_list}")
            else:
                QMessageBox.warning(self, "Внимание", "В выбранном периоде нет дат")
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка генерации дат: {e}")

    def add_single_date(self):
        try:
            date_obj = self.date_to_datetime(self.single_date.date())
            sheet_name = self.find_sheet_for_month(date_obj.month)
            
            if not sheet_name:
                QMessageBox.critical(self, "Ошибка", f"Не найден лист для месяца {date_obj.month}")
                return
            
            # Проверяем, нет ли уже этой даты в списке
            for existing_date in self.selected_dates:
                if (existing_date['day'] == date_obj.day and 
                    existing_date['month'] == date_obj.month and 
                    existing_date['year'] == date_obj.year):
                    QMessageBox.warning(self, "Внимание", "Эта дата уже есть в списке")
                    return
            
            date_info = {
                'date': date_obj, 
                'day': date_obj.day, 
                'month': date_obj.month,
                'year': date_obj.year, 
                'sheet': sheet_name,
                'week_type': self.determine_week_type(date_obj)
            }
            
            self.selected_dates.append(date_info)
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            
            self.update_dates_info()
            self.update_dates_display()
            
            QMessageBox.information(self, "Успех", f"Дата {date_obj.strftime('%d.%m.%Y')} добавлена")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка добавления даты: {e}")

    def remove_selected_date(self):
        selected_index = self.remove_date_combo.currentIndex()
        if selected_index == -1:
            QMessageBox.warning(self, "Внимание", "Выберите дату для удаления")
            return
        
        if 0 <= selected_index < len(self.selected_dates):
            removed_date = self.selected_dates.pop(selected_index)
            self.update_dates_info()
            self.update_dates_display()
            QMessageBox.information(self, "Успех", f"Дата {removed_date['day']}.{removed_date['month']:02d}.{removed_date['year']} удалена")

    def clear_dates(self):
        self.selected_dates.clear()
        self.update_dates_info()
        self.update_dates_display()
        QMessageBox.information(self, "Успех", "Все даты очищены")

    def update_dates_display(self):
        # Обновляем список дат в ListWidget
        self.dates_list.clear()
        for date_info in self.selected_dates:
            display_text = f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']} ({date_info['sheet']}, {date_info['week_type']})"
            self.dates_list.addItem(display_text)
        
        # Обновляем комбобокс для удаления
        date_values = [f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']}" for date_info in self.selected_dates]
        self.remove_date_combo.clear()
        self.remove_date_combo.addItems(date_values)

    def update_dates_info(self):
        count = len(self.selected_dates)
        if count > 0:
            dates_str = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
            sheets_count = {}
            for date in self.selected_dates:
                sheets_count[date['sheet']] = sheets_count.get(date['sheet'], 0) + 1
            
            sheets_info = ", ".join([f"{sheet}: {count}" for sheet, count in sheets_count.items()])
            week_types = set(date['week_type'] for date in self.selected_dates)
            week_types_info = f"Типы: {', '.join(week_types)}" if week_types else ""
            
            self.dates_info_label.setText(f"Выбрано дат: {count} | Даты: {dates_str} | Листы: {sheets_info} {week_types_info}")
        else:
            self.dates_info_label.setText("Выбрано дат: 0")

    def load_workbook(self):
        try:
            self.wb = openpyxl.load_workbook(self.filename)
            sheets = self.wb.sheetnames
            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheets)
            if sheets:
                self.sheet_combo.setCurrentIndex(0)
            
            # Обновляем информацию о файле
            self.file_info_label.setText(f"Файл журнала: {os.path.basename(self.filename)}")
            self.file_info_label.setStyleSheet("color: green; font-weight: bold; padding: 5px;")
            
            self.show_data()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки: {e}")
            self.file_info_label.setText("Файл журнала: ошибка загрузки")
            self.file_info_label.setStyleSheet("color: red; font-weight: bold; padding: 5px;")

    def show_data(self):
        if not self.wb:
            return
        
        sheet_name = self.sheet_combo.currentText()
        if sheet_name and sheet_name in self.wb.sheetnames:
            self.table.setRowCount(0)
            
            try:
                sheet = self.wb[sheet_name]
                row = self.START_ROW
                data_rows = []
                
                while sheet[f'E{row}'].value is not None:
                    day = sheet[f'E{row}'].value
                    if isinstance(day, (int, float)):
                        data_rows.append((
                            int(day),
                            sheet[f'F{row}'].value or '',
                            sheet[f'G{row}'].value or '',
                            sheet[f'H{row}'].value or '',
                            sheet.cell(row=row, column=12).value or '',
                            sheet.cell(row=row, column=13).value or '',
                            sheet.cell(row=row, column=14).value or ''
                        ))
                    row += 1
                
                # Заполняем таблицу
                self.table.setRowCount(len(data_rows))
                for row_idx, row_data in enumerate(data_rows):
                    for col_idx, value in enumerate(row_data):
                        item = QTableWidgetItem(str(value))
                        self.table.setItem(row_idx, col_idx, item)
                
                # Сбрасываем счетчик выбранных записей
                self.selection_info.setText("Выбрано записей: 0")
                
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка данных: {e}")

    def on_table_selection_changed(self):
        """Обновляет информацию о количестве выбранных записей"""
        selected_count = len(self.table.selectedItems()) // self.table.columnCount()
        self.selection_info.setText(f"Выбрано записей: {selected_count}")

    def select_all_entries(self):
        """Выбирает все записи в таблице"""
        self.table.selectAll()

    def deselect_all_entries(self):
        """Снимает выделение со всех записей"""
        self.table.clearSelection()

    def delete_selected_entries(self):
        """Удаляет выбранные записи из таблицы и файла Excel"""
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            QMessageBox.warning(self, "Внимание", "Выберите записи для удаления")
            return
        
        # Получаем индексы выбранных строк
        rows_to_delete = set()
        for range in selected_ranges:
            for row in range(range.topRow(), range.bottomRow() + 1):
                rows_to_delete.add(row)
        
        if not rows_to_delete:
            return
        
        # Подтверждение удаления
        confirm = QMessageBox.question(
            self, 
            "Подтверждение удаления", 
            f"Вы действительно хотите удалить {len(rows_to_delete)} записей?\nЭто действие нельзя отменить.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if confirm != QMessageBox.Yes:
            return
        
        try:
            sheet_name = self.sheet_combo.currentText()
            if sheet_name not in self.wb.sheetnames:
                QMessageBox.critical(self, "Ошибка", "Лист не найден")
                return
            
            sheet = self.wb[sheet_name]
            deleted_count = 0
            
            # Собираем данные выбранных строк для поиска в Excel
            entries_to_delete = []
            for row in sorted(rows_to_delete, reverse=True):
                if row < self.table.rowCount():
                    day_item = self.table.item(row, 0)
                    discipline_item = self.table.item(row, 1)
                    group_item = self.table.item(row, 2)
                    
                    if day_item and discipline_item and group_item:
                        entries_to_delete.append({
                            'row': row,
                            'day': int(day_item.text()),
                            'discipline': discipline_item.text(),
                            'group': group_item.text()
                        })
            
            # Удаляем из Excel
            for entry in entries_to_delete:
                excel_row = self.START_ROW
                found = False
                
                while sheet[f'E{excel_row}'].value is not None and not found:
                    sheet_day = sheet[f'E{excel_row}'].value
                    sheet_discipline = sheet[f'F{excel_row}'].value or ''
                    sheet_group = sheet[f'G{excel_row}'].value or ''
                    
                    if (isinstance(sheet_day, (int, float)) and int(sheet_day) == entry['day'] and
                        sheet_discipline == entry['discipline'] and sheet_group == entry['group']):
                        
                        self.delete_excel_row(sheet, excel_row)
                        deleted_count += 1
                        found = True
                    
                    excel_row += 1
            
            # Сохраняем файл
            self.wb.save(self.filename)
            
            # Обновляем отображение
            self.show_data()
            
            QMessageBox.information(self, "Успех", f"Удалено записей: {deleted_count}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении записей: {e}")

    def delete_excel_row(self, sheet, row_num):
        """Удаляет строку из листа Excel и сдвигает остальные строки вверх"""
        max_row = self.START_ROW
        while sheet[f'E{max_row}'].value is not None:
            max_row += 1
        
        # Сдвигаем строки вверх
        for row in range(row_num, max_row):
            # Копируем значения из следующей строки
            sheet[f'E{row}'] = sheet[f'E{row + 1}'].value
            sheet[f'F{row}'] = sheet[f'F{row + 1}'].value
            sheet[f'G{row}'] = sheet[f'G{row + 1}'].value
            sheet[f'H{row}'] = sheet[f'H{row + 1}'].value
            sheet.cell(row=row, column=12).value = sheet.cell(row=row + 1, column=12).value
            sheet.cell(row=row, column=13).value = sheet.cell(row=row + 1, column=13).value
            sheet.cell(row=row, column=14).value = sheet.cell(row=row + 1, column=14).value
        
        # Очищаем последнюю строку
        sheet[f'E{max_row}'] = None
        sheet[f'F{max_row}'] = None
        sheet[f'G{max_row}'] = None
        sheet[f'H{max_row}'] = None
        sheet.cell(row=max_row, column=12).value = None
        sheet.cell(row=max_row, column=13).value = None
        sheet.cell(row=max_row, column=14).value = None

    def add_entries(self):
        if not all([self.wb, self.selected_dates, self.discipline_edit.text(), 
                   self.group_edit.text(), self.load_type_combo.currentText()]):
            QMessageBox.critical(self, "Ошибка", "Заполните все поля и сгенерируйте даты")
            return
        
        try:
            data = {
                'discipline': self.discipline_edit.text(),
                'group': self.group_edit.text(),
                'load_type': self.load_type_combo.currentText(),
                'lecture': float(self.lecture_edit.text()) if self.lecture_edit.text() else None,
                'practice': float(self.practice_edit.text()) if self.practice_edit.text() else None,
                'lab': float(self.lab_edit.text()) if self.lab_edit.text() else None
            }
            
            dates_by_sheet = {}
            for date_info in self.selected_dates:
                sheet_name = date_info['sheet']
                if sheet_name not in dates_by_sheet:
                    dates_by_sheet[sheet_name] = []
                dates_by_sheet[sheet_name].append(date_info)
            
            results = {}
            for sheet_name, dates in dates_by_sheet.items():
                sheet = self.wb[sheet_name]
                dates.sort(key=lambda x: x['day'])
                
                added_rows = []
                for date_info in dates:
                    row = self.add_entry_to_sheet(sheet, date_info['day'], data)
                    added_rows.append(f"{date_info['day']}.{date_info['month']:02d}(стр.{row})")
                
                results[sheet_name] = added_rows
            
            self.wb.save(self.filename)
            self.show_data()
            
            msg_lines = ["Записи добавлены:"]
            for sheet_name, dates in results.items():
                msg_lines.append(f"{sheet_name}: {', '.join(dates)}")
            
            QMessageBox.information(self, "Успех", "\n".join(msg_lines))
            
            # Очищаем поля ввода
            self.discipline_edit.clear()
            self.group_edit.clear()
            self.lecture_edit.clear()
            self.practice_edit.clear()
            self.lab_edit.clear()
            self.load_type_combo.setCurrentIndex(0)
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка: {e}")

    def add_entry_to_sheet(self, sheet, day, data):
        last_row = self.START_ROW
        while sheet[f'E{last_row}'].value is not None:
            last_row += 1
        
        existing_date_rows = []
        row = self.START_ROW
        while sheet[f'E{row}'].value is not None:
            existing_day = sheet[f'E{row}'].value
            if isinstance(existing_day, (int, float)) and int(existing_day) == day:
                existing_date_rows.append(row)
            row += 1
        
        if existing_date_rows:
            last_date_row = max(existing_date_rows)
            insert_row = last_date_row + 1
            self.shift_rows_down(sheet, insert_row, last_row)
            self.fill_row_data(sheet, insert_row, day, data)
            return insert_row
        else:
            return self.insert_entry_sorted(sheet, day, data, last_row)

    def shift_rows_down(self, sheet, start_row, last_row):
        for row in range(last_row, start_row - 1, -1):
            sheet[f'E{row+1}'] = sheet[f'E{row}'].value
            sheet[f'F{row+1}'] = sheet[f'F{row}'].value
            sheet[f'G{row+1}'] = sheet[f'G{row}'].value
            sheet[f'H{row+1}'] = sheet[f'H{row}'].value
            sheet.cell(row=row+1, column=12).value = sheet.cell(row=row, column=12).value
            sheet.cell(row=row+1, column=13).value = sheet.cell(row=row, column=13).value
            sheet.cell(row=row+1, column=14).value = sheet.cell(row=row, column=14).value

    def insert_entry_sorted(self, sheet, day, data, last_row):
        insert_row = self.START_ROW
        while sheet[f'E{insert_row}'].value is not None:
            existing_day = sheet[f'E{insert_row}'].value
            if isinstance(existing_day, (int, float)) and int(existing_day) > day:
                break
            insert_row += 1
        
        self.shift_rows_down(sheet, insert_row, last_row)
        self.fill_row_data(sheet, insert_row, day, data)
        return insert_row

    def fill_row_data(self, sheet, row, day, data):
        sheet[f'E{row}'] = day
        sheet[f'F{row}'] = data['discipline']
        sheet[f'G{row}'] = data['group']
        sheet[f'H{row}'] = data['load_type']
        sheet.cell(row=row, column=12).value = data['lecture']
        sheet.cell(row=row, column=13).value = data['practice']
        sheet.cell(row=row, column=14).value = data['lab']

if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Установка стиля для лучшего внешнего вида
    app.setStyle('Fusion')
    
    # Пытаемся найти файл журнала рядом с EXE
    journal_filename = "Тетрадь_ППС_2025_2026_каф_NN_Фамилия_ИО_оч_заоч.xltx"
    
    # Проверяем существование файла
    if not os.path.exists(journal_filename):
        # Если файл не найден, создаем приложение без файла
        window = JournalApp()
    else:
        window = JournalApp(journal_filename)
    
    window.show()
    
    sys.exit(app.exec())