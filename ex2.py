import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import Calendar, DateEntry
from datetime import datetime, timedelta, date
import re
from PIL import Image, ImageTk  # Добавляем импорт для работы с изображениями

class JournalApp:
    def __init__(self, filename):
        self.filename = filename
        self.wb = None
        self.START_ROW = 7
        self.HOURS_COLS = {'lecture': 12, 'practice': 13, 'lab': 14}
        self.selected_dates = []
        self.LOAD_TYPES = ["осн.", "почас.", "совм."]
        
        self.setup_gui()
        self.load_workbook()
    
    def setup_gui(self):
        self.root = tk.Tk()
        self.root.title("Журнал преподавателя")
        self.root.geometry("1000x900")  # Увеличил ширину для размещения GIF
        
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Выбор периода
        period_frame = self.create_period_frame(main_frame)
        period_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Управление датами с GIF
        dates_frame = self.create_dates_frame(main_frame)
        dates_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Поля ввода
        self.entries = self.create_input_fields(main_frame)
        
        # Основные кнопки
        ttk.Button(main_frame, text="Добавить записи", command=self.add_entries).grid(row=8, column=0, columnspan=2, pady=15)
        
        # Просмотр данных
        view_frame = self.create_view_frame(main_frame)
        view_frame.grid(row=9, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # Настройка растягивания
        self.configure_grid(main_frame, view_frame)
    
    def create_period_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="Выбор периода", padding="10")
        
        # Даты
        ttk.Label(frame, text="Начало периода:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.start_date = DateEntry(frame, width=12, background='darkblue', foreground='white', 
                                  borderwidth=2, date_pattern='dd.mm.yyyy', locale='ru_RU')
        self.start_date.set_date(datetime(datetime.now().year, 9, 1))
        self.start_date.grid(row=0, column=1, sticky=tk.W, padx=(0, 15))
        
        ttk.Label(frame, text="Конец периода:").grid(row=0, column=2, sticky=tk.W, padx=(0, 5))
        self.end_date = DateEntry(frame, width=12, background='darkblue', foreground='white',
                                borderwidth=2, date_pattern='dd.mm.yyyy', locale='ru_RU')
        self.end_date.set_date(datetime(datetime.now().year, 12, 31))
        self.end_date.grid(row=0, column=3, sticky=tk.W, padx=(0, 15))
        
        # Тип недели
        ttk.Label(frame, text="Тип недели:").grid(row=0, column=4, sticky=tk.W, padx=(0, 5))
        self.period_week_type = tk.StringVar(value="числитель")
        
        week_type_frame = ttk.Frame(frame)
        week_type_frame.grid(row=0, column=5, columnspan=3, sticky=tk.W)
        
        for text, value in [("Числитель", "числитель"), ("Знаменатель", "знаменатель"), ("Обе недели", "обе недели")]:
            ttk.Radiobutton(week_type_frame, text=text, variable=self.period_week_type, value=value).pack(side=tk.LEFT, padx=(0, 5))
        
        # Кнопки
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=1, column=0, columnspan=7, pady=10, sticky=tk.W)
        
        ttk.Button(btn_frame, text="Сгенерировать даты по периоду", command=self.generate_dates_by_period).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Очистить даты", command=self.clear_dates).pack(side=tk.LEFT, padx=5)
        
        # Информация о датах
        self.dates_info_label = ttk.Label(frame, text="Выбрано дат: 0", foreground="blue", font=('Arial', 10, 'bold'))
        self.dates_info_label.grid(row=2, column=0, columnspan=7, sticky=tk.W, pady=2)
        
        return frame
    
    def create_dates_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="Управление датами", padding="10")
        
        # Создаем основной контейнер с двумя колонками
        main_container = ttk.Frame(frame)
        main_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Левая колонка - элементы управления датами
        left_frame = ttk.Frame(main_container)
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.N), padx=(0, 10))
        
        # Правая колонка - GIF
        right_frame = ttk.Frame(main_container)
        right_frame.grid(row=0, column=1, sticky=(tk.E, tk.N))
        
        # Добавление даты
        ttk.Label(left_frame, text="Добавить дату:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.single_date = DateEntry(left_frame, width=12, background='darkblue', foreground='white',
                                   borderwidth=2, date_pattern='dd.mm.yyyy', locale='ru_RU')
        self.single_date.grid(row=0, column=1, sticky=tk.W, padx=(0, 15))
        
        ttk.Button(left_frame, text="Добавить дату", command=self.add_single_date).grid(row=0, column=2, sticky=tk.W, padx=5)
        
        # Удаление даты
        ttk.Label(left_frame, text="Удалить дату:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5), pady=(10, 0))
        self.remove_date_combo = ttk.Combobox(left_frame, width=15, state="readonly")
        self.remove_date_combo.grid(row=1, column=1, sticky=tk.W, padx=(0, 15), pady=(10, 0))
        
        ttk.Button(left_frame, text="Удалить дату", command=self.remove_selected_date).grid(row=1, column=2, sticky=tk.W, padx=5, pady=(10, 0))
        
        # Список выбранных дат
        ttk.Label(left_frame, text="Выбранные даты:").grid(row=2, column=0, sticky=tk.W, pady=(10, 5))
        
        self.dates_listbox = tk.Listbox(left_frame, width=40, height=6)
        self.dates_listbox.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # Кнопки управления списком
        btn_frame = ttk.Frame(left_frame)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=5, sticky=tk.W)
        
        ttk.Button(btn_frame, text="Очистить все даты", command=self.clear_dates).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Обновить список", command=self.update_dates_display).pack(side=tk.LEFT, padx=5)
        
        # Загрузка и отображение GIF
        self.load_and_display_gif(right_frame)
        
        # Настройка растягивания для основного контейнера
        main_container.columnconfigure(0, weight=1)
        main_container.columnconfigure(1, weight=0)
        frame.columnconfigure(0, weight=1)
        
        return frame
    
    def load_and_display_gif(self, parent):
        """Загружает и отображает GIF изображение"""
        try:
            # Пытаемся загрузить GIF
            self.gif_image = Image.open("Без названия.gif")
            
            # Конвертируем в формат, понятный Tkinter
            self.gif_frames = []
            try:
                # Для анимированных GIF
                while True:
                    frame = self.gif_image.copy()
                    # Масштабируем изображение (можно настроить размер)
                    frame = frame.resize((250, 250), Image.Resampling.LANCZOS)
                    self.gif_frames.append(ImageTk.PhotoImage(frame))
                    self.gif_image.seek(len(self.gif_frames))
            except EOFError:
                # Достигнут конец GIF
                pass
                
            # Если это не анимированный GIF, создаем один кадр
            if not self.gif_frames:
                frame = self.gif_image.copy()
                frame = frame.resize((250, 250), Image.Resampling.LANCZOS)
                self.gif_frames = [ImageTk.PhotoImage(frame)]
            
            # Создаем метку для отображения GIF
            self.gif_label = ttk.Label(parent)
            self.gif_label.grid(row=0, column=0, sticky=(tk.E, tk.N), padx=(10, 0))
            
            # Запускаем анимацию если есть несколько кадров
            if len(self.gif_frames) > 1:
                self.current_frame = 0
                self.animate_gif()
            else:
                self.gif_label.configure(image=self.gif_frames[0])
                
        except Exception as e:
            # Если не удалось загрузить GIF, создаем заглушку
            print(f"Не удалось загрузить GIF: {e}")
            error_label = ttk.Label(parent, text="GIF не найден", foreground="red")
            error_label.grid(row=0, column=0, sticky=(tk.E, tk.N), padx=(10, 0))
    
    def animate_gif(self):
        """Анимирует GIF изображение"""
        if hasattr(self, 'gif_frames') and self.gif_frames:
            self.gif_label.configure(image=self.gif_frames[self.current_frame])
            self.current_frame = (self.current_frame + 1) % len(self.gif_frames)
            # Обновляем каждые 100 мс (можно настроить скорость)
            self.root.after(25, self.animate_gif)
    
    def create_input_fields(self, parent):
        fields = [
            ("Дисциплина:", "discipline", "entry"), 
            ("Группа:", "group", "entry"),
            ("Вид нагрузки:", "load_type", "combobox"),
            ("Лекции:", "lecture", "entry"),
            ("Практические:", "practice", "entry"),
            ("Лабораторные:", "lab", "entry")
        ]
        
        entries = {}
        for i, (label, field, field_type) in enumerate(fields):
            ttk.Label(parent, text=label).grid(row=i+2, column=0, sticky=tk.W, pady=2)
            
            if field_type == "combobox":
                entries[field] = ttk.Combobox(parent, values=self.LOAD_TYPES, width=27)
                entries[field].set(self.LOAD_TYPES[0])
            else:
                entries[field] = ttk.Entry(parent, width=30)
            
            entries[field].grid(row=i+2, column=1, sticky=(tk.W, tk.E), padx=5, pady=2)
        
        return entries
    
    def create_view_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="Просмотр данных", padding="10")
        
        # Панель управления просмотром
        control_frame = ttk.Frame(frame)
        control_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(control_frame, text="Лист:").grid(row=0, column=0, sticky=tk.W)
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(control_frame, textvariable=self.sheet_var, width=25)
        self.sheet_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        self.sheet_combo.bind('<<ComboboxSelected>>', self.show_data)
        
        # Кнопки удаления записей
        btn_frame = ttk.Frame(control_frame)
        btn_frame.grid(row=0, column=2, padx=(20, 0))
        
        ttk.Button(btn_frame, text="Удалить выбранные записи", 
                  command=self.delete_selected_entries, style="Danger.TButton").pack(side=tk.LEFT, padx=2)
        
        ttk.Button(btn_frame, text="Выбрать все", 
                  command=self.select_all_entries).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(btn_frame, text="Снять выделение", 
                  command=self.deselect_all_entries).pack(side=tk.LEFT, padx=2)
        
        # Создаем стиль для красной кнопки
        style = ttk.Style()
        style.configure("Danger.TButton", background="#ff6b6b", foreground="black")
        
        # Информация о выборе
        self.selection_info = ttk.Label(control_frame, text="Выбрано записей: 0", foreground="green")
        self.selection_info.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        # Таблица
        columns = ("Число", "Дисциплина", "Группа", "Нагрузка", "Лекции", "Практические", "Лабораторные")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings", height=10, selectmode="extended")
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)
        
        self.tree.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # Привязываем обработчик выбора
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        
        # Полоса прокрутки для таблицы
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.grid(row=2, column=2, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        return frame
    
    def on_tree_select(self, event):
        """Обновляет информацию о количестве выбранных записей"""
        selected_count = len(self.tree.selection())
        self.selection_info.config(text=f"Выбрано записей: {selected_count}")
    
    def select_all_entries(self):
        """Выбирает все записи в таблице"""
        all_items = self.tree.get_children()
        self.tree.selection_set(all_items)
        self.on_tree_select(None)
    
    def deselect_all_entries(self):
        """Снимает выделение со всех записей"""
        self.tree.selection_remove(self.tree.selection())
        self.on_tree_select(None)
    
    def configure_grid(self, main_frame, view_frame):
        view_frame.columnconfigure(1, weight=1)
        view_frame.rowconfigure(2, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(9, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
    
    def delete_selected_entries(self):
        """Удаляет выбранные записи из таблицы и файла Excel"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Внимание", "Выберите записи для удаления")
            return
        
        # Получаем данные выбранных записей
        entries_to_delete = []
        for item in selected_items:
            item_data = self.tree.item(item, 'values')
            if item_data:
                entries_to_delete.append(item_data)
        
        if not entries_to_delete:
            return
        
        # Подтверждение удаления
        confirm = messagebox.askyesno(
            "Подтверждение удаления", 
            f"Вы действительно хотите удалить {len(entries_to_delete)} записей?\n"
            f"Это действие нельзя отменить."
        )
        
        if not confirm:
            return
        
        try:
            sheet_name = self.sheet_var.get()
            if sheet_name not in self.wb.sheetnames:
                messagebox.showerror("Ошибка", "Лист не найден")
                return
            
            sheet = self.wb[sheet_name]
            deleted_count = 0
            
            # Собираем информацию о строках для удаления
            rows_to_delete = []
            for entry_data in entries_to_delete:
                day = int(entry_data[0])
                discipline = entry_data[1]
                group = entry_data[2]
                
                # Ищем строку для удаления
                row = self.START_ROW
                while sheet[f'E{row}'].value is not None:
                    sheet_day = sheet[f'E{row}'].value
                    sheet_discipline = sheet[f'F{row}'].value or ''
                    sheet_group = sheet[f'G{row}'].value or ''
                    
                    if (isinstance(sheet_day, (int, float)) and int(sheet_day) == day and
                        sheet_discipline == discipline and sheet_group == group):
                        rows_to_delete.append(row)
                        break
                    row += 1
            
            # Удаляем строки в обратном порядке (чтобы индексы не сдвигались)
            rows_to_delete.sort(reverse=True)
            for row_num in rows_to_delete:
                self.delete_row(sheet, row_num)
                deleted_count += 1
            
            # Сохраняем файл
            self.wb.save(self.filename)
            
            # Обновляем отображение
            self.show_data()
            
            messagebox.showinfo("Успех", f"Удалено записей: {deleted_count} из {len(entries_to_delete)}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при удалении записей: {e}")
    
    def delete_row(self, sheet, row_num):
        """Удаляет строку из листа Excel и сдвигает остальные строки вверх"""
        max_row = self.START_ROW
        while sheet[f'E{max_row}'].value is not None:
            max_row += 1
        
        # Сдвигаем строки вверх
        for row in range(row_num, max_row):
            # Копируем значения из следующей строки
            sheet[f'E{row}'] = sheet[f'E{row + 1}'].value
            sheet[f'F{row}'] = sheet[f'F{row + 1}'].value
            sheet[f'G{row}'] = sheet[f'G{row + 1}'].value
            sheet[f'H{row}'] = sheet[f'H{row + 1}'].value
            sheet.cell(row=row, column=12).value = sheet.cell(row=row + 1, column=12).value
            sheet.cell(row=row, column=13).value = sheet.cell(row=row + 1, column=13).value
            sheet.cell(row=row, column=14).value = sheet.cell(row=row + 1, column=14).value
        
        # Очищаем последнюю строку
        sheet[f'E{max_row}'] = None
        sheet[f'F{max_row}'] = None
        sheet[f'G{max_row}'] = None
        sheet[f'H{max_row}'] = None
        sheet.cell(row=max_row, column=12).value = None
        sheet.cell(row=max_row, column=13).value = None
        sheet.cell(row=max_row, column=14).value = None
    
    def date_to_datetime(self, date_obj):
        return datetime.combine(date_obj, datetime.min.time()) if isinstance(date_obj, date) and not isinstance(date_obj, datetime) else date_obj
    
    def determine_week_type(self, input_date):
        current_date = self.date_to_datetime(input_date)
        year = current_date.year
        
        september_1 = datetime(year, 9, 1)
        if current_date < september_1:
            september_1 = datetime(year-1, 9, 1)
        
        days_diff = (current_date - september_1).days
        return "числитель" if (days_diff // 7) % 2 == 0 else "знаменатель"
    
    def find_sheet_for_month(self, month):
        month_names = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",
                      7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        
        target_month = month_names.get(month)
        if not target_month:
            return None
        
        for sheet_name in self.wb.sheetnames:
            if re.search(r'\b' + re.escape(target_month) + r'\b', sheet_name):
                return sheet_name
        
        return None
    
    def generate_dates_by_period(self):
        try:
            start_dt = self.date_to_datetime(self.start_date.get_date())
            end_dt = self.date_to_datetime(self.end_date.get_date())
            target_week_type = self.period_week_type.get()
            
            if start_dt >= end_dt:
                messagebox.showerror("Ошибка", "Дата начала должна быть раньше даты окончания")
                return
            
            self.selected_dates.clear()
            generated_count = 0
            
            if target_week_type == "обе недели":
                current_date = start_dt
                while current_date <= end_dt:
                    sheet_name = self.find_sheet_for_month(current_date.month)
                    if sheet_name:
                        self.selected_dates.append({
                            'date': current_date, 'day': current_date.day, 'month': current_date.month,
                            'year': current_date.year, 'sheet': sheet_name,
                            'week_type': self.determine_week_type(current_date)
                        })
                        generated_count += 1
                    current_date += timedelta(days=7)
            else:
                current_date = start_dt
                while current_date <= end_dt:
                    week_type = self.determine_week_type(current_date)
                    if week_type == target_week_type:
                        sheet_name = self.find_sheet_for_month(current_date.month)
                        if sheet_name:
                            self.selected_dates.append({
                                'date': current_date, 'day': current_date.day, 'month': current_date.month,
                                'year': current_date.year, 'sheet': sheet_name, 'week_type': week_type
                            })
                            generated_count += 1
                        current_date += timedelta(days=7)
                    else:
                        current_date += timedelta(days=1)
            
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            self.update_dates_info()
            self.update_dates_display()
            
            if generated_count > 0:
                dates_list = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
                week_type_display = "числитель и знаменатель (каждую неделю)" if target_week_type == "обе недели" else target_week_type
                
                messagebox.showinfo("Успех", 
                    f"Сгенерировано {generated_count} дат\n"
                    f"Период: {start_dt.strftime('%d.%m.%Y')} - {end_dt.strftime('%d.%m.%Y')}\n"
                    f"Тип недели: {week_type_display}\n"
                    f"Даты: {dates_list}")
            else:
                messagebox.showwarning("Внимание", "В выбранном периоде нет дат")
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка генерации дат: {e}")
    
    def add_single_date(self):
        try:
            date_obj = self.date_to_datetime(self.single_date.get_date())
            sheet_name = self.find_sheet_for_month(date_obj.month)
            
            if not sheet_name:
                messagebox.showerror("Ошибка", f"Не найден лист для месяца {date_obj.month}")
                return
            
            # Проверяем, нет ли уже этой даты в списке
            for existing_date in self.selected_dates:
                if (existing_date['day'] == date_obj.day and 
                    existing_date['month'] == date_obj.month and 
                    existing_date['year'] == date_obj.year):
                    messagebox.showwarning("Внимание", "Эта дата уже есть в списке")
                    return
            
            date_info = {
                'date': date_obj, 
                'day': date_obj.day, 
                'month': date_obj.month,
                'year': date_obj.year, 
                'sheet': sheet_name,
                'week_type': self.determine_week_type(date_obj)
            }
            
            self.selected_dates.append(date_info)
            self.selected_dates.sort(key=lambda x: (x['month'], x['day']))
            
            self.update_dates_info()
            self.update_dates_display()
            
            messagebox.showinfo("Успех", f"Дата {date_obj.strftime('%d.%m.%Y')} добавлена")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка добавления даты: {e}")
    
    def remove_selected_date(self):
        selected_index = self.remove_date_combo.current()
        if selected_index == -1:
            messagebox.showwarning("Внимание", "Выберите дату для удаления")
            return
        
        if 0 <= selected_index < len(self.selected_dates):
            removed_date = self.selected_dates.pop(selected_index)
            self.update_dates_info()
            self.update_dates_display()
            messagebox.showinfo("Успех", f"Дата {removed_date['day']}.{removed_date['month']:02d}.{removed_date['year']} удалена")
    
    def update_dates_display(self):
        # Обновляем список дат в Listbox
        self.dates_listbox.delete(0, tk.END)
        for date_info in self.selected_dates:
            display_text = f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']} ({date_info['sheet']}, {date_info['week_type']})"
            self.dates_listbox.insert(tk.END, display_text)
        
        # Обновляем комбобокс для удаления
        date_values = [f"{date_info['day']:02d}.{date_info['month']:02d}.{date_info['year']}" for date_info in self.selected_dates]
        self.remove_date_combo['values'] = date_values
        if date_values:
            self.remove_date_combo.set(date_values[0])
        else:
            self.remove_date_combo.set('')
    
    def clear_dates(self):
        self.selected_dates.clear()
        self.update_dates_info()
        self.update_dates_display()
        messagebox.showinfo("Успех", "Все даты очищены")
    
    def update_dates_info(self):
        count = len(self.selected_dates)
        if count > 0:
            dates_str = ", ".join([f"{date['day']}.{date['month']:02d}" for date in self.selected_dates])
            sheets_count = {}
            for date in self.selected_dates:
                sheets_count[date['sheet']] = sheets_count.get(date['sheet'], 0) + 1
            
            sheets_info = ", ".join([f"{sheet}: {count}" for sheet, count in sheets_count.items()])
            week_types = set(date['week_type'] for date in self.selected_dates)
            week_types_info = f"Типы: {', '.join(week_types)}" if week_types else ""
            
            self.dates_info_label.config(text=f"Выбрано дат: {count} | Даты: {dates_str} | Листы: {sheets_info} {week_types_info}")
        else:
            self.dates_info_label.config(text="Выбрано дат: 0")
    
    def load_workbook(self):
        try:
            self.wb = openpyxl.load_workbook(self.filename)
            sheets = self.wb.sheetnames
            self.sheet_combo['values'] = sheets
            if sheets:
                self.sheet_var.set(sheets[0])
            self.show_data()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка загрузки: {e}")
    
    def show_data(self, event=None):
        if not self.wb:
            return
        
        sheet_name = self.sheet_var.get()
        if sheet_name in self.wb.sheetnames:
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            try:
                sheet = self.wb[sheet_name]
                row = self.START_ROW
                
                while sheet[f'E{row}'].value is not None:
                    day = sheet[f'E{row}'].value
                    if isinstance(day, (int, float)):
                        self.tree.insert("", "end", values=(
                            int(day),
                            sheet[f'F{row}'].value or '',
                            sheet[f'G{row}'].value or '',
                            sheet[f'H{row}'].value or '',
                            sheet.cell(row=row, column=12).value or '',
                            sheet.cell(row=row, column=13).value or '',
                            sheet.cell(row=row, column=14).value or ''
                        ))
                    row += 1
                    
                # Сбрасываем счетчик выбранных записей
                self.selection_info.config(text="Выбрано записей: 0")
                
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка данных: {e}")
    
    def add_entries(self):
        if not all([self.wb, self.selected_dates, self.entries['discipline'].get(), 
                   self.entries['group'].get(), self.entries['load_type'].get()]):
            messagebox.showerror("Ошибка", "Заполните все поля и сгенерируйте даты")
            return
        
        try:
            data = {
                'discipline': self.entries['discipline'].get(),
                'group': self.entries['group'].get(),
                'load_type': self.entries['load_type'].get(),
                'lecture': float(self.entries['lecture'].get()) if self.entries['lecture'].get() else None,
                'practice': float(self.entries['practice'].get()) if self.entries['practice'].get() else None,
                'lab': float(self.entries['lab'].get()) if self.entries['lab'].get() else None
            }
            
            dates_by_sheet = {}
            for date_info in self.selected_dates:
                sheet_name = date_info['sheet']
                if sheet_name not in dates_by_sheet:
                    dates_by_sheet[sheet_name] = []
                dates_by_sheet[sheet_name].append(date_info)
            
            results = {}
            for sheet_name, dates in dates_by_sheet.items():
                sheet = self.wb[sheet_name]
                dates.sort(key=lambda x: x['day'])
                
                added_rows = []
                for date_info in dates:
                    row = self.add_entry_to_sheet(sheet, date_info['day'], data)
                    added_rows.append(f"{date_info['day']}.{date_info['month']:02d}(стр.{row})")
                
                results[sheet_name] = added_rows
            
            self.wb.save(self.filename)
            self.show_data()
            
            msg_lines = ["Записи добавлены:"]
            for sheet_name, dates in results.items():
                msg_lines.append(f"{sheet_name}: {', '.join(dates)}")
            
            messagebox.showinfo("Успех", "\n".join(msg_lines))
            
            for entry in self.entries.values():
                if isinstance(entry, ttk.Entry):
                    entry.delete(0, tk.END)
                elif isinstance(entry, ttk.Combobox):
                    entry.set(self.LOAD_TYPES[0])
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка: {e}")
    
    def add_entry_to_sheet(self, sheet, day, data):
        last_row = self.START_ROW
        while sheet[f'E{last_row}'].value is not None:
            last_row += 1
        
        existing_date_rows = []
        row = self.START_ROW
        while sheet[f'E{row}'].value is not None:
            existing_day = sheet[f'E{row}'].value
            if isinstance(existing_day, (int, float)) and int(existing_day) == day:
                existing_date_rows.append(row)
            row += 1
        
        if existing_date_rows:
            last_date_row = max(existing_date_rows)
            insert_row = last_date_row + 1
            self.shift_rows_down(sheet, insert_row, last_row)
            self.fill_row_data(sheet, insert_row, day, data)
            return insert_row
        else:
            return self.insert_entry_sorted(sheet, day, data, last_row)
    
    def shift_rows_down(self, sheet, start_row, last_row):
        for row in range(last_row, start_row - 1, -1):
            sheet[f'E{row+1}'] = sheet[f'E{row}'].value
            sheet[f'F{row+1}'] = sheet[f'F{row}'].value
            sheet[f'G{row+1}'] = sheet[f'G{row}'].value
            sheet[f'H{row+1}'] = sheet[f'H{row}'].value
            sheet.cell(row=row+1, column=12).value = sheet.cell(row=row, column=12).value
            sheet.cell(row=row+1, column=13).value = sheet.cell(row=row, column=13).value
            sheet.cell(row=row+1, column=14).value = sheet.cell(row=row, column=14).value
    
    def insert_entry_sorted(self, sheet, day, data, last_row):
        insert_row = self.START_ROW
        while sheet[f'E{insert_row}'].value is not None:
            existing_day = sheet[f'E{insert_row}'].value
            if isinstance(existing_day, (int, float)) and int(existing_day) > day:
                break
            insert_row += 1
        
        self.shift_rows_down(sheet, insert_row, last_row)
        self.fill_row_data(sheet, insert_row, day, data)
        return insert_row
    
    def fill_row_data(self, sheet, row, day, data):
        sheet[f'E{row}'] = day
        sheet[f'F{row}'] = data['discipline']
        sheet[f'G{row}'] = data['group']
        sheet[f'H{row}'] = data['load_type']
        sheet.cell(row=row, column=12).value = data['lecture']
        sheet.cell(row=row, column=13).value = data['practice']
        sheet.cell(row=row, column=14).value = data['lab']
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = JournalApp("Тетрадь_ППС_2025_2026_каф_NN_Фамилия_ИО_оч_заоч.xltx")
    app.run()