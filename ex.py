import openpyxl
from openpyxl import Workbook
import os

class JournalFiller:
    def __init__(self, filename):
        self.filename = filename
        self.wb = None
        self.START_COLUMN = 5  # Столбец E
        self.START_ROW = 7     # Строка 7
        self.load_workbook()
        
    def load_workbook(self):
        """Загрузка файла Excel"""
        try:
            self.wb = openpyxl.load_workbook(self.filename)
            print(f"Файл '{self.filename}' успешно загружен")
        except FileNotFoundError:
            print(f"Файл '{self.filename}' не найден.")
            return False
        return True
    
    def get_available_sheets(self):
        """Получить список доступных листов"""
        return self.wb.sheetnames if self.wb else []
    
    def get_existing_days(self, sheet):
        """Получить список существующих чисел и их строк"""
        days_data = []
        row = self.START_ROW
        
        while sheet[f'E{row}'].value is not None:
            day_number = sheet[f'E{row}'].value
            if isinstance(day_number, (int, float)):
                days_data.append({
                    'row': row,
                    'day': int(day_number),
                    'discipline': sheet[f'F{row}'].value,
                    'group': sheet[f'G{row}'].value,
                    'load_type': sheet[f'H{row}'].value
                })
            row += 1
        
        # Сортируем по числам
        days_data.sort(key=lambda x: x['day'])
        return days_data
    
    def find_insert_position(self, sheet, new_day):
        """Найти позицию для вставки нового числа (с сохранением сортировки)"""
        existing_days = self.get_existing_days(sheet)
        
        if not existing_days:
            return self.START_ROW  # Первая запись
        
        # Проверяем, существует ли уже такое число
        for day_data in existing_days:
            if day_data['day'] == new_day:
                return day_data['row']  # Заменяем существующую запись
        
        # Ищем позицию для вставки
        for i, day_data in enumerate(existing_days):
            if new_day < day_data['day']:
                # Вставляем перед текущим элементом
                if i == 0:
                    return self.START_ROW
                else:
                    return existing_days[i-1]['row'] + 1
        
        # Если число больше всех существующих, вставляем в конец
        return existing_days[-1]['row'] + 1
    
    def insert_sorted_entry(self, sheet, insert_row, schedule_data):
        """Вставить запись в указанную строку, сдвигая существующие данные вниз"""
        # Получаем все данные ниже точки вставки
        existing_data = []
        row = insert_row
        
        # Собираем все существующие данные начиная с строки вставки
        while sheet[f'E{row}'].value is not None:
            existing_data.append({
                'day': sheet[f'E{row}'].value,
                'discipline': sheet[f'F{row}'].value,
                'group': sheet[f'G{row}'].value,
                'load_type': sheet[f'H{row}'].value
            })
            # Очищаем ячейки для сдвига
            sheet[f'E{row}'] = None
            sheet[f'F{row}'] = None
            sheet[f'G{row}'] = None
            sheet[f'H{row}'] = None
            row += 1
        
        # Вставляем новую запись
        sheet[f'E{insert_row}'] = schedule_data['day_number']
        sheet[f'F{insert_row}'] = schedule_data['discipline']
        sheet[f'G{insert_row}'] = schedule_data['group']
        sheet[f'H{insert_row}'] = schedule_data['load_type']
        
        # Восстанавливаем сдвинутые данные
        current_row = insert_row + 1
        for data in existing_data:
            sheet[f'E{current_row}'] = data['day']
            sheet[f'F{current_row}'] = data['discipline']
            sheet[f'G{current_row}'] = data['group']
            sheet[f'H{current_row}'] = data['load_type']
            current_row += 1
        
        return insert_row
    
    def add_schedule_entry(self, month, schedule_data):
        """
        Добавление записи расписания с автоматической сортировкой по числам
        """
        if not self.wb or month not in self.wb.sheetnames:
            print(f"Лист '{month}' не найден!")
            return False
        
        sheet = self.wb[month]
        
        try:
            # Находим позицию для вставки с учетом сортировки
            insert_row = self.find_insert_position(sheet, schedule_data['day_number'])
            
            # Вставляем запись
            result_row = self.insert_sorted_entry(sheet, insert_row, schedule_data)
            
            print(f"Запись добавлена на лист '{month}':")
            print(f"  Строка {result_row}:")
            print(f"    E: {schedule_data['day_number']} (число)")
            print(f"    F: {schedule_data['discipline']} (дисциплина)")
            print(f"    G: {schedule_data['group']} (группа)")
            print(f"    H: {schedule_data['load_type']} (нагрузка)")
            return True
            
        except Exception as e:
            print(f"Ошибка при добавлении записи: {e}")
            return False
    
    def add_multiple_entries(self, entries_list):
        """Добавление нескольких записей"""
        if not self.wb:
            print("Файл не загружен!")
            return 0
            
        success_count = 0
        for entry in entries_list:
            if self.add_schedule_entry(entry['month'], entry['schedule_data']):
                success_count += 1
        
        print(f"Успешно добавлено {success_count} из {len(entries_list)} записей")
        return success_count
    
    def save_workbook(self):
        """Сохранение файла"""
        if not self.wb:
            print("Файл не загружен для сохранения!")
            return False
            
        try:
            self.wb.save(self.filename)
            print(f"Файл '{self.filename}' успешно сохранен")
            return True
        except Exception as e:
            print(f"Ошибка при сохранении: {e}")
            return False
    
    def show_current_data(self, month):
        """Показать текущие данные на листе"""
        if not self.wb or month not in self.wb.sheetnames:
            print(f"Лист '{month}' не найден!")
            return
        
        sheet = self.wb[month]
        existing_days = self.get_existing_days(sheet)
        
        print(f"\nТекущие данные на листе '{month}':")
        print("-" * 80)
        print("Строка | Число (E) | Дисциплина (F) | Группа (G) | Нагрузка (H)")
        print("-" * 80)
        
        if not existing_days:
            print("Нет данных")
            return
        
        for day_data in existing_days:
            print(f"{day_data['row']:6} | {day_data['day']:9} | {day_data['discipline']:13} | {day_data['group']:10} | {day_data['load_type']}")

def validate_day_number(day, month=None):
    """Проверка корректности числа месяца"""
    try:
        day_int = int(day)
        if day_int < 1 or day_int > 31:
            return False, "Число должно быть в диапазоне от 1 до 31"
        
        # Дополнительная проверка по месяцам (опционально)
        if month:
            month_days = {
                'Январь': 31, 'Февраль': 29, 'Март': 31, 'Апрель': 30,
                'Май': 31, 'Июнь': 30, 'Июль': 31, 'Август': 31,
                'Сентябрь': 30, 'Октябрь': 31, 'Ноябрь': 30, 'Декабрь': 31
            }
            max_days = month_days.get(month, 31)
            if day_int > max_days:
                return False, f"В месяце '{month}' не может быть {day_int} числа"
        
        return True, day_int
    except ValueError:
        return False, "Введите корректное число"

def main():
    # Инициализация
    journal = JournalFiller("ПРимер.xlsx")
    
    if not journal.wb:
        print("Не удалось загрузить файл. Программа завершена.")
        return
    
    while True:
        print("\n" + "="*60)
        print("ЖУРНАЛ ПРЕПОДАВАТЕЛЯ - ЗАПОЛНЕНИЕ РАСПИСАНИЯ")
        print("="*60)
        print("Доступные листы:", ", ".join(journal.get_available_sheets()))
        print(f"Заполнение начинается с позиции: E{journal.START_ROW}")
        print("\nФормат заполнения:")
        print("  E - число месяца | F - дисциплина | G - группа | H - вид нагрузки")
        print("\nОсобенности:")
        print("  • Числа автоматически сортируются по возрастанию")
        print("  • При добавлении существующего числа запись заменяется")
        print("  • Проверка корректности чисел (1-31)")
        print("\n1. Добавить запись расписания")
        print("2. Добавить несколько записей")
        print("3. Показать текущие данные")
        print("4. Сохранить и выйти")
        print("5. Выйти без сохранения")
        
        choice = input("\nВыберите действие (1-5): ").strip()
        
        if choice == '1':
            add_single_entry(journal)
        elif choice == '2':
            add_multiple_entries_interactive(journal)
        elif choice == '3':
            month = input("Введите название листа для просмотра: ").strip()
            journal.show_current_data(month)
        elif choice == '4':
            if journal.save_workbook():
                print("До свидания!")
                break
        elif choice == '5':
            print("Изменения не сохранены. До свидания!")
            break
        else:
            print("Неверный выбор!")

def add_single_entry(journal):
    """Добавление одной записи через интерфейс"""
    print("\n--- ДОБАВЛЕНИЕ ЗАПИСИ РАСПИСАНИЯ ---")
    
    month = input("Введите месяц (название листа): ").strip()
    
    # Проверка числа месяца
    while True:
        day_input = input("Число месяца (1-31): ").strip()
        is_valid, day_result = validate_day_number(day_input, month)
        if is_valid:
            day_number = day_result
            break
        else:
            print(f"Ошибка: {day_result}")
    
    discipline = input("Дисциплина: ").strip()
    group = input("Группа (например: 606-41 (АСОИУ-24-1)): ").strip()
    load_type = input("Вид нагрузки (осн./совм./почас.): ").strip()
    
    schedule_data = {
        'day_number': day_number,
        'discipline': discipline,
        'group': group,
        'load_type': load_type
    }
    
    if journal.add_schedule_entry(month, schedule_data):
        save_now = input("Сохранить файл сейчас? (да/нет): ").strip().lower()
        if save_now == 'да':
            journal.save_workbook()

def add_multiple_entries_interactive(journal):
    """Добавление нескольких записей через интерфейс"""
    entries = []
    
    print("\n--- ДОБАВЛЕНИЕ НЕСКОЛЬКИХ ЗАПИСЕЙ ---")
    print("Вводите данные для каждой записи. Для завершения введите 'готово' в поле месяца.")
    
    while True:
        month = input("\nМесяц (или 'готово' для завершения): ").strip()
        if month.lower() == 'готово':
            break
        
        # Проверка числа месяца
        while True:
            day_input = input("Число месяца (1-31): ").strip()
            is_valid, day_result = validate_day_number(day_input, month)
            if is_valid:
                day_number = day_result
                break
            else:
                print(f"Ошибка: {day_result}")
        
        discipline = input("Дисциплина: ").strip()
        group = input("Группа: ").strip()
        load_type = input("Вид нагрузки (осн./совм./почас.): ").strip()
        
        schedule_data = {
            'day_number': day_number,
            'discipline': discipline,
            'group': group,
            'load_type': load_type
        }
        
        entries.append({
            'month': month,
            'schedule_data': schedule_data
        })
        
        more = input("Добавить еще одну запись для этого месяца? (да/нет): ").strip().lower()
        if more != 'да':
            break
    
    if entries:
        journal.add_multiple_entries(entries)
        save = input("Сохранить файл? (да/нет): ").strip().lower()
        if save == 'да':
            journal.save_workbook()

# Демонстрация работы сортировки
def demo_sorted_fill():
    """Демонстрация работы с сортировкой"""
    try:
        wb = openpyxl.load_workbook("ПРимер.xlsx")
        sheet = wb['Сентябрь']
        
        # Очищаем старые данные для демонстрации
        for row in range(7, 20):
            sheet[f'E{row}'] = None
            sheet[f'F{row}'] = None
            sheet[f'G{row}'] = None
            sheet[f'H{row}'] = None
        
        # Данные вразнобой для демонстрации сортировки
        demo_data = [
            [15, "Математический анализ", "606-41", "совм."],
            [3, "Физика", "607-42", "осн."],
            [25, "Информатика", "608-43", "совм."],
            [8, "Программирование", "609-44", "почас."],
            [1, "Химия", "610-45", "осн."],
        ]
        
        print("Демонстрация автоматической сортировки:")
        print("Исходный порядок:", [data[0] for data in demo_data])
        
        for data in demo_data:
            # Находим позицию для вставки
            row = 7
            existing_days = []
            while sheet[f'E{row}'].value is not None:
                existing_days.append(sheet[f'E{row}'].value)
                row += 1
            
            # Простая вставка для демонстрации
            insert_row = 7
            for i, existing_day in enumerate(existing_days):
                if data[0] < existing_day:
                    break
                insert_row = 8 + i
            
            # Сдвигаем и вставляем
            temp_data = []
            current_row = insert_row
            while sheet[f'E{current_row}'].value is not None:
                temp_data.append([
                    sheet[f'E{current_row}'].value,
                    sheet[f'F{current_row}'].value,
                    sheet[f'G{current_row}'].value,
                    sheet[f'H{current_row}'].value
                ])
                current_row += 1
            
            # Вставляем новую запись
            sheet[f'E{insert_row}'] = data[0]
            sheet[f'F{insert_row}'] = data[1]
            sheet[f'G{insert_row}'] = data[2]
            sheet[f'H{insert_row}'] = data[3]
            
            # Восстанавливаем сдвинутые данные
            for i, temp in enumerate(temp_data):
                sheet[f'E{insert_row + 1 + i}'] = temp[0]
                sheet[f'F{insert_row + 1 + i}'] = temp[1]
                sheet[f'G{insert_row + 1 + i}'] = temp[2]
                sheet[f'H{insert_row + 1 + i}'] = temp[3]
            
            print(f"Добавлено число {data[0]} в строку {insert_row}")
        
        # Показываем результат
        print("\nРезультат после сортировки:")
        row = 7
        while sheet[f'E{row}'].value is not None:
            print(f"Строка {row}: {sheet[f'E{row}'].value} - {sheet[f'F{row}'].value}")
            row += 1
        
        wb.save("ПРимер_с_сортировкой.xlsx")
        print("\nФайл сохранен как 'ПРимер_с_сортировкой.xlsx'")
        
    except Exception as e:
        print(f"Ошибка: {e}")

if __name__ == "__main__":
    print("Программа для заполнения журнала преподавателя")
    print("Заполнение начинается с позиции: E7")
    print("\nФормат заполнения:")
    print("  E - число месяца | F - дисциплина | G - группа | H - вид нагрузки")
    print("\n1. Интерактивный режим (с сортировкой)")
    print("2. Демонстрация сортировки")
    
    choice = input("Выберите режим (1/2): ").strip()
    
    if choice == '1':
        main()
    elif choice == '2':
        demo_sorted_fill()
    else:
        print("Запуск интерактивного режима...")
        main()